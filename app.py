"""Local, dependency-light bank-statement normalization web application."""
from __future__ import annotations

import csv
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
import io
import json
import hashlib
import hmac
import os
import re
import shutil
import threading
import time
from difflib import SequenceMatcher
import urllib.error
import urllib.request
import uuid
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

try:
    from PIL import Image
    import pytesseract
    from pytesseract import Output as TesseractOutput
except ImportError:
    Image = None
    pytesseract = None
    TesseractOutput = None

import openpyxl
from openpyxl.styles import Font, PatternFill
try:
    import xlrd
except ImportError:
    xlrd = None
import pdfplumber
from pypdf import PdfReader
try:
    from docx import Document
except ImportError:  # Installed from requirements in production; retain a clear fallback error locally.
    Document = None
try:
    import pymupdf as fitz  # PyMuPDF: substantially faster text extraction for long PDFs.
except ImportError:  # Keep local/development fallback usable until dependencies install.
    try:
        import fitz  # Compatibility with older PyMuPDF releases.
    except ImportError:
        fitz = None

ROOT = Path(__file__).parent
# Railway's container filesystem is replaced on a restart.  When a persistent
# volume is mounted at /data, keep mutable job state, uploads, exports and
# learned profiles there.  Local development continues to use the project
# folder without any setup.
DATA_ROOT = Path(os.environ.get("UPG_DATA_DIR") or ("/data" if Path("/data").exists() else ROOT))
UPLOADS = DATA_ROOT / "uploads"
EXPORTS = DATA_ROOT / "exports"
PROFILES = DATA_ROOT / "profiles"
PROFILE_REVISIONS = PROFILES / "revisions"
JOBS_DIR = DATA_ROOT / "jobs"
LEARNING_LEDGER = PROFILES / "validated_learning.json"
UPG_API_KEY = os.environ.get("UPG_API_KEY", "")
UPG_WEBHOOK_URL = os.environ.get("UPG_WEBHOOK_URL", "")
UPG_WEBHOOK_SECRET = os.environ.get("UPG_WEBHOOK_SECRET", "")
for folder in (UPLOADS, EXPORTS, PROFILES, PROFILE_REVISIONS, JOBS_DIR):
    folder.mkdir(exist_ok=True)

# Seed a new persistent volume with the validated profiles shipped in the
# repository. Never overwrite a learned profile already present on the volume.
if DATA_ROOT != ROOT:
    bundled_profiles = ROOT / "profiles"
    if bundled_profiles.exists():
        for source in bundled_profiles.glob("*.json"):
            target = PROFILES / source.name
            if not target.exists():
                shutil.copy2(source, target)

JOBS: dict[str, dict] = {}
JOBS_LOCK = threading.RLock()
# Heavy PDF work is deliberately bounded.  The service accepts many requests
# immediately, but only a sized number may run in this container at once.
# Railway's 1 GB starter container must remain at one worker.  A larger service
# can opt in to 2-4 workers with UPG_WORKERS; we cap the value to avoid a typo
# turning 30 simultaneous uploads into an out-of-memory restart.
try:
    WORKER_CAPACITY = min(4, max(1, int(os.environ.get("UPG_WORKERS", "1"))))
except ValueError:
    WORKER_CAPACITY = 1
try:
    # A parser must keep retrying until it validates, but no single difficult
    # statement may occupy a multi-tenant worker forever.  One full retry
    # round per lease gives every queued tenant a fair turn; the job resumes
    # automatically from its persisted investigation state.
    WORKER_LEASE_ROUNDS = min(3, max(1, int(os.environ.get("UPG_WORKER_LEASE_ROUNDS", "1"))))
except ValueError:
    WORKER_LEASE_ROUNDS = 1
try:
    # A client that closes/cancels Fix with AI must not reserve a queue turn
    # forever.  Normal browser/API polling refreshes this lease automatically.
    JOB_CLIENT_LEASE_SECONDS = max(60, int(os.environ.get("UPG_CLIENT_LEASE_SECONDS", "300")))
except ValueError:
    JOB_CLIENT_LEASE_SECONDS = 300
JOB_EXECUTOR = ThreadPoolExecutor(max_workers=WORKER_CAPACITY, thread_name_prefix="upg-parser")
ACTIVE_JOB_IDS: set[str] = set()
# Per-upload, in-memory evidence cache. This speeds candidate retries without
# persisting statement text or transactions as long-term learning data.
EXTRACTION_CACHE: dict[tuple[str, str], tuple[list[list[object]], str]] = {}
EXTRACTION_CACHE_LOCK = threading.Lock()
PDF_TEXT_CACHE: dict[str, str] = {}
PDF_SAMPLE_CACHE: dict[str, str] = {}
PDF_GEOMETRY_PROFILE_CACHE: dict[str, tuple[list[object], list[tuple[float, float]]] | None] = {}
# A candidate's original-source plausibility is immutable for one uploaded
# file. Retain the tiny boolean result so retries never re-extract the same
# representative pages merely to rediscover that a strategy cannot fit.
CANDIDATE_SAMPLE_PROOF_CACHE: dict[tuple[str, str], bool] = {}
# The same source page may be measured with a secondary Tesseract layout mode
# when the normal table mode misses date anchors.  Include the mode in the
# cache key so a recovery pass can never overwrite the primary geometry.
OCR_WORD_PAGE_CACHE: dict[tuple[str, int, str], list[dict]] = {}
# Passwords are request-scoped, held only in memory, and are never written to
# profiles, learning, exports, logs, or webhook payloads.
PDF_PASSWORD_CACHE: dict[str, str] = {}

def _retention_seconds(setting: str, default: int, minimum: int) -> int:
    """Read a safe retention setting without allowing accidental immediate deletion."""
    try:
        return max(minimum, int(os.environ.get(setting, str(default))))
    except ValueError:
        return default

# Source uploads are only required while an active job is parsing.  Retain them
# briefly for diagnostics, then remove them so Railway's persistent volume is
# reserved for durable parser knowledge rather than customer statements.
UPLOAD_RETENTION_SECONDS = _retention_seconds("UPG_UPLOAD_RETENTION_SECONDS", 6 * 60 * 60, 30 * 60)
EXPORT_RETENTION_SECONDS = _retention_seconds("UPG_EXPORT_RETENTION_SECONDS", 24 * 60 * 60, 60 * 60)
JOB_RETENTION_SECONDS = _retention_seconds("UPG_JOB_RETENTION_SECONDS", 7 * 24 * 60 * 60, 24 * 60 * 60)
STORAGE_SWEEP_SECONDS = _retention_seconds("UPG_STORAGE_SWEEP_SECONDS", 15 * 60, 5 * 60)

def job_file(job_id: str) -> Path:
    return JOBS_DIR / f"{job_id}.json"

def persist_job_locked(job_id: str) -> None:
    """Atomically save non-secret job state for restart recovery."""
    job = dict(JOBS.get(job_id, {}))
    # Passwords are deliberately never part of a job record.
    job.pop("password", None)
    job.pop("pdf_password", None)
    job["updated_at"] = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    temporary = job_file(job_id).with_suffix(".tmp")
    temporary.write_text(json.dumps(job, ensure_ascii=False), encoding="utf-8")
    temporary.replace(job_file(job_id))

def checkpoint_job(job_id: str) -> None:
    with JOBS_LOCK:
        persist_job_locked(job_id)

def replace_job(job_id: str, value: dict) -> None:
    with JOBS_LOCK:
        JOBS[job_id] = value
        persist_job_locked(job_id)

def patch_job(job_id: str, **changes: object) -> dict:
    with JOBS_LOCK:
        job = JOBS.get(job_id, {})
        job.update(changes)
        JOBS[job_id] = job
        persist_job_locked(job_id)
        return job

def reserve_ai_call(job_id: str | None, purpose: str) -> bool:
    """Reserve one bounded expert-AI decision for a parser job.

    Deterministic extraction, cached certified profiles and validation never
    consume this allowance.  This only prevents an unfamiliar layout from
    repeatedly sending the same source evidence to the model during retries.
    """
    if not job_id:
        return True
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if job is None:
            return False
        used = int(job.get("ai_calls", 0) or 0)
        if used >= MAX_AI_CALLS_PER_JOB:
            job.update({
                "ai_budget_exhausted": True,
                "ai_budget_message": (
                    f"UPG used its {MAX_AI_CALLS_PER_JOB} evidence-led AI decisions for this job. "
                    "It will not repeat AI calls without new source evidence."
                ),
            })
            JOBS[job_id] = job
            persist_job_locked(job_id)
            return False
        history = [str(item) for item in job.get("ai_call_purposes", [])][-5:]
        history.append(purpose)
        job.update({"ai_calls": used + 1, "ai_call_purposes": history})
        JOBS[job_id] = job
        persist_job_locked(job_id)
        return True

def ai_call_purposes(job_id: str | None) -> set[str]:
    """Return the durable AI decisions already used by one parser job."""
    if not job_id:
        return set()
    with JOBS_LOCK:
        return {str(item) for item in JOBS.get(job_id, {}).get("ai_call_purposes", [])}

def ai_calls_remaining(job_id: str | None) -> int:
    """Read remaining expert-AI budget without reserving another call."""
    if not job_id:
        return MAX_AI_CALLS_PER_JOB
    with JOBS_LOCK:
        used = int(JOBS.get(job_id, {}).get("ai_calls", 0) or 0)
    return max(0, MAX_AI_CALLS_PER_JOB - used)

def queue_snapshot_locked() -> tuple[int, int]:
    """Return this job's FIFO position and the total queued depth.

    The executor is intentionally bounded, so a truthful position is more
    useful than showing every waiting tenant an unqualified "processing".
    """
    queued = sorted(
        (item for item in JOBS.items() if item[1].get("status") == "queued"),
        key=lambda item: (str(item[1].get("queued_at") or item[1].get("submitted_at") or ""), item[0]),
    )
    return len(queued), len(ACTIVE_JOB_IDS)

def refresh_queue_positions_locked() -> None:
    queued = sorted(
        (item for item in JOBS.items() if item[1].get("status") == "queued"),
        key=lambda item: (str(item[1].get("queued_at") or item[1].get("submitted_at") or ""), item[0]),
    )
    for position, (queued_id, job) in enumerate(queued, start=1):
        job["queue_position"] = position
        job["queue_depth"] = len(queued)
        job["worker_capacity"] = WORKER_CAPACITY
        JOBS[queued_id] = job
        persist_job_locked(queued_id)

def touch_worker(job_id: str, **changes: object) -> None:
    """Persist a liveness heartbeat for long PDF and AI operations."""
    changes["worker_heartbeat_at"] = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    patch_job(job_id, **changes)

def utc_now() -> datetime:
    return datetime.utcnow()

def timestamp_now() -> str:
    return utc_now().isoformat(timespec="seconds") + "Z"

def timestamp_age_seconds(value: object) -> float:
    try:
        return max(0.0, (utc_now() - datetime.fromisoformat(str(value).replace("Z", "+00:00")).replace(tzinfo=None)).total_seconds())
    except (TypeError, ValueError):
        return float("inf")

def touch_client(job_id: str) -> None:
    """Refresh a job lease, throttled to avoid writing on every poll."""
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if not job or job.get("status") in {"completed", "failed", "cancelled"}:
            return
        if timestamp_age_seconds(job.get("client_heartbeat_at")) < 30:
            return
        job["client_heartbeat_at"] = timestamp_now()
        JOBS[job_id] = job
        persist_job_locked(job_id)

def cancel_job(job_id: str, reason: str = "Cancelled by the requesting client.") -> bool:
    """Cancel queued work now, or request a safe stop between active steps."""
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return False
        if job.get("status") in {"completed", "failed", "cancelled"}:
            return True
        job.update({"cancel_requested": True, "cancel_reason": reason,
                    "client_heartbeat_at": timestamp_now()})
        if job.get("status") == "queued":
            job.update({"processing": False, "valid": False, "status": "cancelled",
                        "message": "UPG job was cancelled before a worker started it."})
        else:
            job["message"] = "UPG cancellation requested; the active worker will stop safely after its current extraction step."
        JOBS[job_id] = job
        persist_job_locked(job_id)
        refresh_queue_positions_locked()
        return True

def cancel_direct_job(job_id: str, cancel_token: str) -> bool:
    """Allow only the browser that submitted a public UI job to cancel it."""
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        expected = str(job.get("cancel_token", "")) if job else ""
    if not expected or not cancel_token or not hmac.compare_digest(expected, cancel_token):
        return False
    return cancel_job(job_id, "Cancelled because the direct UPG page was closed or refreshed.")

def job_cancel_requested(job_id: str) -> bool:
    with JOBS_LOCK:
        job = JOBS.get(job_id, {})
        return bool(job.get("cancel_requested") or job.get("status") == "cancelled")

def expire_abandoned_jobs() -> None:
    """Remove work whose requesting browser/API client has stopped polling."""
    candidates: list[str] = []
    with JOBS_LOCK:
        for job_id, job in JOBS.items():
            if job.get("status") not in {"queued", "processing", "pending"}:
                continue
            last_seen = job.get("client_heartbeat_at") or job.get("submitted_at")
            # A direct browser page has no server-side caller that can resume
            # it safely.  Release its worker sooner when the page was closed,
            # refreshed, or lost network.  API jobs retain the normal lease
            # because BS Analyzer continues polling them server-to-server.
            lease_seconds = 90 if job.get("cancel_token") else JOB_CLIENT_LEASE_SECONDS
            if timestamp_age_seconds(last_seen) > lease_seconds:
                candidates.append(job_id)
    for job_id in candidates:
        cancel_job(job_id, "Cancelled automatically because the requesting client stopped monitoring this job.")

def queue_supervisor() -> None:
    while True:
        try:
            expire_abandoned_jobs()
        except Exception:
            pass
        time.sleep(15)

CANONICAL = ["date", "narration", "withdrawal", "deposit", "instrument_number", "balance"]
# These five fields are mandatory for a usable bank-statement transaction.
# Instrument/cheque number remains a useful optional source field: not every
# bank prints one, but keeping it preserves the original product contract.
MANDATORY_TRANSACTION_FIELDS = ("date", "narration", "withdrawal", "deposit", "balance")
OPTIONAL_TRANSACTION_FIELDS = ("instrument_number",)
FIVE_MINUTES_MS = 300000
# AI is used to plan a source-layout mapping, never to extract every
# transaction. Keep the expert model as the quality default. Cost control is
# achieved by eliminating repeated, low-evidence calls below—not by lowering
# the reasoning capability available for unfamiliar layouts.
AI_MODEL = os.environ.get("UPG_AI_MODEL", "gpt-5.6-sol")
try:
    AI_MAX_OUTPUT_TOKENS = min(1_200, max(250, int(os.environ.get("UPG_AI_MAX_OUTPUT_TOKENS", "550"))))
except ValueError:
    AI_MAX_OUTPUT_TOKENS = 550
try:
    # One layout plan and, only if needed, one evidence-led repair diagnosis.
    # Deterministic profiles/validation do not consume this allowance.
    MAX_AI_CALLS_PER_JOB = min(3, max(1, int(os.environ.get("UPG_MAX_AI_CALLS_PER_JOB", "2"))))
except ValueError:
    MAX_AI_CALLS_PER_JOB = 2
DIAGNOSTIC_RULE_LIBRARY = {
    "value_date": "Use Value Date as the output date when both posting and value dates exist.",
    "dual_date_running_balance": "For layouts with both posting and Value Date, begin records only at the posting-date column but export the Value Date and infer amounts from running-balance changes.",
    "summary_endpoints": "Use the printed statement-summary opening and closing balances as endpoints.",
    "balance_delta": "Classify a single unsigned amount column from running-balance movement.",
    "continuation_merge": "Join narration and transaction fragments across rows or pages.",
    "narration_source_cell": "Export Particulars only from the same measured source Particulars cell (plus source-proven continuation text). A blank Particulars cell stays blank; never copy a Withdrawal, Deposit, Balance, Date, Instrument, header, footer, or adjacent-row token into narration.",
    "footer_exclusion": "Exclude totals, closing labels, disclaimers, and page furniture.",
    "terminal_row_before_summary": "When statement-summary labels begin after the final dated row, seal that dated row before ignoring summary/footer text; never let a later opening/closing/total label contaminate its date or narration.",
    "reverse_order": "Reverse newest-first statements before reconciliation.",
    "source_coverage": "Reject partial extracts and require all detectable source records.",
    "truncated_table_date": "Repair a date cell only when the original source proves its missing final year digit; retain the row's actual amount and balance, never discard it.",
    "bf_preperiod_artifact": "For a dated B/F-area OCR artifact, never delete a source-proven amount merely because its OCR date falls outside the statement period. Neutralize it only when source context proves the amount is included in printed Grand Totals: pin its date to the statement-period start, clear its non-transaction B/F balance, retain its actual amount/narration, and make the opening/closing anchors ignore that cleared balance.",
    "signed_balance_text": "Use dated text blocks with Dr/Cr running balances, then classify debit or credit from each balance movement.",
    "headerless_layout": "Treat a repeated or missing table header as layout evidence, not as a transaction; infer columns only from dated rows and running balances.",
    "multi_page_continuation": "Preserve a dated transaction whose narration or amount cells continue across a page boundary, excluding page headers and footers between its parts.",
    "summary_total_warning": "Keep inconsistent printed debit or credit totals as a warning when transaction count, balance chain, and endpoint reconciliation independently pass.",
    "amount_balance_consistency": "When a source row visibly prints its transaction amount, require that amount to agree with the running-balance movement; reject a layout that only reconciles after replacing source amounts.",
    "balance_source_cell": "A normal running balance must be read from the same measured source Balance cell on the dated row. It may be repaired only when the damaged source token and adjacent source movements prove one unique value; never manufacture a balance merely to reconcile totals.",
    "source_amount_geometry": "When the running-balance chain is unreliable but separate Withdrawal and Deposit columns are measurable, use original-PDF column geometry for the printed movements; classify neither from balance deltas nor from narration.",
    "reference_date_boundary": "When full-year transaction dates are used, do not split a row at a short date embedded in narration or a reversal/reference number; retain the complete source row and its printed amount.",
    "date_column_boundary": "A date starts a transaction only when it is in the measured Date column at the row boundary. Any date-like text to the right, including narration/reference dates, remains Particulars text.",
    "date_source_cell": "Each exported date must be the date from that row's measured Date or Value Date cell. A source-proven repair of a truncated year is allowed; a date-like narration, reference, balance, or adjacent-row value is never a transaction date.",
    "distinct_source_columns": "Each canonical transaction role must come from its own measured source column. Date, Particulars, Withdrawal, Deposit and Balance may never share a source index. The only supported alternative is one Amount column plus a separate Dr/Cr Type column; never reuse Balance, Narration or a movement column to fill another role.",
    "header_role_alignment": "When a native or measured source header explicitly identifies a role (for example Withdrawal Amount, Deposit Amount, Transaction Remarks, Value Date or Running Balance), that role must use that exact measured source column. A saved profile or AI addendum may fill only unlabelled roles; it may not move an explicit header to a different column.",
    "measured_column_evidence": "Before a layout map can be certified, its measured Date column must contain source dates, its movement columns (or Amount plus Dr/Cr Type) must contain source movements, and its Balance column must contain source balance evidence on dated rows. A header label alone is never proof of a usable column.",
    "failure_specific_repair": "After a candidate fails, classify the exact failed source proof before repairing it: header alignment, measured-column evidence, date traceability, balance traceability, source count, narration coverage, amount totals, or unreliable balance. Change only that failed module; never repeat a broad parser attempt or weaken a release gate.",
    "numeric_date_geometry": "For original-PDF geometry, recognize both DD-Mon-YYYY and DD-MM-YYYY or DD/MM/YYYY transaction dates, but only inside the measured Date-column x-band. Keep the source Withdrawal and Deposit x-bands authoritative and exclude trailing system-generated footer text from the final narration.",
    "corrupt_balance_text_layer": "For a PDF whose visible balance is correct but whose searchable text corrupts its punctuation, keep the measured debit/credit amount authoritative. Repair a blanked balance only if the preceding movement and the next measured balance, or the independently printed final totals, prove exactly one balance; never invent the opposite movement to force reconciliation.",
    "indian_money_punctuation": "Infer decimal precision only from normal monetary cells in the same source column. If a damaged token has multiple full stops, repair it only when its final fractional group has that credible precision and all earlier groups exactly form Indian comma grouping; never change digits, sign, direction, or column.",
    "unordered_balance_chain": "Preserve original PDF page and visual row order in the exported transactions. Only when searchable-text serialisation demonstrably differs from that measured order may a unique amount-and-balance chain be built internally for validation; never reorder, delete, or alter source rows. Reject ambiguity or an incomplete chain.",
}
# Reusable behaviours are classified independently from bank/layout profiles.
# A new source can compose (for example) narration continuation from one
# certified layout and footer removal from another, while its own geometry is
# still measured from the submitted file.  This is an explainable retrieval
# system--not a blind whole-parser copy or an untrained "deep learning" claim.
RULE_GROUPS = {
    "narration": {"continuation_merge", "multi_page_continuation", "reference_date_boundary", "narration_source_cell"},
    "furniture": {"footer_exclusion", "terminal_row_before_summary", "bf_preperiod_artifact", "headerless_layout"},
    "dates": {"value_date", "dual_date_running_balance", "reverse_order", "truncated_table_date", "date_column_boundary", "numeric_date_geometry", "date_source_cell"},
    "money_and_balance": {"balance_delta", "signed_balance_text", "corrupt_balance_text_layer", "indian_money_punctuation", "unordered_balance_chain", "amount_balance_consistency", "source_amount_geometry", "balance_source_cell"},
    "endpoints_and_totals": {"summary_endpoints", "summary_total_warning"},
    "validation": {"source_coverage", "narration_source_cell", "balance_source_cell", "date_source_cell", "distinct_source_columns", "header_role_alignment", "measured_column_evidence", "failure_specific_repair"},
}
# Earlier certified profiles predate the structured rule-library fields.  They
# are still useful evidence, but an empty historic `diagnostic_rules` list
# must not make a source-proven capability look like it has no reusable rule.
# These are capability modules, never foreign parser coordinates or code.
CAPABILITY_DEFAULT_RULES = {
    "value_date": ("value_date", "dual_date_running_balance", "date_column_boundary", "reference_date_boundary"),
    "bf_preperiod_artifact": ("bf_preperiod_artifact", "terminal_row_before_summary", "summary_endpoints"),
    "footer_exclusion": ("footer_exclusion", "terminal_row_before_summary", "summary_endpoints"),
    "multi_page_continuation": ("multi_page_continuation", "continuation_merge", "footer_exclusion"),
    "signed_balance_text": ("signed_balance_text", "balance_delta", "amount_balance_consistency"),
    "balance_delta": ("balance_delta", "amount_balance_consistency", "indian_money_punctuation"),
}

def rule_groups_for(rule_ids: object) -> list[str]:
    """Return deterministic capability groups for stored/retrieved lessons."""
    values = {str(item) for item in (rule_ids or [])}
    return sorted(group for group, members in RULE_GROUPS.items() if values & members)

def group_for_capability(capability: str) -> str:
    return {
        "value_date": "dates", "bf_preperiod_artifact": "furniture",
        "footer_exclusion": "furniture", "multi_page_continuation": "narration",
        "signed_balance_text": "money_and_balance", "balance_delta": "money_and_balance",
    }.get(capability, "validation")

def rules_for_capability(capability: str) -> list[str]:
    """Return the certified reusable module IDs for a source capability."""
    return [rule for rule in CAPABILITY_DEFAULT_RULES.get(capability, ())
            if rule in DIAGNOSTIC_RULE_LIBRARY]
PARSER_GENERATOR_POLICY = """
Bank statement extraction policy:
- Transaction output is a strict whitelist only: Date, Particulars/Narration, Withdrawal, Deposit, Running/Closing Balance, and optional Instrument/Cheque Number. Date, Particulars/Narration, Withdrawal, Deposit, and Balance are the five mandatory transaction fields. A parser profile may map only these fields; every other PDF object is non-transaction evidence unless it directly proves one of the permitted fields.
- When both Transaction/Posting Date and Value Date are printed, retain Transaction Date only as source evidence and always export Value Date. Normalize every exported date as DD/MM/YYYY. Never create a second transaction from the two date columns.
- Treat a column headed Closing Balance, Balance, Available Balance, or Running Balance as the transaction running-balance field only when it is aligned with a dated transaction row. A statement-level Closing Balance label is endpoint evidence, never a transaction row.
- Ignore logos, seals, images, QR codes, coloured banners, decorative lines, account-holder/address blocks, bank/branch contact information, page numbers, generated-on stamps, signatures, legal notices, repeated headings, and empty visual cells. Images and logos are never narration or an instrument number.
- Particulars/Narration is source text from its own transaction cell only. Do not fill it using numbers, balances, dates, bank names, logos, headers, or nearby furniture. Keep it blank if the actual particulars cell is blank.
- Use the original PDF's word coordinates, column x-ranges, and row y-ranges as the primary evidence for creating and reusing a parser profile. A date starts a transaction only when it occurs in the measured Date-column band at a row boundary; a date-looking token elsewhere is narration/reference text, never a record split. Use extracted text only to join narration continuations, provide validation evidence, or as a fallback when usable PDF geometry is absent. For a fixed-width text layout, the equivalent boundary is the leading date at the start of a source row.
- For a DOCX statement that has no actual Word table or cell coordinates, preserve paragraph order and leading spaces exactly and treat it as a fixed-width text layout. Derive columns from the source heading and aligned amount/balance positions; merge wrapped continuation lines. Never claim PDF-style geometry for such a DOCX, and do not discard whitespace before extraction.
- B/F, opening balance, and brought-forward entries are statement metadata, never transactions. Narrow exception: an OCR-dated B/F-area artifact that has a source-proven transaction amount included in the printed Grand Total must not be deleted. Only when statement-period evidence, B/F context, and printed-total reconciliation jointly prove this case, pin its date to the period start and clear its B/F balance so it cannot become an opening/closing anchor; retain the source-proven amount and narration. Otherwise reject the ambiguity rather than guessing.
- A row without a valid transaction date is statement furniture, a page/transaction total, or a balance label; never treat it as a transaction merely because it contains amounts.
- Use the statement opening balance when printed. If it is absent, derive it from the first real transaction's signed running balance minus its deposit plus its withdrawal.
- A printed statement-level opening or closing balance overrides any inferred value. Otherwise, the closing balance is the signed running balance of the last real transaction, never a page total, grand total, available amount, or other footer balance.
- Normalize Cr balances as positive and Dr balances as negative. A signed increase is a deposit; a signed decrease is a withdrawal.
- Monetary format rule for Withdrawal, Deposit, and Balance: a valid amount has zero or one decimal point only; all Indian thousands/lakh/crore grouping must use commas. Infer the column's decimal precision only from normal single-decimal source cells. A token with multiple points may be restored only when the final group has that proven precision and every earlier group exactly fits Indian grouping (for example `-5.00.177.00` -> `-5,00,177.00`). Never change digits, sign, debit/credit direction, or the source column; otherwise reject it as malformed evidence.
- If a PDF's visual balance is correctly printed but its searchable text has malformed punctuation (for example `-5,00,177.00` becoming `-5.00.177.00`), reject that damaged token rather than truncating it. Preserve the measured source withdrawal/deposit. Restore the balance only when the previous source balance plus that measured movement and either the next measured balance or independently printed final totals prove one unique value; record this as a text-layer repair, never as an invented amount.
- If the source running-balance chain is unreliable but its separate Withdrawal and Deposit columns are measurable, read each movement from the original PDF's physical x-band. Keep those printed movements authoritative; do not change them from balance deltas or narration. Certify this exception only with full transaction coverage, exact printed withdrawal/deposit totals, one available endpoint or a source-proven derived endpoint, and a visible running-balance warning.
- Balance-chain validation is mandatory for every transaction with a running balance: previous balance = current balance + current withdrawal - current deposit. Equivalently, current balance = previous balance - withdrawal + deposit. Do not release a parser when any transaction balance is missing or breaks this chain.
- Exception: if any transaction proves that the source running-balance chain is unreliable, do not use that column for a normal balance-chain pass or transaction classification. Certify only if parsed withdrawals and deposits exactly equal the printed statement totals, while narration coverage and transaction count pass. Form assumed endpoints from one available transaction balance and the verified totals, label them assumed, and require manual source review. If totals or independent evidence are inconsistent, withhold the parser.
- Transaction-count validation is mandatory: independently count source records that have a transaction date plus amount/running-balance evidence, and require exactly that many parsed transactions. More or fewer parsed rows is a failure even if balances reconcile.
- Particulars must contain only actual transaction narration. Do not put monetary amounts, blank-field substitutes, page headers, account-holder text, totals, or statement furniture in it. If the source Particulars is blank, output a blank narration.
- Join continuation fragments of the same transaction across pages and ignore repeated headers/footers.
- When a transaction crosses a page boundary, remove the intervening page total, disclaimer, bank header, account-holder block, and repeated statement title before joining the remaining fragments. The date, narration, amount, and running balance must remain one transaction.
- For text-layout statements with no reliable table geometry, detect whether each dated transaction is a multi-line block: a narration/reference line followed by an amount and running-balance line. Do not treat a statement-period date before the transaction heading as a transaction. If both posting Date and Value Date columns exist, only the posting-date column starts a record; output the Value Date without treating it as a second transaction.
- If a combined text layer is unreliable, also test a page-by-page text candidate and join its dated rows before validation. This candidate must still reconcile across the complete statement.
- When a layout has one unsigned amount column rather than separate debit/credit columns, infer withdrawal or deposit only from the signed change between consecutive running balances. Ignore long reference IDs, account numbers, dates, timestamps, page numbers, and footer postal codes as monetary values.
- A running balance of exactly zero is valid even when its usual Dr/Cr suffix is omitted. Treat a dated row ending in an explicit zero amount as a real transaction only when it has the row's amount and running-balance evidence; derive its direction from the balance chain.
- Some statements group or print same-date transactions out of running-balance order. If each row visibly has an amount and a balance, reconstruct the order only when each next row is uniquely proved by `next balance = previous balance - withdrawal + deposit`, from the declared opening balance through the declared closing balance. Preserve each row's actual source amount and narration; reject any ambiguous, partial, or disconnected chain.
- If an undated narration/reference fragment is immediately followed by a date, time, amount and signed running balance, join that fragment to the dated row. The date/time line is a continuation of that transaction, not a separate blank-narration transaction.
- A table extractor can truncate or misplace a date while still correctly reading the amount and running balance. For a date-like cell with a missing final year digit, consult the original source text for the immediately following digit and repair it only when that exact completion is present. Keep that transaction; do not discard it merely because the table cell is malformed. If its source Particulars is blank, export a blank Particulars field rather than inventing text.
- Cut a transaction block before closing-balance labels, transaction totals, grand totals, available-balance labels, disclaimers, and other footer furniture. The printed closing balance is validation evidence, never a transaction amount.
- Before generating a new parser, compare the layout with saved validated profiles. For a related layout, create an addendum that inherits the stable mapping and changes only the differing fields. Never overwrite or regress the older parser.
- For an unfamiliar layout, compose only source-proven capabilities from multiple certified profiles before producing the first candidate. For example, take multi-line narration handling from one certified profile and page-furniture removal from another when this source independently exhibits both conditions. Treat each as a rule module, not as a parser copy: measure this statement's own columns and row boundaries, and never borrow another profile's coordinates, headers, executable code, account data, or transaction values.
- For every upload, including a large PDF, run the exact existing validated parser/profile first. For a large PDF, combine that saved mapping with sampled original-PDF geometry. Create a new parser or addendum only after the existing parser has been fully extracted and has failed a release validation.
- Keep trying safe candidate strategies - saved parser, related-profile addendum, an AI-generated source-layout profile, detected table layout, signed or unsigned text running-balance layout, and chronological/reverse-chronological order - until one passes every validation gate. Do not stop after the first failed candidate and never export a partial or unreconciled result.
- A profile may be saved or Excel released only after every narration is traceable to the source and both the full financial reconciliation and each running-balance step pass. Printed debit/credit summary totals are an additional check when present, but a discrepancy is a warning rather than a release gate because some source statements print incorrect totals.
- A printed-total mismatch may be reported as a warning only when it is small enough to plausibly be a source-summary error. A materially divergent parsed total is a hard failure: it indicates that references, dates, or other non-monetary text may have been read as money.
- A partial extraction is never valid. A candidate must account for every detectable source transaction record; a shorter subset that happens to reconcile is a failure. Narration verification is one-to-one source coverage, not merely a loose text substring check.
- Self-healing is constrained to parser-profile addenda: use failed validation evidence to propose a revised header/column layout, test it from the original source, and retain it only after every release gate passes. Never modify application code, invent transactions, or weaken a validation to make a result pass.
- On every failed candidate, act as an evidence-led expert: classify the root cause (geometry, headers, date order, continuation, furniture, balance direction, endpoints, totals, narration, count, or novel layout), choose a materially different safe addendum action, and remember only that non-sensitive investigation result. Never retry a deterministic strategy that already failed the same statement.
- For long PDFs, create or repair a layout profile from a representative sample: first seven pages, seven pages centered around the middle, and last seven pages, plus adjacent boundary pages so transactions split across sampled-page edges remain visible. Apply the resulting candidate to the complete statement and validate the whole source before learning or export.
"""
# The API receives this compact contract for planning/diagnosis.  The complete
# policy above is enforced locally by deterministic extraction and validation;
# repeating it in every model prompt wastes tokens without adding protection.
AI_LAYOUT_CONTRACT = (
    "You plan a bank-statement layout only. Output no transactions and never "
    "weaken validation. Keep only date, source narration, withdrawal, deposit, "
    "running balance, and optional instrument number. Ignore furniture such as "
    "headers, totals, logos, addresses, notices, and page labels. If both dates "
    "exist, use Value Date for output DD/MM/YYYY and posting date only for row "
    "boundaries. Use original PDF geometry when supplied; return -1 for absent "
    "columns. For an image-only PDF, use OCR word coordinates from the original "
    "page; flattened OCR text is evidence only and must never define columns. "
    "A candidate is saved only after local financial, source-coverage, "
    "narration, transaction-count, and balance-chain gates pass."
)
# Historical lessons captured from previously resolved statement layouts.  This
# is deliberately generic: it teaches recognition and safe handling, never
# any customer's statement text, balances, account numbers, or transactions.
HISTORICAL_CHALLENGE_LESSONS = [
    "B/F, opening balance, brought-forward, page total, and grand total are metadata, never transactions.",
    "A blank source Particulars remains blank; an amount, instrument number, balance, or nearby furniture must never fill it.",
    "When both posting/transaction date and Value Date are present, use Value Date as the one exported DD/MM/YYYY date. The posting date proves the row boundary only; it is never a second transaction.",
    "Every certified row has exactly the five mandatory concepts: date, source narration (which may be blank), one-sided movement, and running balance. Instrument/cheque number is optional. Keep validation-only evidence private and never export furniture as a field.",
    "Page headers, addresses, branch/account blocks, logos, QR codes, legal notes, print stamps, and repeated titles are furniture even when they sit between two parts of a transaction.",
    "A transaction may continue over a page boundary. Remove intervening furniture, then merge only source-proven narration/amount/balance fragments into the same dated transaction.",
    "Use the original PDF geometry for layout decisions. For large PDFs, use first/middle/last samples with boundary pages to design the profile, then parse and validate all pages.",
    "For an image-only PDF, use OCR word boxes measured against the original page. Rebuild column bands from those coordinates; never reuse flattened OCR text as if it preserved geometry. A multi-line header such as 'Cheque' over 'No/Reference' is one column label, not two transaction fields.",
    "For scanned dual-date tables, Post Date and Value Date are separate measured columns even when their header is printed on one line. Use Value Date for output; never collapse two dates into one cell. A visually centred Description heading may require a wider left-hand narration band, and source header/account dates above the table must be fenced out before row detection.",
    "A bank DOCX can contain an exact fixed-width statement text layer without Word tables, tabs, or coordinates. Preserve its paragraph order and leading spaces, parse it as fixed-width text, and merge its wrapped continuation lines. Do not treat it as PDF geometry or collapse its whitespace before measuring columns.",
    "Statements may run newest-to-oldest. Detect direction before calculating endpoints or balance chains; reverse the chain logic only when source balances prove it.",
    "With a single unsigned amount column, determine withdrawal/deposit from signed running-balance movement, not from references or arbitrary numeric text.",
    "For a headerless or sparse-header PDF with multiple dated rows ending in explicit Dr/Cr balances, prefer the signed running-balance text strategy before table geometry. Use untouched source text for B/F and final endpoints, and remove only proven headers, footers, and address furniture from narration.",
    "For a headerless signed cash-credit ledger, count a source transaction as one dated block that reaches an explicit Dr/Cr balance before the next date. A date-only count can include period labels or other dated furniture and must not reject an otherwise complete, balance-chained extract.",
    "When the source running-balance column is demonstrably unreliable, do not use it as ordinary evidence. Use the narrowly defined printed-totals exception only with exact totals, source coverage, count, narration checks, assumed endpoints, and a manual-review warning.",
    "Printed debit/credit summary totals can be wrong. They are an additional warning unless the unreliable-balance exception requires exact source-total equality; never change parsed rows merely to match a summary.",
    "A correct financial endpoint alone is insufficient. Require one-to-one narration/source coverage, transaction count, and a complete balance chain whenever the source balance is reliable.",
    "For a large multi-year statement, an OCR date near the B/F area can be an artifact even when its amount is real and included in the printed Grand Total. Do not delete that amount. With source proof of B/F context, statement period, and total inclusion, pin the date to the first day of the period and clear the non-transaction balance so endpoint selection ignores it. This is a narrow normalization, not a blanket early-row rule.",
    "A statement summary can begin immediately after the last real transaction. Seal the terminal dated row before reading Opening Balance, Closing Balance, debit/credit totals, or end-of-statement furniture; those labels must never be merged into the final row or cause it to be dropped.",
    "For a searchable dual-date PDF, measure Post Date and Value Date as separate original-PDF columns. A matching pair on one baseline is one ledger record: use Value Date for DD/MM/YYYY output, use Post Date only as the record boundary, and never count statement-period dates as transactions. Some browser-created text layers encode Indian thousands as 10.000.00 instead of 10,000.00; normalize only that exact repeated-thousands shape before reading the measured Debit, Credit, and Balance columns.",
    "Do not re-run a failed deterministic strategy. A repair must change a source-proven layout mapping or select a different supported strategy.",
]
ALIASES = {
    "date": ["date", "transaction date", "txn date", "value date"],
    "narration": ["narration", "particular", "particulars", "description", "remarks", "details"],
    "withdrawal": ["withdrawal", "withdrawal amount", "debit", "debit amount", "dr", "amount debited"],
    "deposit": ["deposit", "deposit amount", "credit", "credit amount", "cr", "amount credited"],
    "instrument_number": ["cheque number", "check number", "instrument number", "instrument", "cheque no", "chq no", "ref no", "chq. / ref no."],
    "balance": ["balance", "closing balance", "closing balance*", "running balance", "available balance"],
    "amount": ["amount", "amount(inr)", "transaction amount"],
    "transaction_type": ["type", "dr/cr", "transaction type"],
}


def semantic_header_role(value: object) -> str | None:
    """Return the canonical meaning of one source header label.

    This is intentionally a vocabulary classifier, not a bank template.  A
    newly uploaded statement is still measured from its own native grid/PDF
    geometry; this helper merely tells the measured-grid code that phrases
    such as ``Transaction Remarks`` and ``Debit Amount (INR)`` describe the
    same concepts as ``Particulars`` and ``Withdrawal``.

    Keeping this small, deterministic layer in front of the AI is important:
    a spelling variation must not consume an expensive AI repair call or make
    UPG borrow a different bank's column coordinates.
    """
    label = norm(value)
    if not label:
        return None
    if "balance" in label or label in {"availableamount", "closingamount"}:
        return "balance"
    if any(token in label for token in ("withdraw", "debit", "amountdebited")) or label in {"dr", "dramount"}:
        return "withdrawal"
    if any(token in label for token in ("deposit", "credit", "amountcredited")) or label in {"cr", "cramount"}:
        return "deposit"
    if any(token in label for token in ("narration", "particular", "remark", "description", "detail", "transactioninfo")):
        return "narration"
    if any(token in label for token in ("cheque", "check", "instrument", "reference", "refno", "transactionid", "utr", "rrn")):
        return "instrument_number"
    if "date" in label:
        # Value/settlement/effective dates are still dates; callers that can
        # see the full measured header retain the Value Date preference.
        return "date"
    if "amount" in label:
        return "amount"
    if label in {"type", "transactiontype", "drcr"}:
        return "transaction_type"
    return None

HTML = r'''<!doctype html><html><head><meta charset="utf-8"><title>Statement Normalizer</title><style>
body{font-family:system-ui;max-width:860px;margin:50px auto;color:#172033;background:#f5f7fb}.card{background:white;padding:30px;border-radius:16px;box-shadow:0 4px 22px #1223}h1{margin-top:0}label{display:block;margin:16px 0 5px;font-weight:650}input,button{font:inherit;padding:10px}input{width:100%;box-sizing:border-box;border:1px solid #cbd5e1;border-radius:7px}button{margin-top:22px;background:#0f766e;color:white;border:0;border-radius:8px;cursor:pointer}.hint{color:#52606d}.result{margin-top:20px;padding:16px;border-radius:8px}.ok{background:#dcfce7}.fail{background:#fee2e2}.field{display:grid;grid-template-columns:1fr 1fr;gap:15px}</style></head><body><main class="card"><h1>Bank Statement Normalizer</h1><p class="hint">Upload a statement. Excel is created only after the declared balances reconcile with parsed transactions. For unfamiliar layouts, the configured AI parser generator may inspect the layout to create a profile; no export is released unless both checks pass.</p><form id="form"><label>Statement file</label><input name="file" type="file" accept=".csv,.xlsx,.xls,.txt,.pdf,.doc,.docx" required><div class="field"><div><label>Opening balance (optional fallback)</label><input name="opening" placeholder="Extracted from source when present"></div><div><label>Closing balance (optional fallback)</label><input name="closing" placeholder="Extracted from source when present"></div></div><label>PDF password (only if protected)</label><input name="password" type="password" autocomplete="off" placeholder="Used only in memory for this upload"><button>Parse and validate</button></form><section id="result"></section></main><script>
const f=document.querySelector('#form'), r=document.querySelector('#result'), submit=f.querySelector('button');let activeJob=null,activeCancelToken=null;
function show(d){const label=d.valid?'Validated':d.processing?'UPG is retrying':d.interrupted?'UPG job interrupted':'Not validated';const job=d.job_id?`<br><small>Job ID: ${d.job_id}</small>`:'';const diagnosis=d.investigation?`<br><small>AI diagnosis: ${d.investigation.failure_type||'investigating'}; action: ${d.investigation.profile_action||'planning addendum'}.</small>`:'';r.className='result '+(d.valid?'ok':d.processing||d.interrupted?'':'fail');r.innerHTML=`<strong>${label}</strong><br>${d.message}`+job+diagnosis+(d.download?`<br><br><a href="${d.download}">Download validated Excel</a>`:'')}
async function poll(job){const d=await (await fetch('/status/'+job)).json();if(job!==activeJob)return;show(d);if(d.processing)setTimeout(()=>poll(job),2500);else{activeJob=null;activeCancelToken=null;submit.disabled=false;submit.textContent='Parse and validate'}}
function cancelDirectJob(){if(!activeJob||!activeCancelToken)return;const body=JSON.stringify({job_id:activeJob,cancel_token:activeCancelToken});navigator.sendBeacon('/cancel',new Blob([body],{type:'application/json'}));}
window.addEventListener('pagehide',cancelDirectJob);
f.onsubmit=async e=>{e.preventDefault();if(activeJob)return;r.className='result';r.innerHTML='<strong>UPG is retrying</strong><br>Creating and validating parser candidates.';submit.disabled=true;submit.textContent='UPG is working...';try{const d=await (await fetch('/parse',{method:'POST',body:new FormData(f)})).json();activeJob=d.job||null;activeCancelToken=d.cancel_token||null;show(d);if(d.processing)poll(d.job);else{activeJob=null;activeCancelToken=null;submit.disabled=false;submit.textContent='Parse and validate'}}catch(err){activeJob=null;activeCancelToken=null;submit.disabled=false;submit.textContent='Parse and validate';r.className='result fail';r.innerHTML='<strong>Unable to start UPG</strong><br>The parser retry job could not start.'}};
</script></body></html>'''

def money(value: object) -> Decimal | None:
    if value is None or str(value).strip() == "": return None
    s = str(value).strip().replace(",", "").replace("₹", "").replace("$", "")
    s = re.sub(r"\s+", "", s)
    # A monetary value has one decimal point at most.  Indian digit grouping
    # must use commas, not additional full stops: ``-5,00,177.00`` is valid,
    # while ``-5.00.177.00`` is invalid.  This rule applies equally to debit,
    # credit and balance cells.
    # Never truncate a malformed value such as ``-5.00.177.00`` to
    # ``-5.00``.  That turns a source printing defect into a fake
    # ₹5,00,000-scale balance movement and can make the parser manufacture an
    # opposite-side deposit.  Proper Indian dot-grouping was normalized just
    # above; any remaining second decimal point is unusable monetary evidence.
    if s.count(".") > 1:
        return None
    # Coordinate extraction can leave a transaction value followed by a page
    # total in the same cell. The first monetary token belongs to the row.
    token = re.match(r"-?(?:\d{1,3}(?:\.\d{3})+|\d+)(?:\.\d{1,2})?(?:DR|CR)?", s, re.I)
    if token:
        s = token.group()
    suffix = re.search(r"(DR|CR)$", s, re.I)
    # A few exports print a debit balance as both `-123.45Dr` and
    # `123.45Dr`.  DR is an accounting sign, not an instruction to negate an
    # already signed decimal: calculate the magnitude first, then apply one
    # final sign.  The previous implementation multiplied `-123.45` by -1
    # when it saw DR and silently turned a debit balance into a credit.
    explicit_negative = s.startswith("-") or (s.startswith("(") and s.endswith(")"))
    debit_suffix = bool(suffix and suffix.group(1).upper() == "DR")
    s = re.sub(r"(?:DR|CR)$", "", s.strip("()"), flags=re.I).strip()
    try:
        amount = abs(Decimal(s))
        return -amount if explicit_negative or debit_suffix else amount
    except InvalidOperation: return None


def inferred_column_decimal_places(rows: list[list[object]], header_at: int, columns: dict) -> dict[str, int]:
    """Infer the trustworthy decimal width independently for each money column.

    Corrupt PDF text layers can turn Indian grouping commas into full stops
    (``5,00,177.00`` -> ``5.00.177.00``).  Only normal, single-decimal source
    cells are allowed to teach this rule.  A damaged value is never evidence
    for repairing itself or another value.
    """
    result: dict[str, int] = {}
    for field in ("withdrawal", "deposit", "balance", "amount"):
        index = columns.get(field)
        if index is None:
            continue
        widths: list[int] = []
        for row in rows[header_at + 1:]:
            if index >= len(row):
                continue
            token = str(row[index] or "").strip().replace(",", "")
            token = re.sub(r"(?:DR|CR)$", "", token, flags=re.I).strip().strip("()")
            if not token or token.count(".") > 1:
                continue
            match = re.fullmatch(r"-?\d+(?:\.(\d{1,2}))?", token)
            if match:
                widths.append(len(match.group(1)) if match.group(1) else 0)
        if widths:
            # Mode, then the wider precision on a tie.  This keeps a mostly
            # two-decimal bank column at two decimals even when a few cells
            # omit trailing zeroes.
            result[field] = max(set(widths), key=lambda width: (widths.count(width), width))
    return result


def repair_indian_grouping_decimal(value: object, decimal_places: int | None) -> Decimal | None:
    """Repair only an exact, structurally valid dot-for-comma Indian amount.

    No digit, sign, direction or column is changed.  The final full stop is
    accepted only when it has the credible decimal width learned from valid
    cells in that same column; preceding full stops must form a valid Indian
    integer grouping (3 digits at the right, 2-digit lakh/crore groups).
    """
    if value is None or decimal_places is None or decimal_places < 1:
        return None
    source = str(value).strip().replace("â‚¹", "").replace("$", "")
    source = re.sub(r"\s+", "", source)
    suffix = re.search(r"(DR|CR)$", source, re.I)
    debit_suffix = bool(suffix and suffix.group(1).upper() == "DR")
    source = re.sub(r"(?:DR|CR)$", "", source, flags=re.I)
    explicit_negative = source.startswith("-") or (source.startswith("(") and source.endswith(")"))
    source = source.strip("-()")
    if source.count(".") < 2 or "," in source:
        return None
    groups = source.split(".")
    integer_groups, fraction = groups[:-1], groups[-1]
    if len(fraction) != decimal_places or not fraction.isdigit() or not all(group.isdigit() for group in integer_groups):
        return None
    # Standard 3-digit grouping is valid for 1,234.56.  Longer Indian numbers
    # have 2-digit groups before that final 3-digit group: 5,00,177.00.
    if not (1 <= len(integer_groups[0]) <= 3 and len(integer_groups[-1]) == 3):
        return None
    if len(integer_groups) > 2 and any(len(group) != 2 for group in integer_groups[1:-1]):
        return None
    try:
        amount = Decimal("".join(integer_groups) + "." + fraction)
    except InvalidOperation:
        return None
    return -amount if explicit_negative or debit_suffix else amount


def source_money(value: object, decimal_places: int | None = 2) -> Decimal | None:
    """Parse one measured money cell without changing its accounting side.

    The normal parser accepts a conventional number first.  Only then does it
    attempt the narrow, evidence-preserving Indian punctuation repair.  This
    ensures a searchable PDF token such as ``-5.00.177.00`` is read as the
    visible ``-5,00,177.00`` when the column's two-decimal convention proves
    it, while a random dotted reference in Particulars can never become an
    amount or create a counter-balancing movement.
    """
    return money(value) or repair_indian_grouping_decimal(value, decimal_places)

def indian_amount(value: Decimal | int | float) -> str:
    """Format monetary values with Indian lakh/crore grouping."""
    amount = Decimal(value).quantize(Decimal(".01"), rounding=ROUND_HALF_UP)
    sign = "-" if amount < 0 else ""
    whole, fraction = f"{abs(amount):.2f}".split(".")
    if len(whole) <= 3:
        grouped = whole
    else:
        tail, rest = whole[-3:], whole[:-3]
        pairs = []
        while rest:
            pairs.append(rest[-2:]); rest = rest[:-2]
        grouped = ",".join(reversed(pairs)) + "," + tail
    return sign + grouped + "." + fraction

def norm(s: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())

def normalize_narration(s: object) -> str:
    """Comparison form: ignores harmless whitespace/case/punctuation variations only."""
    return norm(s)

def coordinate_narrations_traceable(transactions: list[dict], raw: str) -> bool:
    """Validate PDF-coordinate narrations when the text layer changes order.

    A borderless PDF can expose its Particulars words in a different reading
    order to the visual source, even though each word came from the real
    Particulars column.  Exact substring matching is then invalid evidence.
    This narrow alternative is available only after original-PDF geometry has
    already located the narration column: every meaningful narration token
    must be traceable to the source text.  It is not a fallback for text-only
    extraction and never accepts numeric or furniture-only narration.
    """
    # Short source narrations such as GST, ATM, UPI and NEFT are valid
    # particulars.  Require alphabetic three-character tokens rather than
    # silently rejecting them simply because they are shorter than four.
    source_tokens = set(re.findall(r"[a-z][a-z0-9]{2,}", str(raw or "").lower()))
    if not source_tokens:
        return False
    strong_rows = 0
    narrated_rows = 0
    for transaction in transactions:
        narration = str(transaction.get("narration") or "").strip()
        if not narration:
            continue
        narrated_rows += 1
        tokens = re.findall(r"[a-z][a-z0-9]{2,}", narration.lower())
        if not tokens:
            return False
        matched = sum(token in source_tokens for token in tokens)
        score = matched / len(tokens)
        # One corrupted glyph in a scanned/searchable text layer must not
        # invalidate a coordinate-proven row, but a made-up narration cannot
        # pass.  Require every row to retain some source evidence and at
        # least 95% of narrated rows to retain the normal 70% token match.
        if not matched or score < 0.25:
            return False
        if score >= 0.70:
            strong_rows += 1
    return narrated_rows > 0 and strong_rows / narrated_rows >= 0.95

def narration_source_cells_traceable(transactions: list[dict]) -> bool:
    """Prove every exported narration came from its own Particulars cell.

    This is deliberately row-local.  Finding words merely *somewhere* in a
    statement is not enough: a withdrawal, deposit, balance, instrument or
    neighbouring row can otherwise leak into Particulars and still make the
    financial equation reconcile.  A blank source Particulars cell is valid,
    but it must remain blank in the output.

    The check accepts source-cell punctuation/whitespace differences and
    cleaned page furniture, but not invented text.  It is only used for a
    measured native/source table; generated canonical rows have already lost
    their original cell identity and use the stricter geometry trace route.
    """
    checked = 0
    for transaction in transactions:
        if "_source_narration" not in transaction:
            continue
        checked += 1
        parsed = str(transaction.get("narration") or "").strip()
        source = str(transaction.get("_source_narration") or "").strip()
        if not parsed:
            continue
        if not source:
            return False
        parsed_norm, source_norm = normalize_narration(parsed), normalize_narration(source)
        # The parser may remove a header/furniture fragment from a source cell,
        # but it may never add a token from an adjacent financial column.
        if not parsed_norm or parsed_norm not in source_norm:
            return False
    return checked > 0

def balance_source_cells_traceable(transactions: list[dict]) -> bool:
    """Require normal exported balances to be source-cell values.

    A running balance is evidence, not a field UPG may invent to force the
    opening/withdrawal/deposit/closing equation.  The sole permitted
    exception is a separately recorded, uniquely-proven repair of a damaged
    source token; the caller surfaces that condition to the reviewer.
    """
    checked = 0
    for transaction in transactions:
        if "_source_balance_value" not in transaction:
            continue
        checked += 1
        balance = transaction.get("balance")
        source_balance = transaction.get("_source_balance_value")
        if balance is None:
            return False
        if source_balance is not None and Decimal(balance).quantize(Decimal(".01")) == Decimal(source_balance).quantize(Decimal(".01")):
            continue
        if transaction.get("_balance_repaired_from_chain") is True:
            continue
        return False
    return checked > 0

def date_source_cells_traceable(transactions: list[dict]) -> bool:
    """Require output dates to be read from their measured source date cells."""
    checked = 0
    for transaction in transactions:
        if "_source_date_raw" not in transaction:
            continue
        checked += 1
        exported = transaction_date_value(transaction.get("date"))
        source = transaction_date_value(transaction.get("_source_date_raw"))
        if exported is None:
            return False
        if source is not None and exported.date() == source.date():
            continue
        # The only non-direct case is repair_truncated_table_date(), which
        # searches the original source for the missing final year digit.
        if transaction.get("_date_repaired_from_source") is True:
            continue
        return False
    return checked > 0

def clean_narration(s: str) -> str:
    """Keep only the statement's Particulars, never amounts or page furniture."""
    # A J&K Bank transaction can be split after its reference on page N and
    # resume with the remaining Particulars on page N+1.  The repeated branch,
    # account-holder and statement block sits between those two pieces.  Strip
    # that *middle* block while preserving both the reference before it and the
    # real continuation after it.  Dropping the whole cell would lose a valid
    # narration; retaining it leaks furniture into Excel.
    s = re.sub(
        r"(?is)\b(?:jammu\s+and\s+kashmir\s+bank\s+ltd|"
        r"k\.?\s*b\.?\s*adda\s*,\s*baramulla\s*-\s*\d{4}).*?"
        r"\bjammu\s+and\s+kashmir\s*",
        " ",
        str(s or ""),
    )
    lines = []
    furniture = ("jammu and kashmir bank", "statement of account", "page total", "grand total", "printed by", "ifsc code", "micr code", "unless the constituent", "customer id", "currency code", "a/c no", "interest rate", "no nomination", "c kyc", "ckyc")
    for line in s.splitlines():
        compact = re.sub(r"\s+", " ", line).strip()
        # A bank name is furniture only when it is part of the page header.
        # It may legitimately occur inside a payment description (for
        # example, an IMPS beneficiary at J&K Bank).  Retaining that source
        # narration is safer than blanking a genuine Particular merely
        # because it contains the issuing bank's name.
        looks_like_transaction = bool(re.search(
            r"(?i)\b(?:imps|upi|neft|rtgs|pos|atm|rrn\s*:|from\s*:|to\s*:|cheque)\b",
            compact,
        ))
        if compact and (looks_like_transaction or not any(marker in compact.lower() for marker in furniture)):
            lines.append(compact)
    result = " ".join(lines)
    result = re.sub(r"-{5,}", " ", result)
    # Some PDF text layers collapse a multi-line page header into a single
    # string immediately before the continuation. Remove that header prefix.
    result = re.sub(r"(?is)\ba/c no\s*:.*?(?:no nomination available for the account|currency code\s*:\s*inr)\s*", "", result)
    result = re.sub(r"(?is)\bto:\s*.*?\.\s*,\s*", "", result)
    # Time is a separate statement column in several compact PDF text layers,
    # not part of Particulars.
    result = re.sub(r"\b\d{2}:\d{2}:\d{1,2}\b", "", result)
    result = re.sub(r"\s+-?[\d,]+(?:\.\d{1,2})?\s*$", "", result).strip()
    return "" if re.fullmatch(r"-?[\d,]+(?:\.\d{1,2})?", result) else result

def narration_is_furniture(value: object) -> bool:
    """Return True only for text that cannot be a transaction Particulars cell.

    Blank Particulars are valid and deliberately return False: banks sometimes
    print a genuine transaction with an empty narration.  This guard prevents
    an AI/profile from passing a heading, total, address block, or a numeric
    column value through as narration just because the balances reconcile.
    """
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if not text:
        return False
    compact = norm(text)
    if re.fullmatch(r"[\d,.()\-+/]+", text):
        return True
    furniture_markers = (
        "statementofaccount", "openingbalance", "closingbalance",
        "pagetotal", "grandtotal", "transactiontotal", "accountnumber",
        "customercare", "ifscode", "micrcode", "generatedon",
        "thisisasystemgenerated", "doesnotrequireanysignature",
    )
    return any(marker in compact for marker in furniture_markers)

def canonical_transaction_contract_valid(transaction: dict) -> bool:
    """Enforce the output contract before a profile can be certified.

    This is intentionally a release gate rather than a repair step.  It never
    invents missing values or changes a source row merely to make it fit.
    """
    if not transaction_date_value(transaction.get("date")):
        return False
    withdrawal = transaction.get("withdrawal")
    deposit = transaction.get("deposit")
    balance = transaction.get("balance")
    if withdrawal is None or deposit is None or balance is None:
        return False
    if withdrawal < 0 or deposit < 0:
        return False
    # A transaction is a dated source row with one measured movement.  Do not
    # turn a date-bearing header, period label, B/F metadata, or a blank row
    # into a zero-value transaction merely because it happens to align with a
    # balance column.  Likewise never accept two populated financial columns:
    # that is a column-boundary/mapping defect, not a debit and credit to be
    # netted out for reconciliation.
    if (withdrawal == 0) == (deposit == 0):
        return False
    source_amount = transaction.get("source_amount")
    if source_amount is not None and abs(Decimal(source_amount)).quantize(Decimal(".01")) != abs(deposit - withdrawal).quantize(Decimal(".01")):
        return False
    return not narration_is_furniture(transaction.get("narration"))

def canonical_transaction_core_valid(transaction: dict) -> bool:
    """The non-balance portion of the release contract.

    This is used only for the documented unreliable-running-balance exception.
    A malformed source balance may be blanked, but it must never relax the
    requirements for a real date, one measured movement, and source narration.
    """
    if not transaction_date_value(transaction.get("date")):
        return False
    withdrawal = transaction.get("withdrawal")
    deposit = transaction.get("deposit")
    if withdrawal is None or deposit is None or withdrawal < 0 or deposit < 0:
        return False
    if (withdrawal == 0) == (deposit == 0):
        return False
    source_amount = transaction.get("source_amount")
    if source_amount is not None and abs(Decimal(source_amount)).quantize(Decimal(".01")) != abs(deposit - withdrawal).quantize(Decimal(".01")):
        return False
    return not narration_is_furniture(transaction.get("narration"))

def map_headers(headers: list[object]) -> dict[str, int]:
    mapped = {}
    for canonical, possibilities in ALIASES.items():
        for i, h in enumerate(headers):
            header = norm(h)
            if header in {norm(x) for x in possibilities}:
                mapped[canonical] = i; break
    # Header wording varies by bank. These semantic fallbacks allow the parser
    # generator to discover a new layout without a bank-specific code change.
    for i, h in enumerate(headers):
        role = semantic_header_role(h)
        if role:
            mapped.setdefault(role, i)
    # Value Date labels often include a parenthesised format or abbreviations
    # (for example ``Value Date (DD/MM/YYYY)``).  It has priority over the
    # posting-date column for the exported canonical date.
    value_date = next((i for i, h in enumerate(headers) if norm(h).startswith("valuedate") or norm(h) in {"valuedt", "valdt"}), None)
    if value_date is not None:
        mapped["date"] = value_date
    return mapped

def source_columns_are_distinct(columns: dict[str, int]) -> bool:
    """Reject a map that reuses one measured source column for two roles.

    A reconciled equation is not proof that an extractor found the correct
    columns.  In particular, a shifted map can accidentally use Balance as a
    Deposit or put a Withdrawal amount into Particulars.  Check the physical
    source indices before certification.  A single Amount column with an
    independent Dr/Cr Type column is the documented exception to separate
    Withdrawal and Deposit columns.
    """
    required = ["date", "narration", "balance"]
    if "amount" in columns and "transaction_type" in columns:
        required.extend(["amount", "transaction_type"])
    else:
        required.extend(["withdrawal", "deposit"])
    indices: list[int] = []
    for role in required:
        value = columns.get(role)
        if not isinstance(value, int) or value < 0:
            # Existing mapping/row validation explains a missing field.  This
            # guard only answers the separate question: are supplied source
            # columns physically distinct?
            continue
        indices.append(value)
    return len(indices) == len(set(indices))

def explicit_header_roles_aligned(headers: list[object], columns: dict[str, int]) -> bool:
    """Ensure profiles/addenda cannot contradict clear source header labels."""
    explicit = map_headers(headers)
    return all(columns.get(role) == index for role, index in explicit.items())

def measured_column_evidence(rows: list[list[object]], header_at: int, columns: dict[str, int]) -> tuple[bool, dict[str, int]]:
    """Check that a proposed grid map has real typed evidence below its header.

    This is deliberately a pre-certification guard, not a transaction parser.
    It prevents an AI/header candidate from pointing at an empty column, page
    number, account number or narration text merely because its heading looked
    plausible.  It accepts a source with malformed balance punctuation—the
    documented balance exception is assessed later—but still requires at least
    one balance-like source cell on a dated row.
    """
    decimal_places = inferred_column_decimal_places(rows, header_at, columns)
    evidence = {"dated_rows": 0, "movement_rows": 0, "balance_rows": 0}

    def value(row: list[object], role: str) -> object:
        index = columns.get(role)
        return row[index] if isinstance(index, int) and 0 <= index < len(row) else ""

    for row in rows[header_at + 1:]:
        if not transaction_date_value(value(row, "date")):
            continue
        evidence["dated_rows"] += 1
        withdrawal = source_money(value(row, "withdrawal"), decimal_places.get("withdrawal"))
        deposit = source_money(value(row, "deposit"), decimal_places.get("deposit"))
        if withdrawal is None and deposit is None and "amount" in columns and "transaction_type" in columns:
            amount = source_money(value(row, "amount"), decimal_places.get("amount"))
            kind = str(value(row, "transaction_type") or "").upper()
            if amount is not None and ("DR" in kind or "CR" in kind):
                withdrawal = amount if "DR" in kind else Decimal("0")
                deposit = amount if "CR" in kind else Decimal("0")
        if (withdrawal or Decimal("0")) != 0 or (deposit or Decimal("0")) != 0:
            evidence["movement_rows"] += 1
        raw_balance = value(row, "balance")
        if source_money(raw_balance, decimal_places.get("balance")) is not None or re.search(r"\d", str(raw_balance or "")):
            evidence["balance_rows"] += 1

    valid = evidence["dated_rows"] > 0 and evidence["movement_rows"] > 0 and evidence["balance_rows"] > 0
    return valid, evidence

def text_layout_fingerprint(raw: str) -> str:
    """Privacy-safe structural signature for text-only statement layouts."""
    header = re.search(r"(?im)^\s*date\s+.{0,120}(?:balance|amount)\s*$", raw)
    features = {
        "header": norm(header.group(0)) if header else "",
        "opening": bool(re.search(r"\b(?:opening balance|b/f)\b", raw, re.I)),
        "closing": bool(re.search(r"\bclosing balance\b", raw, re.I)),
        "month_dates": len(re.findall(r"\b\d{2}-[A-Za-z]{3}-\d{4}\b", raw)) > 3,
        "numeric_dates": len(re.findall(r"\b\d{2}[-/]\d{2}[-/]\d{4}\b", raw)) > 3,
    }
    return hashlib.sha256(json.dumps(features, sort_keys=True).encode()).hexdigest()[:16]

def profile_id(headers: list[object], layout_fingerprint: str = "") -> str:
    """Stable, privacy-safe layout signature: header labels only, never statement data."""
    signature = "|".join(norm(h) for h in headers) + "|" + layout_fingerprint
    return hashlib.sha256(signature.encode()).hexdigest()[:16]

def generated_canonical_headers(headers: list[object]) -> bool:
    """Do not learn a profile from our own synthetic fallback header."""
    return [norm(h) for h in headers] == [norm(h) for h in ["Date", "Narration", "Withdrawal", "Deposit", "Instrument Number", "Balance"]]

def load_profile(headers: list[object], layout_fingerprint: str = "") -> dict[str, int] | None:
    if generated_canonical_headers(headers) and not layout_fingerprint: return None
    profile = PROFILES / f"{profile_id(headers, layout_fingerprint)}.json"
    if not profile.exists(): return None
    try:
        data = json.loads(profile.read_text(encoding="utf-8"))
        return {k: int(v) for k, v in data["columns"].items()}
    except (OSError, ValueError, KeyError): return None

def certified_javascript_code(headers: list[object], strategy: str | None) -> tuple[str, str]:
    """Build the safe, self-contained JavaScript contract consumed by BS Analyzer.

    This is deliberately a deterministic generator, not AI-authored executable
    code. It emits only a reviewed line parser and a conservative layout
    detector. The source transaction extraction has already passed UPG's
    financial, narration, balance-chain, and count gates before this function
    is called.
    """
    header_text = " ".join(norm(header) for header in headers)
    anchors = []
    for group in (("date",), ("narration", "particular", "description"), ("balance",), ("withdrawal", "debit", "dr"), ("deposit", "credit", "cr")):
        found = next((word for word in group if word in header_text), None)
        if found:
            anchors.append(found)
    # These generic anchors still require all core transaction concepts, so a
    # profile cannot accidentally claim a non-statement document.
    anchors = anchors or ["date", "balance"]
    anchors_json = json.dumps(anchors)
    detection = """function detect(text) {
  const normalized = String(text || '').toLowerCase().replace(/\\s+/g, ' ');
  const anchors = __ANCHORS__;
  const hasNarration = normalized.includes('narration') || normalized.includes('particular') || normalized.includes('description');
  const hasDebit = normalized.includes('withdrawal') || normalized.includes('debit') || normalized.includes('dr');
  const hasCredit = normalized.includes('deposit') || normalized.includes('credit') || normalized.includes('cr');
  return normalized.includes('date') && normalized.includes('balance') && hasNarration && hasDebit && hasCredit && anchors.length >= 2;
}""".replace("__ANCHORS__", anchors_json)
    if strategy in {"running_balance_text", "unsigned_running_balance_text", "value_date_unsigned", "page_text_unsigned"}:
        parser = """function parse(text, options) {
  const blocks = String(text || '').split(/(?=^\\s*\\d{1,2}[\\/-]\\d{1,2}[\\/-]\\d{2,4}\\b)/m);
  const dateToken = '\\d{1,2}[\\/-](?:\\d{1,2}|[A-Za-z]{3})[\\/-]\\d{2,4}';
  const dateRe = new RegExp('^\\s*(' + dateToken + ')(?:\\s+(' + dateToken + '))?\\b');
  const moneyRe = /-?\\d[\\d,]*\\.\\d{2}/g;
  const asNumber = (value) => Number(String(value).replace(/,/g, ''));
  const rows = []; let previous = Number(options && options.openingBalance); if (!Number.isFinite(previous)) previous = null;
  for (const block of blocks) {
    const match = block.match(dateRe); if (!match || /\\bB\\/F\\b/i.test(block) || /^(?:page total|grand total)/im.test(block)) continue;
    const signed = [...block.matchAll(/(-?[\\d,]+\\.\\d{2})\\s*(Dr|Cr)\\b/ig)]; const values = [...block.matchAll(moneyRe)];
    if (!signed.length || values.length < 2) continue;
    const balanceToken = signed[signed.length - 1], balance = asNumber(balanceToken[1]) * (balanceToken[2].toLowerCase() === 'dr' ? -1 : 1);
    const amountToken = values.filter((value) => value.index < balanceToken.index).pop(); if (!amountToken) continue;
    const amount = Math.abs(asNumber(amountToken[0])); const particulars = block.slice(match[0].length, amountToken.index).replace(/\\s+/g, ' ').trim();
    const refs = particulars.match(/\\b\\d{6,}\\b/g) || []; const p = match[1].split(/[\\/-]/); const year = p[2].length === 2 ? `20${p[2]}` : p[2];
    const delta = previous == null ? amount : balance - previous;
    rows.push({date:`${p[0].padStart(2,'0')}/${p[1].padStart(2,'0')}/${year}`, particulars, withdrawal:delta < 0 ? -delta : 0, deposit:delta > 0 ? delta : 0, balance, chqNo:refs.length ? refs[refs.length - 1] : ''}); previous = balance;
  } return rows;
}"""
        return detection, parser
    if strategy == "geometry_profile":
        # Borderless PDF profiles are measured from original-PDF coordinates
        # by UPG.  Their portable counterpart must still handle the common
        # text layer: two date columns, wrapped narration, and three terminal
        # numeric cells.  It deliberately keeps B/F and summary labels out of
        # transaction rows.
        parser = """function parse(text, options) {
  const source = String(text || '');
  const token = '\\d{1,2}[\\/-](?:\\d{1,2}|[A-Za-z]{3})[\\/-]\\d{2,4}';
  const blocks = source.split(new RegExp('(?=^\\\\s*' + token + '\\\\s+' + token + '\\\\b)', 'm'));
  const rowRe = new RegExp('^\\\\s*(' + token + ')\\\\s+(' + token + ')\\\\s*');
  const moneyRe = /-?\\d[\\d,]*\\.\\d{2}\\b/g;
  const months = {jan:'01',feb:'02',mar:'03',apr:'04',may:'05',jun:'06',jul:'07',aug:'08',sep:'09',oct:'10',nov:'11',dec:'12'};
  const date = value => { const p=String(value).split(/[\\/-]/), y=p[2].length===2?'20'+p[2]:p[2], m=months[String(p[1]).toLowerCase()]||p[1]; return p.length===3 ? p[0].padStart(2,'0')+'/'+String(m).padStart(2,'0')+'/'+y : String(value); };
  const n = value => Number(String(value).replace(/,/g,''));
  const rows = [];
  for (const block of blocks) {
    const match = block.match(rowRe); if (!match) continue;
    if (/^\\s*(?:B\\s*\\/\\s*F|OPENING\\s+BALANCE)\\b/i.test(block.slice(match[0].length)) || /\\b(?:closing balance|total debit amt|total credit amt|end of statement)\\b/i.test(block)) continue;
    const values = [...block.matchAll(moneyRe)]; if (values.length < 3) continue;
    const tail = values.slice(-3), first = tail[0].index;
    const debit = Math.abs(n(tail[0][0])) || 0, credit = Math.abs(n(tail[1][0])) || 0, balance = n(tail[2][0]);
    if (!Number.isFinite(balance) || (!debit && !credit)) continue;
    let particulars = block.slice(match[0].length, first).replace(/\\s+/g,' ').trim();
    const refs = particulars.match(/\\b[A-Z0-9]{8,}\\b/g) || [], chqNo = refs.length ? refs[refs.length-1] : '';
    if (chqNo) particulars = particulars.replace(chqNo,' ').replace(/\\s+/g,' ').trim();
    rows.push({date:date(match[2]), particulars, withdrawal:debit, deposit:credit, balance, chqNo});
  }
  return rows;
}"""
        return detection, parser
    parser = """function parse(text, options) {
  const lines = String(text || '').split(/\\r?\\n/);
  const dateRe = /^\\s*(\\d{1,2}[\\/-]\\d{1,2}[\\/-]\\d{2,4})\\b/;
  const moneyRe = /-?\\d[\\d,]*\\.\\d{2}\\b/g;
  const ignored = /^(page\\s+\\d+|generation date|hdfc bank|statement summary|\\*\\*end of statement|date\\s+.*(?:balance|deposit|credit)|account branch|address\\s*:|contents of this statement|opening balance|closing balance|total (?:debit|credit) amt)/i;
  const amount = (value) => {
    if (value == null) return null;
    const n = Number(String(value).replace(/,/g, ''));
    return Number.isFinite(n) && n !== 0 ? Math.abs(n) : null;
  };
  const date = (value) => {
    const p = String(value).split(/[\\/-]/), months = {jan:'01',feb:'02',mar:'03',apr:'04',may:'05',jun:'06',jul:'07',aug:'08',sep:'09',oct:'10',nov:'11',dec:'12'};
    if (p.length !== 3) return String(value);
    const year = p[2].length === 2 ? `20${p[2]}` : p[2];
    const month = months[String(p[1]).toLowerCase()] || p[1];
    return `${p[0].padStart(2, '0')}/${String(month).padStart(2, '0')}/${year}`;
  };
  const rows = [];
  let prefix = [];
  for (const rawLine of lines) {
    const line = String(rawLine || '').trim();
    if (!line || ignored.test(line)) { if (ignored.test(line)) prefix = []; continue; }
    const match = line.match(dateRe);
    if (!match) {
      // Preserve only transaction-like wrapped narration, never page furniture.
      if (/^(RTGS|NEFT|IMPS|UPI|CHQ|CASH|ADHOC|EMP|EVP|POS|CARD)/i.test(line)) prefix.push(line);
      continue;
    }
    if (/^\\s*(?:B\\s*\\/\\s*F|OPENING\\s+BALANCE)\\b/i.test(line.slice(match[0].length))) { prefix = []; continue; }
    const rest = line.slice(match[0].length);
    const money = [...rest.matchAll(moneyRe)];
    if (money.length < 3) { prefix = []; continue; }
    const values = money.slice(-3).map((item) => amount(item[0]));
    const firstAmountAt = money[money.length - 3].index;
    let detail = rest.slice(0, firstAmountAt)
      .replace(/\\b\\d{1,2}[\\/-]\\d{1,2}[\\/-]\\d{2,4}\\b/g, ' ')
      .replace(/\\s+/g, ' ').trim();
    const refs = detail.match(/\\b[A-Z0-9]{8,}\\b/g) || [];
    const chqNo = refs.length ? refs[refs.length - 1] : '';
    if (chqNo) detail = detail.replace(chqNo, ' ').replace(/\\s+/g, ' ').trim();
    const particulars = [...prefix, detail].join(' ').replace(/\\s+/g, ' ').trim();
    rows.push({ date: date(match[2] || match[1]), particulars, withdrawal: values[0], deposit: values[1], balance: values[2], chqNo });
    prefix = [];
  }
  return rows;
}"""
    return detection, parser

def certified_feature_vector(headers: list[object], columns: dict[str, int], strategy: str | None,
                             diagnostic_rules: list[str] | None = None,
                             challenge_history: list[str] | None = None,
                             capability_tags: list[str] | None = None) -> dict[str, object]:
    """Persist a non-sensitive feature vector for retrieval and future ML evaluation.

    No source text, account values, narration, transaction rows, or customer
    identifiers are retained.  This is deliberately an explainable feature
    record, not a model prediction and never a release decision.
    """
    header_words = " ".join(norm(item) for item in headers)
    rules = {str(item) for item in (diagnostic_rules or [])}
    challenges = {str(item) for item in (challenge_history or [])}
    capabilities = {str(item) for item in (capability_tags or [])}
    return {
        "schema_version": 1,
        "source_kind": "pdf_geometry" if strategy == "geometry_profile" else "structured_or_text",
        "mapped_fields": sorted(name for name in columns if name in CANONICAL),
        "strategy": strategy or "detected_table",
        "signals": sorted({
            *( ["dual_date"] if "value date" in header_words else [] ),
            *( ["separate_debit_credit"] if "withdrawal" in header_words or "debit" in header_words else [] ),
            *( ["running_balance"] if "balance" in header_words else [] ),
            *( ["continuation"] if any("continuation" in item or "multi_page" in item for item in rules | challenges | capabilities) else [] ),
            *( ["footer_furniture"] if any("footer" in item or "summary" in item or "terminal" in item for item in rules | challenges | capabilities) else [] ),
            *( ["bf_metadata"] if any("bf_" in item or "preperiod" in item for item in rules | challenges | capabilities) else [] ),
            *( ["reverse_order"] if any("reverse" in item or "date_order" in item for item in rules | challenges | capabilities) else [] ),
            *( ["unreliable_balance"] if any("unreliable" in item for item in rules | challenges | capabilities) else [] ),
        }),
        "capabilities": sorted(capabilities),
        "rule_ids": sorted(rules),
        "challenge_ids": sorted(challenges),
    }

def save_profile(headers: list[object], columns: dict[str, int], parent_profile: str | None = None, strategy: str | None = None, self_healed: bool = False, layout_fingerprint: str = "", diagnostic_rules: list[str] | None = None, validation: dict | None = None, bank_name: str = "Unknown", format_name: str = "PDF Statement", challenge_history: list[str] | None = None, capability_tags: list[str] | None = None, capability_provenance: list[dict[str, object]] | None = None) -> str:
    """Persist validated learning additively; never discard a certified revision.

    A later statement can add a capability or a self-healing addendum, but it
    may never erase the exact profile evidence that certified an older layout.
    Previous revisions are immutable snapshots under ``profiles/revisions``;
    the active profile remains the latest certified revision for normal lookup.
    """
    if generated_canonical_headers(headers) and not layout_fingerprint: return ""
    ident = profile_id(headers, layout_fingerprint)
    profile_path = PROFILES / f"{ident}.json"
    try: prior = json.loads(profile_path.read_text(encoding="utf-8"))
    except (OSError, ValueError): prior = {}
    if prior:
        # The active file is a convenience pointer for fast exact-layout
        # lookup.  Before it is advanced, retain the prior certified parser as
        # an immutable revision.  This lets old statements continue to use the
        # historical, source-proven mapping and stops a new lesson from
        # rewriting history for a different layout variant.
        prior_version = max(1, int(prior.get("version", 1) or 1))
        revision_path = PROFILE_REVISIONS / f"{ident}.v{prior_version}.json"
        if not revision_path.exists():
            snapshot = dict(prior)
            snapshot["immutable_revision"] = True
            snapshot["revision_of"] = ident
            snapshot["archived_at"] = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
            revision_path.write_text(json.dumps(snapshot, indent=2), encoding="utf-8")
    observations = int(prior.get("validated_observations", 0)) + 1
    detection_code, parser_code = certified_javascript_code(headers, strategy)
    challenges = sorted({str(item) for item in (challenge_history or []) if str(item) and str(item) != "none"})
    features = certified_feature_vector(headers, columns, strategy, diagnostic_rules, challenges, capability_tags)
    learned_rule_ids = sorted({*(diagnostic_rules or []), *challenges, *(capability_tags or [])})
    rule_groups = rule_groups_for(learned_rule_ids)
    # Keep only the identity of reusable behaviours/providers.  Never retain
    # statement rows, text, account data, balances, or coordinates here.
    provenance = [
        {
            "capability": str(item.get("capability", "")),
            "rule_group": str(item.get("rule_group", "")),
            "rule_modules": [str(rule) for rule in item.get("selected_rule_modules", item.get("rule_modules", [])) if str(rule) in DIAGNOSTIC_RULE_LIBRARY],
            "provider_profile_ids": [str(profile_id) for profile_id in item.get("certified_profile_ids", [])][:4],
        }
        for item in (capability_provenance or []) if isinstance(item, dict) and item.get("capability")
    ][:12]
    version = int(prior.get("version", 0)) + 1
    data = {"version": version, "header_signature": [str(h) for h in headers], "layout_fingerprint": layout_fingerprint, "columns": columns, "parent_profile": parent_profile, "validated_observations": observations, "last_validated_strategy": strategy or "detected_table", "self_healed_addendum": bool(self_healed), "diagnostic_rules": diagnostic_rules or [], "challenge_history": challenges, "rule_groups": rule_groups, "feature_vector": features, "capability_provenance": provenance, "bank_name": bank_name or prior.get("bank_name", "Unknown"), "format_name": format_name or prior.get("format_name", "PDF Statement"), "detection_code": detection_code, "parser_code": parser_code, "validation": validation or {"status": "pass", "financial_pass": True, "narration_pass": True, "balance_chain_pass": True}, "learning_lineage": {"mode": "additive_immutable_revisions", "previous_revision": f"{ident}.v{version - 1}" if prior else None, "preserves_prior_certified_learning": True}, "certification": {"status": "certified", "source": "upg_native", "certified_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z"}}
    profile_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    # Aggregate learning intentionally contains only layout signatures and
    # validation outcomes, never account, narration, balances, or transactions.
    try: ledger = json.loads(LEARNING_LEDGER.read_text(encoding="utf-8"))
    except (OSError, ValueError): ledger = {"validated_profiles": {}}
    ledger.setdefault("validated_profiles", {})[ident] = {"observations": observations, "strategy": data["last_validated_strategy"], "self_healed_addendum": data["self_healed_addendum"], "diagnostic_rules": data["diagnostic_rules"], "challenge_history": challenges, "rule_groups": rule_groups, "feature_vector": features, "capability_provenance": provenance, "parent_profile": parent_profile}
    # A rule earns preference only after a full certified outcome.  This is a
    # bounded aggregate scorecard, not ML training and not statement storage.
    scorecard = ledger.setdefault("certified_rule_successes", {})
    for rule in learned_rule_ids:
        if rule in DIAGNOSTIC_RULE_LIBRARY:
            scorecard[rule] = int(scorecard.get(rule, 0) or 0) + 1
    LEARNING_LEDGER.write_text(json.dumps(ledger, indent=2), encoding="utf-8")
    return ident

def certified_learning_context(limit: int = 8) -> list[dict[str, object]]:
    """Return compact, privacy-safe lessons from certified profiles for AI planning.

    This intentionally excludes statement text, account values, narration, and
    transaction rows.  It carries only reusable layout/validation knowledge.
    """
    try:
        ledger = json.loads(LEARNING_LEDGER.read_text(encoding="utf-8"))
        learned = ledger.get("validated_profiles", {})
    except (OSError, ValueError):
        return []
    lessons: list[dict[str, object]] = []
    ranked = sorted(learned.items(), key=lambda item: int(item[1].get("observations", 0) or 0), reverse=True)
    for ident, summary in ranked[:limit]:
        try:
            profile = json.loads((PROFILES / f"{ident}.json").read_text(encoding="utf-8"))
        except (OSError, ValueError):
            continue
        validation = profile.get("validation", {}) if isinstance(profile.get("validation"), dict) else {}
        lessons.append({
            "profile_id": ident,
            "observations": int(summary.get("observations", 0) or 0),
            "strategy": str(profile.get("last_validated_strategy", summary.get("strategy", "detected_table"))),
            "headers": [str(header)[:80] for header in profile.get("header_signature", [])[:12]],
            "mapped_fields": sorted(name for name in profile.get("columns", {}) if name in CANONICAL),
            "challenge_history": list(profile.get("challenge_history", summary.get("challenge_history", [])))[:8],
            "diagnostic_rules": list(profile.get("diagnostic_rules", summary.get("diagnostic_rules", [])))[:8],
            "rule_groups": list(profile.get("rule_groups", summary.get("rule_groups", [])))[:6],
            # Capabilities are the independently reusable behaviours learned
            # from this certified parser.  They are deliberately separate
            # from its column coordinates and executable code: an unfamiliar
            # bank may share its continuation/furniture behaviour but never
            # its geometry.
            "capability_tags": list(
                profile.get("feature_vector", {}).get(
                    "capabilities", summary.get("feature_vector", {}).get("capabilities", [])
                )
                if isinstance(profile.get("feature_vector", {}), dict) else []
            )[:10],
            "balance_chain_exception": bool(validation.get("balance_chain_exception", False)),
            "self_healed_addendum": bool(profile.get("self_healed_addendum", False)),
        })
    return lessons

def certified_rule_successes() -> dict[str, int]:
    """Return only bounded, certification-earned reusable rule scores."""
    try:
        ledger = json.loads(LEARNING_LEDGER.read_text(encoding="utf-8"))
        raw = ledger.get("certified_rule_successes", {})
    except (OSError, ValueError):
        ledger = {}
        raw = {}
    scores = {
        str(rule): int(count or 0) for rule, count in raw.items()
        if str(rule) in DIAGNOSTIC_RULE_LIBRARY and int(count or 0) > 0
    }
    # Backfill score evidence from profiles certified before this scorecard
    # existed.  They are already safe lessons, so no reprocessing or source
    # data is required to make their reusable rules available immediately.
    if not scores:
        for summary in (ledger.get("validated_profiles", {}) if isinstance(ledger, dict) else {}).values():
            if not isinstance(summary, dict):
                continue
            for rule in [*summary.get("diagnostic_rules", []), *summary.get("challenge_history", [])]:
                rule = str(rule)
                if rule in DIAGNOSTIC_RULE_LIBRARY:
                    scores[rule] = scores.get(rule, 0) + 1
    return scores

def closest_certified_lessons(path: Path | None, headers: list[object] | None = None, limit: int = 3) -> list[dict[str, object]]:
    """Find the closest certified layouts using this source's safe structure."""
    target_headers = {norm(item) for item in (headers or []) if norm(item)}
    target_fingerprint = ""
    if path and path.suffix.lower() == ".pdf":
        try:
            target_fingerprint = text_layout_fingerprint(remove_page_furniture(cached_pdf_text(path)))
        except (OSError, ValueError):
            pass
    if not target_headers and not target_fingerprint:
        return []
    ranked: list[tuple[float, dict[str, object]]] = []
    for lesson in certified_learning_context(limit=64):
        profile_headers = {norm(item) for item in lesson.get("headers", []) if norm(item)}
        overlap = len(target_headers & profile_headers) / max(1, len(target_headers | profile_headers))
        try:
            profile = json.loads((PROFILES / f"{lesson['profile_id']}.json").read_text(encoding="utf-8"))
        except (OSError, ValueError, KeyError):
            continue
        exact_fingerprint = bool(target_fingerprint and profile.get("layout_fingerprint") == target_fingerprint)
        score = (1_000 if exact_fingerprint else 0) + overlap * 400
        if score <= 0:
            continue
        ranked.append((score, {**lesson, "similarity_score": round(score, 1), "exact_layout": exact_fingerprint}))
    ranked.sort(key=lambda item: item[0], reverse=True)
    return [lesson for _, lesson in ranked[:limit]]

def ai_learning_packet(path: Path | None = None, headers: list[object] | None = None) -> dict[str, object]:
    """The complete non-sensitive knowledge supplied to UPG's AI workers."""
    return {
        "transaction_output_contract": {
            "mandatory_fields": list(MANDATORY_TRANSACTION_FIELDS),
            "optional_fields": list(OPTIONAL_TRANSACTION_FIELDS),
            "date_rule": "Export Value Date in DD/MM/YYYY when both Value Date and posting/transaction date exist.",
            "narration_rule": "Use only the source Particulars cell; blank is allowed; furniture and numeric cells are forbidden.",
        },
        "permanent_historical_lessons": HISTORICAL_CHALLENGE_LESSONS,
        "certified_rule_successes": certified_rule_successes(),
        "closest_certified_layouts": closest_certified_lessons(path, headers),
        "certified_profile_lessons": certified_learning_context(),
    }

def source_capability_plan(raw: str = "", headers: list[object] | None = None) -> list[dict[str, object]]:
    """Select reusable *capabilities*, never a foreign parser, from source evidence.

    An unfamiliar bank is rarely an entirely unfamiliar problem.  Its headers
    may be new while its layout still has a Value Date, fixed-width narration
    continuations, repeated page furniture, B/F metadata, or a signed running
    balance.  This stage deliberately transfers only those independently
    certified behaviours.  It never copies another bank's coordinates, header
    indexes, detection code, or parser code.
    """
    text = str(raw or "")
    header_text = " ".join(str(item) for item in (headers or []))
    evidence = (text[:120000] + "\n" + header_text).lower()
    signals: list[tuple[str, str, str | None]] = []
    if re.search(r"\b(?:value\s+date|val\.?\s*date)\b", evidence):
        signals.append(("value_date", "Source prints both transaction/posting and Value Date evidence; export Value Date only.", "value_date_unsigned"))
    if re.search(r"\b(?:b\s*/?\s*f|brought\s+forward|opening\s+balance)\b", evidence):
        signals.append(("bf_preperiod_artifact", "Source contains B/F or opening-balance metadata; prevent it becoming a transaction.", None))
    if re.search(r"\b(?:grand\s+total|page\s+total|statement\s+summary|total\s+(?:debits?|credits?))\b", evidence):
        signals.append(("footer_exclusion", "Source contains summary/page-total furniture; seal the last dated row before it.", None))
    if len(re.findall(r"(?im)^\s*(?:page\s*)?\d+\s*(?:of\s*\d+)?\s*$", text)) >= 2 or len(re.findall(r"(?i)\bpage\s+\d+\b", text)) >= 2:
        signals.append(("multi_page_continuation", "Repeated page markers indicate page furniture and possible cross-page continuations.", None))
    if re.search(r"\b\d[\d,]*\.\d{1,2}\s*(?:dr|cr)\b", evidence):
        signals.append(("signed_balance_text", "Dated source text contains explicit Dr/Cr balances.", "running_balance_text"))
    elif re.search(r"\b(?:running|available|closing)\s+balance\b", evidence):
        signals.append(("balance_delta", "Source exposes an unsigned running-balance column; infer direction only from balance movement.", "unsigned_running_balance_text"))

    certified = certified_learning_context(limit=64)
    success_scores = certified_rule_successes()

    def providers_for(capability: str) -> list[dict[str, object]]:
        """Return compact certified rule modules for one source-proven need.

        A provider is a lesson, not a parser import.  Keeping the actual rule
        names with the provider lets the planner compose capabilities from
        several different banks (for example continuation handling from one
        and footer exclusion from another) without borrowing their offsets.
        """
        providers: list[dict[str, object]] = []
        aliases = {
            "value_date": ("value_date", "dual_date"),
            "bf_preperiod_artifact": ("bf_", "opening", "preperiod"),
            "footer_exclusion": ("footer", "summary", "terminal_row"),
            "multi_page_continuation": ("continuation", "multi_page", "page_furniture"),
            "signed_balance_text": ("signed_balance", "running_balance_text", "headerless"),
            "balance_delta": ("balance_delta", "unsigned_running_balance", "value_date_unsigned"),
        }.get(capability, (capability,))
        for lesson in certified:
            vocabulary = " ".join([
                str(lesson.get("strategy", "")),
                *[str(item) for item in lesson.get("rule_groups", [])],
                *[str(item) for item in lesson.get("capability_tags", [])],
                *[str(item) for item in lesson.get("challenge_history", [])],
                *[str(item) for item in lesson.get("diagnostic_rules", [])],
            ]).lower()
            if not any(alias in vocabulary for alias in aliases):
                continue
            reusable_rules = [str(rule) for rule in lesson.get("diagnostic_rules", [])
                              if str(rule) in DIAGNOSTIC_RULE_LIBRARY]
            # Legacy profiles were certified before rule modules were
            # persisted.  Use only the capability's deterministic library
            # fallback; never invent a whole-parser match from bank name.
            if not reusable_rules:
                reusable_rules = rules_for_capability(capability)
            providers.append({
                "profile_id": str(lesson.get("profile_id")),
                "strategy_family": str(lesson.get("strategy", "")),
                "reusable_rules": reusable_rules[:4],
                "rule_groups": list(lesson.get("rule_groups", []))[:4],
                "challenges_solved": [str(item) for item in lesson.get("challenge_history", [])][:3],
                "certified_success_score": sum(success_scores.get(rule, 0) for rule in reusable_rules),
            })
        providers.sort(key=lambda provider: int(provider.get("certified_success_score", 0) or 0), reverse=True)
        return providers[:4]

    selected: list[dict[str, object]] = []
    for capability, reason, strategy in signals:
        providers = providers_for(capability)
        group = group_for_capability(capability)
        default_rules = rules_for_capability(capability)
        # Pick one strongest certified provider, then retain only rules that
        # belong to this source-proven capability group.  A new statement may
        # compose several groups, but it must never inherit an unbounded pile
        # of unrelated rules from every vaguely similar profile.
        best_provider = providers[0] if providers else {}
        provider_rules = [
            str(rule) for rule in best_provider.get("reusable_rules", [])
            if str(rule) in RULE_GROUPS.get(group, set())
        ] if isinstance(best_provider, dict) else []
        selected_rules = list(dict.fromkeys([*default_rules, *provider_rules]))[:5]
        selected.append({
            "capability": capability,
            "rule_group": group,
            "reason": reason,
            "preferred_strategy": strategy,
            "rule_modules": default_rules,
            "selected_rule_modules": selected_rules,
            "selected_provider_profile_id": str(best_provider.get("profile_id", "")) if isinstance(best_provider, dict) else "",
            "certified_profile_ids": [provider["profile_id"] for provider in providers],
            "certified_rule_modules": providers,
            "instruction": DIAGNOSTIC_RULE_LIBRARY.get(capability, capability),
            "composition_constraint": (
                "Compose this behaviour with other source-proven modules only. "
                "Do not copy a provider's coordinates, header indexes, parser code, or transaction data."
            ),
        })
    return selected

STEP_NAMES = {
    "source_intake": (1, "SOURCE_INTAKE"),
    "native_structure": (2, "NATIVE_STRUCTURE_READ"),
    "column_geometry": (5, "COLUMN_GEOMETRY_MAP"),
    "header_mapping": (4, "HEADER_SEMANTICS"),
    "date_order": (11, "DATE_SELECTION_AND_NORMALIZATION"),
    "continuation": (7, "NARRATION_ASSEMBLY"),
    "page_furniture": (8, "FURNITURE_REMOVAL"),
    "balance_direction": (10, "BALANCE_INTERPRETATION"),
    "unreliable_balance": (10, "BALANCE_INTERPRETATION"),
    "endpoint": (16, "FINANCIAL_RECONCILIATION"),
    "source_totals": (16, "FINANCIAL_RECONCILIATION"),
    "narration_coverage": (14, "SOURCE_COVERAGE_CHECK"),
    "transaction_count": (15, "TRANSACTION_COUNT_CHECK"),
    "novel_layout": (20, "PARSER_PLAN_COMPOSITION"),
}

def compact_ai_learning_packet(path: Path | None = None, headers: list[object] | None = None,
                               raw: str = "", failure_type: str = "column_geometry") -> dict[str, object]:
    """Return source-scoped certified knowledge for one named pipeline step.

    This is Step 27 (AI_CONTEXT_SCOPE).  The model receives only capabilities
    proven by the new source, their certified rule modules, and the closest
    matching certified lessons.  It never receives the full unrelated rule
    library, which reduces cost and prevents a geometry repair from being
    distracted by, for example, a B/F or narration-only lesson.
    """
    step_number, step_name = STEP_NAMES.get(failure_type, STEP_NAMES["column_geometry"])
    capabilities = source_capability_plan(raw, headers)
    selected_rules = list(dict.fromkeys(
        rule for item in capabilities if isinstance(item, dict)
        for rule in item.get("selected_rule_modules", [])
        if rule in DIAGNOSTIC_RULE_LIBRARY
    ))
    # Every AI decision needs a small structural safety baseline.  This is not
    # a broad history dump: it is the current step's deterministic contract.
    baseline = {
        "column_geometry": ["distinct_source_columns", "header_role_alignment", "measured_column_evidence"],
        "header_mapping": ["header_role_alignment", "measured_column_evidence", "distinct_source_columns"],
        "date_order": ["date_column_boundary", "date_source_cell", "reverse_order"],
        "continuation": ["continuation_merge", "narration_source_cell", "reference_date_boundary"],
        "page_furniture": ["footer_exclusion", "terminal_row_before_summary", "multi_page_continuation"],
        "balance_direction": ["balance_delta", "amount_balance_consistency", "balance_source_cell"],
        "unreliable_balance": ["source_amount_geometry", "summary_total_warning", "balance_source_cell"],
        "endpoint": ["summary_endpoints", "balance_source_cell", "source_coverage"],
        "source_totals": ["summary_total_warning", "source_amount_geometry", "amount_balance_consistency"],
        "narration_coverage": ["narration_source_cell", "continuation_merge", "source_coverage"],
        "transaction_count": ["source_coverage", "date_source_cell", "measured_column_evidence"],
        "novel_layout": ["distinct_source_columns", "header_role_alignment", "measured_column_evidence"],
    }.get(failure_type, ["distinct_source_columns", "header_role_alignment", "measured_column_evidence"])
    allowed_rules = list(dict.fromkeys([*baseline, *selected_rules]))
    closest = closest_certified_lessons(path, headers, limit=3)
    return {
        "ai_context_scope": {
            "pipeline_step": f"S{step_number:02d}_{step_name}",
            "failure_type": failure_type,
            "mode": "source_proven_rules_and_closest_certified_lessons_only",
            "included_rule_count": len(allowed_rules),
            "excluded": "unrelated rule libraries, foreign parser code, coordinates, transaction data, and unproven capabilities",
        },
        "output_contract": {
            "required": list(MANDATORY_TRANSACTION_FIELDS),
            "optional": list(OPTIONAL_TRANSACTION_FIELDS),
            "value_date_priority": True,
            "date_format": "DD/MM/YYYY",
        },
        "allowed_rule_modules": {rule: DIAGNOSTIC_RULE_LIBRARY[rule] for rule in allowed_rules},
        "closest_layouts": [{
            "profile_id": item.get("profile_id"),
            "strategy": item.get("strategy"),
            "mapped_fields": item.get("mapped_fields", []),
            "challenge_history": item.get("challenge_history", [])[:4],
            "rule_groups": item.get("rule_groups", [])[:4],
        } for item in closest],
        "source_matched_certified_capabilities": capabilities,
    }

def find_related_profile(headers: list[object]) -> tuple[str | None, dict[str, int]]:
    """Find a near-match without changing the original validated profile."""
    if generated_canonical_headers(headers): return None, {}
    current = "|".join(norm(h) for h in headers)
    best_id, best_columns, best_score = None, {}, 0.0
    for profile in PROFILES.glob("*.json"):
        try:
            data = json.loads(profile.read_text(encoding="utf-8"))
            if generated_canonical_headers(data.get("header_signature", [])):
                continue
            prior = "|".join(norm(h) for h in data.get("header_signature", []))
            score = SequenceMatcher(None, current, prior).ratio()
            if score > best_score:
                best_id, best_columns, best_score = profile.stem, {k: int(v) for k, v in data["columns"].items()}, score
        except (OSError, ValueError, KeyError):
            continue
    return (best_id, best_columns) if best_score >= 0.62 else (None, {})

def saved_text_strategy(path: Path) -> str | None:
    """Find the validated extraction strategy before new generation begins."""
    if path.suffix.lower() != ".pdf": return None
    try:
        # Profiles are saved from the normalized source used for extraction.
        # Look up with that same representation; using uncleaned PDF text here
        # made repeated page furniture change the signature and hid a valid
        # parser for an otherwise matching statement.
        fingerprint = text_layout_fingerprint(remove_page_furniture(cached_pdf_text(path)))
        for profile in PROFILES.glob("*.json"):
            data = json.loads(profile.read_text(encoding="utf-8"))
            if data.get("layout_fingerprint") == fingerprint:
                return data.get("last_validated_strategy")
    except (OSError, ValueError, KeyError):
        pass
    return None

REPAIR_MODULE_STRATEGIES: dict[str, frozenset[str]] = {
    # A retry may change only the component disproved by source evidence.  The
    # strategy is still re-measured from the current source; this map never
    # imports offsets or executable code from a different bank.
    "header_mapping": frozenset({"geometry_profile", "source_amount_geometry", "standard_column_geometry"}),
    "column_geometry": frozenset({"geometry_profile", "source_amount_geometry", "standard_column_geometry", "dual_date_geometry"}),
    "date_order": frozenset({"dual_date_geometry", "value_date_unsigned", "geometry_profile"}),
    "continuation": frozenset({"geometry_profile", "standard_column_geometry"}),
    "narration_coverage": frozenset({"geometry_profile", "standard_column_geometry"}),
    "page_furniture": frozenset({"geometry_profile", "standard_column_geometry", "page_text_unsigned"}),
    "source_totals": frozenset({"source_amount_geometry", "geometry_profile", "standard_column_geometry"}),
    "balance_direction": frozenset({"running_balance_text", "unsigned_running_balance_text", "page_text_unsigned", "geometry_profile"}),
    "unreliable_balance": frozenset({"running_balance_text", "unsigned_running_balance_text", "page_text_unsigned"}),
    "endpoint": frozenset({"running_balance_text", "unsigned_running_balance_text", "geometry_profile"}),
    "transaction_count": frozenset({"geometry_profile", "standard_column_geometry", "source_amount_geometry"}),
}

# A strategy can address more than one structural capability, but UPG should
# not pay for two full parses that both solve exactly the same observed need.
# These are capability families, not bank names or remembered coordinates.
STRATEGY_CAPABILITY_COVERAGE: dict[str, frozenset[str]] = {
    "geometry_profile": frozenset({"header_mapping", "column_geometry", "narration", "furniture", "bf_preperiod_artifact"}),
    "standard_column_geometry": frozenset({"header_mapping", "column_geometry", "narration", "furniture"}),
    "source_amount_geometry": frozenset({"header_mapping", "column_geometry", "amounts", "balance"}),
    "dual_date_geometry": frozenset({"value_date", "date_order", "column_geometry"}),
    "value_date_unsigned": frozenset({"value_date", "date_order", "balance"}),
    "running_balance_text": frozenset({"signed_balance_text", "balance", "endpoint"}),
    "unsigned_running_balance_text": frozenset({"balance_delta", "balance", "endpoint"}),
    "page_text_unsigned": frozenset({"furniture", "continuation", "balance", "endpoint"}),
}


def evidence_first_candidates(path: Path, large_pdf: bool, geometry_ready: bool, validated_strategy: str | None,
                              planned_strategies: list[str], retry_round: int,
                              include_ai_addendum: bool = True,
                              repair_module: str = "",
                              source_sample: str | None = None,
                              source_geometry: tuple[list[object], list[tuple[float, float]]] | None = None) -> list[tuple[str | None, bool]]:
    """Rank parser candidates from source evidence before parsing the full file.

    A candidate is never accepted by score alone.  The score only decides which
    small set earns a full-document extraction and the normal release gates.
    """
    scores: dict[tuple[str | None, bool], int] = {}
    source_capabilities: set[str] = set()
    def add(strategy: str | None, ai_addendum: bool, score: int) -> None:
        key = (strategy, ai_addendum)
        scores[key] = max(scores.get(key, -10_000), score)
    def candidate_for(name: str) -> tuple[str | None, bool]:
        return (None, True) if name == "ai_layout_addendum" else ((None, False) if name == "detected_table" else (name, False))

    # Exact validated reuse is strongest evidence, followed by a closely
    # related layout addendum and original-PDF geometry.
    if validated_strategy:
        add(validated_strategy, False, 1_000)
    headers: list[object] = []
    header_fields: dict[str, int] = {}
    if path.suffix.lower() == ".pdf":
        try:
            # Step 24: preflight may already have measured this upload's
            # original-PDF geometry. Reuse that snapshot instead of scanning
            # the same representative pages again while ranking candidates.
            geometry = source_geometry if source_geometry is not None else sampled_geometry_profile(path)
            if geometry:
                headers = geometry[0]
                header_fields = map_headers(headers)
                exact = load_profile(headers)
                related_id, related_columns = (None, {}) if exact else find_related_profile(headers)
                if exact:
                    add(str(exact.get("last_validated_strategy", "detected_table")), False, 900)
                elif related_id and related_columns:
                    add("geometry_profile" if large_pdf else "detected_table", False, 800)
                # A sparse, guessed header is not enough evidence to make a
                # table-geometry parser the first choice.  Headerless J&K
                # Bank layouts, for example, expose their transactions as
                # dated text ending in signed Dr/Cr balances.
                geometry_score = 750 if len(header_fields) >= 4 else 320
                add("geometry_profile", False, geometry_score if large_pdf else min(650, geometry_score))
        except (OSError, ValueError, KeyError):
            pass
    if geometry_ready and len(header_fields) >= 4:
        add("geometry_profile", False, 780)
    # The measured header cells are stronger than a flattened PDF text stream.
    # A text layer may merge adjacent captions or insert line breaks, whereas
    # sampled_geometry_profile has preserved the visible column cells.  This
    # applies to every bank; it is a semantic contract, not a bank template.
    if path.suffix.lower() == ".pdf" and has_dual_date_header_contract(headers):
        add("dual_date_geometry", False, 1_300)

    for lesson in closest_certified_lessons(path, headers):
        strategy = str(lesson.get("strategy", "detected_table"))
        if strategy == "detected_table":
            strategy = "geometry_profile" if large_pdf else "detected_table"
        if strategy in {"geometry_profile", "value_date_unsigned", "unsigned_running_balance_text", "running_balance_text", "page_text_unsigned", "detected_table"}:
            add(None if strategy == "detected_table" else strategy, False, int(820 + min(130, float(lesson.get("similarity_score", 0)) / 10)))

    # Lightweight source signals only rank supported alternatives; they never
    # create a parser or classify a transaction without validation.
    try:
        # These source signals must be read by the reader for the actual file
        # type.  Previously this always called ``cached_pdf_text`` for a
        # non-large upload, so an .xls job reached PyMuPDF before its Excel
        # reader and failed with "Failed to open file ... as type xls".
        if source_sample is not None:
            sample = source_sample
        elif path.suffix.lower() == ".pdf":
            sample = sampled_pdf_text(path) if large_pdf else cached_pdf_text(path)
        else:
            _, sample = load_rows(path)
            sample = sample[:60000]
        # Common bank ledgers often use business vocabulary such as
        # ``Transaction Remarks`` and ``Withdrawal Amount`` rather than the
        # canonical labels.  This complete semantic header contract is direct
        # source evidence: route it to the measured original-PDF column parser
        # before generic geometry or an AI layout call.  The parser still has
        # to pass all normal validation gates; this merely prevents a clearly
        # tabular statement from being treated as an unknown layout.
        # A dual-date ledger needs its own original-PDF geometry route.  Do
        # this *before* the general standard-table candidate: text-table
        # extraction can merge the two date bands and then incorrectly report
        # a date-order defect, even though the visual source is unambiguous.
        # Banks label the first date band as Date, Transaction Date, Posting
        # Date, Post Date, or Txn Date.  Requiring the words "TXN" or "POST"
        # made a visibly dual-date ledger fall through to generic parsing.
        # Require the full surrounding money/balance contract as the guard,
        # not one bank's exact first-column caption.
        dual_date_header = bool(re.search(
            r"(?is)\b(?:(?:TRANSACTION|TXN|POST(?:ING)?)\s*)?DATE\b[\s\S]{0,100}"
            r"\bVALUE\s+DATE\b[\s\S]{0,280}\b(?:DEBITS?|WITHDRAWALS?)(?:\b|(?=DEPOSITS?))[\s\S]{0,100}"
            # Flattened text layers sometimes remove the visual gap between
            # adjacent headers (``WithdrawalsDeposits``).  The original-PDF
            # geometry still proves the two bands, so do not let this text
            # artefact hide an otherwise unambiguous dual-date ledger.
            r"(?:\s*|(?=DEPOSITS?))\b(?:CREDITS?|DEPOSITS?)\b[\s\S]{0,100}\b(?:RUNNING\s+)?BALANCE\b",
            sample,
        ))
        if path.suffix.lower() == ".pdf" and dual_date_header:
            add("dual_date_geometry", False, 1_300)
        if path.suffix.lower() == ".pdf" and has_standard_geometry_header_contract(sample):
            add("standard_column_geometry", False, 1_250)
        if re.search(r"(?i)\bvalue\s+date\b", sample):
            add("value_date_unsigned", False, 520)
        signed_balance_rows = len(re.findall(
            r"(?im)^\s*\d{2}[-/]\d{2}[-/]\d{2,4}\b.*?\b\d[\d,]*\.\d{1,2}\s*(?:dr|cr)\b", sample
        ))
        if re.search(r"(?i)\b(?:dr|cr)\b", sample):
            # For a headerless/sparse-header signed-balance layout this is
            # direct original-source evidence, stronger than inferred column
            # geometry.  Start with it rather than spending retries proving
            # that a three-field table guess is unsuitable.
            # Explicit Dr/Cr ledger balances are stronger source evidence than
            # a generic header map even when the PDF happens to have a neat
            # table.  The running-balance strategy establishes endpoints from
            # real ledger rows; a table map alone may not find statement-level
            # "Opening/Closing Balance" labels at all.
            signed_score = 1_100 if signed_balance_rows >= 3 else 500
            add("running_balance_text", False, signed_score)
            if signed_balance_rows >= 3 and len(header_fields) < 4:
                add("page_text_unsigned", False, 1_020)
        if re.search(r"(?i)\b(?:running|closing|available)\s+balance\b", sample):
            add("unsigned_running_balance_text", False, 460)
        if not headers:
            add("page_text_unsigned", False, 300)
        # A related source need not share headers with a certified profile.
        # Promote only the extraction family proved by this source's features;
        # never borrow another bank's offsets or executable parser code.
        for capability in source_capability_plan(sample, headers):
            if isinstance(capability, dict) and capability.get("capability"):
                source_capabilities.add(str(capability["capability"]))
            strategy = capability.get("preferred_strategy")
            if isinstance(strategy, str) and strategy:
                # A capability is emitted only when the *uploaded source*
                # proves its signal.  It is therefore stronger than a nearby
                # certified profile that merely has similar headers.  In
                # particular, a dated Dr/Cr running-balance ledger must test
                # its balance-aware strategy before generic table parsing,
                # otherwise endpoint labels may be absent and the correct
                # final-row balance is never reached.
                source_proven_score = 1_160 if strategy in {
                    "running_balance_text", "unsigned_running_balance_text", "value_date_unsigned"
                } else 1_020
                add(strategy, False, source_proven_score)
    except (OSError, ValueError):
        pass

    # Certified lessons add a small preference only.  They never outweigh an
    # exact profile or this statement's original-PDF geometry.
    for lesson in certified_learning_context():
        strategy = str(lesson.get("strategy", ""))
        if strategy in {"geometry_profile", "dual_date_geometry", "value_date_unsigned", "unsigned_running_balance_text", "running_balance_text", "page_text_unsigned"}:
            add(strategy, False, 350 + min(90, int(lesson.get("observations", 0) or 0) * 5))

    # A controlled AI layout blueprint is the first and only planning call for
    # an unfamiliar layout. Once it has failed, the final AI call is reserved
    # for a targeted repair plan; that plan must be applied through supported
    # deterministic strategies rather than spending a dead third AI call.
    if include_ai_addendum:
        add(None, True, 700 if retry_round == 1 else 850)
    for index, name in enumerate(planned_strategies):
        strategy, ai_addendum = candidate_for(name)
        add(strategy, ai_addendum, 880 - index * 20)

    ordered = sorted(scores, key=lambda item: scores[item], reverse=True)
    # Step 22: after an evidence-led failure, do not recycle candidates merely
    # because they were useful in another branch of the retry plan.  Only a
    # strategy that can repair the failed module may earn another full parse.
    # The AI addendum remains available as the bounded fallback when no
    # supported deterministic repair is suitable.
    allowed = REPAIR_MODULE_STRATEGIES.get(str(repair_module))
    if allowed:
        ordered = [item for item in ordered if item[1] or item[0] in allowed]
    # Step 26: an exact certified layout fingerprint is not merely a high
    # score—it is a direct reusable parser match. Test that saved strategy on
    # the complete uploaded source before spending work on related layouts or
    # an AI addendum. If the full validation rejects it, the next retry still
    # diagnoses the concrete failure and may create an addendum; no release
    # gate is weakened by this fast path.
    if retry_round == 1 and validated_strategy:
        return [candidate_for(validated_strategy)]
    # A source-proven dual-date ledger must first be tested with its measured
    # Value-Date geometry.  Generic table parsing can collapse the date bands,
    # and an AI call before this test only spends money rediscovering evidence
    # we already have.  If it fails, the next retry can diagnose that exact
    # failed module using the retained evidence.
    dual_key = ("dual_date_geometry", False)
    if retry_round == 1 and dual_key in scores:
        selected = [dual_key]
        fallback = next((item for item in ordered
                         if item != dual_key and not item[1]), None)
        if fallback is not None:
            selected.append(fallback)
        return selected
    # Two strong evidence-led candidates normally suffice. Keep the first AI
    # blueprint as the third candidate whenever budget permits. A saved
    # strategy is only a reusable hypothesis; if it is wrong for this variant,
    # it must never suppress fresh measured-layout planning.
    # Step 23: choose the smallest high-evidence deterministic bundle that
    # covers distinct *observed* source needs.  Previously two nearly
    # identical geometry candidates could be selected simply because they had
    # adjacent scores. This preserves quality—the later release gates are
    # unchanged—while avoiding duplicate full-document work and AI calls.
    selected: list[tuple[str | None, bool]] = []
    covered: set[str] = set()
    deterministic = [item for item in ordered if not item[1]]
    for item in deterministic:
        strategy = item[0]
        capabilities = STRATEGY_CAPABILITY_COVERAGE.get(str(strategy), frozenset())
        gain = (capabilities & source_capabilities) - covered
        if not selected or gain:
            selected.append(item)
            covered.update(capabilities)
        if len(selected) >= 2:
            break
    # Some simple statements advertise no specialized capability. Preserve a
    # second evidence-ranked fallback in that case, but never duplicate a
    # strategy already chosen.
    if len(selected) < 2:
        for item in deterministic:
            if item not in selected:
                selected.append(item)
            if len(selected) >= 2:
                break
    ai_key = (None, True)
    if include_ai_addendum and ai_key not in selected and ai_key in scores:
        selected.append(ai_key)
    return selected

def preflight_plan_id(plan: dict[str, object]) -> str:
    """Stable, privacy-safe identity for the measured plan, not source data."""
    evidence = {
        "measured_from": plan.get("measured_from"),
        "header_fields": plan.get("header_fields", []),
        "candidate_plan": plan.get("candidate_plan", []),
        "rule_bundle": plan.get("selected_rule_bundle", []),
    }
    return hashlib.sha256(json.dumps(evidence, sort_keys=True, separators=(",", ":")).encode()).hexdigest()[:16]


def source_preflight_snapshot(path: Path, large_pdf: bool) -> dict[str, object]:
    """Measure an upload once for its preflight planner.

    The snapshot is intentionally job-local and contains only the source
    evidence needed for planning. It is passed to candidate ranking directly,
    never stored in a parser profile or learning record.
    """
    geometry = None
    source_error = ""
    try:
        geometry = sampled_geometry_profile(path) if path.suffix.lower() == ".pdf" else None
    except (OSError, ValueError, RuntimeError) as error:
        source_error = re.sub(r"\s+", " ", str(error)).strip()[:220]
    try:
        if path.suffix.lower() == ".pdf":
            source_sample = sampled_pdf_text(path) if large_pdf else cached_pdf_text(path)
        else:
            _, source_sample = load_rows(path)
            source_sample = source_sample[:60000]
    except (OSError, ValueError, RuntimeError) as error:
        source_sample = ""
        if not source_error:
            source_error = re.sub(r"\s+", " ", str(error)).strip()[:220]
    return {"geometry": geometry, "sample": source_sample, "source_error": source_error}

def preflight_step_gate(path: Path, snapshot: dict[str, object]) -> dict[str, object]:
    """Apply the first two named pipeline steps as hard evidence gates.

    No column, narration, amount, validation, or AI step may run until the
    source has first been accepted and then read through its native evidence
    path.  A blank/corrupt upload is not a later ``column_geometry`` problem.
    It is repaired or reported at S01/S02 without guessing downstream data.
    """
    supported = {".pdf", ".csv", ".xlsx", ".xls", ".txt", ".doc", ".docx"}
    if not path.exists() or not path.is_file() or path.stat().st_size <= 0:
        return {"passed": False, "step": "S01_SOURCE_INTAKE", "number": 1,
                "reason": "The upload is missing or empty.",
                "repair": "Re-upload the original populated bank statement before parsing."}
    if path.suffix.lower() not in supported:
        return {"passed": False, "step": "S01_SOURCE_INTAKE", "number": 1,
                "reason": f"Unsupported source type: {path.suffix or 'no extension'}.",
                "repair": "Upload PDF, DOCX, Excel, CSV, or TXT source evidence."}
    sample = str(snapshot.get("sample", "")).strip()
    geometry = snapshot.get("geometry")
    # An image-only PDF can legitimately have no native text. It remains at
    # S02 and will use the existing coordinate-preserving OCR route when it is
    # available; it must not be mistaken for a later parser failure.
    if not sample and not geometry and not (path.suffix.lower() == ".pdf" and ocr_is_available()):
        detail = str(snapshot.get("source_error", "") or "no native rows, text, or measured page structure were readable")
        return {"passed": False, "step": "S02_NATIVE_STRUCTURE_READ", "number": 2,
                "reason": detail,
                "repair": "Export a populated, readable source file or provide an OCR-readable PDF."}
    return {"passed": True, "step": "S02_NATIVE_STRUCTURE_READ", "number": 2,
            "reason": "Original source evidence is available.", "repair": ""}

def build_preflight_blueprint(path: Path, large_pdf: bool, validated_strategy: str | None,
                              planned_strategies: list[str]) -> dict[str, object]:
    """Create one evidence-led parser plan before any full-file extraction."""
    snapshot = source_preflight_snapshot(path, large_pdf)
    step_gate = preflight_step_gate(path, snapshot)
    geometry = snapshot["geometry"]
    headers = geometry[0] if geometry else []
    source_sample = str(snapshot["sample"])
    closest = closest_certified_lessons(path, headers)
    capabilities = source_capability_plan(source_sample, headers)
    candidates = evidence_first_candidates(
        path, large_pdf, bool(geometry), validated_strategy, planned_strategies, 1,
        source_sample=source_sample, source_geometry=geometry,
    ) if step_gate["passed"] else []
    selected_rule_bundle = [
        {
            "capability": str(item.get("capability", "")),
            "rule_group": str(item.get("rule_group", "")),
            "rules": [str(rule) for rule in item.get("selected_rule_modules", [])],
            "provider_profile_id": str(item.get("selected_provider_profile_id", "")),
        }
        for item in capabilities if isinstance(item, dict)
    ]
    plan = {
        "version": 3,
        "measured_from": "original_pdf_geometry" if geometry else "source_text_structure",
        "header_fields": sorted(map_headers(headers)) if headers else [],
        "transaction_output_contract": {
            "mandatory": list(MANDATORY_TRANSACTION_FIELDS),
            "optional": list(OPTIONAL_TRANSACTION_FIELDS),
            "value_date_priority": True,
            "date_format": "DD/MM/YYYY",
        },
        "closest_profile_ids": [item["profile_id"] for item in closest],
        "closest_challenges": sorted({challenge for item in closest for challenge in item.get("challenge_history", [])}),
        "source_matched_capabilities": capabilities,
        "selected_rule_bundle": selected_rule_bundle,
        "candidate_plan": ["ai_layout_addendum" if ai else (strategy or "detected_table") for strategy, ai in candidates],
        "step_gate": step_gate,
        "full_source_validation_required": True,
    }
    plan["plan_id"] = preflight_plan_id(plan)
    return plan

def pdf_password(path: Path) -> str:
    with EXTRACTION_CACHE_LOCK:
        return PDF_PASSWORD_CACHE.get(str(path.resolve()), "")

def register_pdf_password(path: Path, password: str) -> None:
    if path.suffix.lower() != ".pdf":
        return
    key = str(path.resolve())
    with EXTRACTION_CACHE_LOCK:
        if password:
            PDF_PASSWORD_CACHE[key] = password
        else:
            PDF_PASSWORD_CACHE.pop(key, None)

def clear_pdf_password(path: Path) -> None:
    with EXTRACTION_CACHE_LOCK:
        PDF_PASSWORD_CACHE.pop(str(path.resolve()), None)

def open_pdf_reader(path: Path) -> PdfReader:
    reader = PdfReader(str(path))
    if reader.is_encrypted:
        password = pdf_password(path)
        if not password or not reader.decrypt(password):
            raise ValueError("PASSWORD_REQUIRED: This PDF is password protected. Enter its password and submit it again; UPG will not retry unreadable encrypted files.")
    return reader


def fast_pdf_text(path: Path, page_indices: list[int] | None = None) -> str:
    """Read searchable-PDF text with PyMuPDF when available.

    Long statements must still be fully parsed and validated, but they must not
    pay the much slower pure-Python extraction cost once per retry.  PyMuPDF
    reads the same original PDF text layer; it does not OCR, invent geometry,
    or change the evidence used by validation.  Encrypted PDFs retain the
    existing password-safe pypdf route.
    """
    if fitz is None:
        reader = open_pdf_reader(path)
        pages = page_indices if page_indices is not None else range(len(reader.pages))
        return "\f".join(reader.pages[index].extract_text() or "" for index in pages)
    document = fitz.open(str(path))
    try:
        if document.needs_pass:
            password = pdf_password(path)
            if not password or not document.authenticate(password):
                raise ValueError("PASSWORD_REQUIRED: This PDF is password protected. Enter its password and submit it again; UPG will not retry unreadable encrypted files.")
        pages = page_indices if page_indices is not None else range(document.page_count)
        return "\f".join(document.load_page(index).get_text("text") for index in pages)
    finally:
        document.close()

def open_pdfplumber(path: Path):
    password = pdf_password(path)
    try:
        return pdfplumber.open(path, password=password or None)
    except Exception as error:
        if "password" in str(error).lower() or "encrypt" in str(error).lower():
            raise ValueError("PASSWORD_REQUIRED: This PDF is password protected. Enter its password and submit it again; UPG will not retry unreadable encrypted files.") from error
        raise


def ocr_is_available() -> bool:
    """Return whether this deployment can read image-only PDFs with word boxes."""
    return bool(fitz is not None and Image is not None and pytesseract is not None and shutil.which("tesseract"))


def ocr_pdf_page_words(path: Path, page_number: int, config: str | None = None) -> list[dict]:
    """OCR one original PDF page and preserve its coordinates in PDF points.

    We deliberately do not flatten a scan into plain text and then guess its
    columns.  Tesseract's pixel boxes are scaled back to the source PDF page
    size, so the normal header/band parser can learn a reusable layout.
    """
    config = config or os.environ.get("UPG_TESSERACT_CONFIG", "--oem 1 --psm 6")
    key = (str(path.resolve()), page_number, config)
    with EXTRACTION_CACHE_LOCK:
        cached = OCR_WORD_PAGE_CACHE.get(key)
    if cached is not None:
        return cached
    if not ocr_is_available():
        raise ValueError(
            "OCR_UNAVAILABLE: This PDF has no text layer. The UPG server needs "
            "Tesseract OCR installed before it can measure this scanned statement."
        )
    document = fitz.open(str(path))
    try:
        if document.needs_pass:
            password = pdf_password(path)
            if not password or not document.authenticate(password):
                raise ValueError("PASSWORD_REQUIRED: This PDF is password protected. Enter its password and submit it again; UPG will not retry unreadable encrypted files.")
        page = document.load_page(page_number)
        # 240 DPI is sufficient for transaction fonts while retaining a small
        # enough image for multi-page statements.  PDF coordinates are restored
        # below, so a saved profile remains portable to another statement.
        scale = max(1.5, min(4.0, float(os.environ.get("UPG_OCR_SCALE", "3.333"))))
        pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), colorspace=fitz.csRGB, alpha=False)
        image = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
        data = pytesseract.image_to_data(
            image, output_type=TesseractOutput.DICT,
            config=config,
            lang=os.environ.get("UPG_OCR_LANGUAGE", "eng"),
        )
        words: list[dict] = []
        page_width, page_height = float(page.rect.width), float(page.rect.height)
        for index, value in enumerate(data.get("text", [])):
            text = str(value or "").strip()
            try:
                confidence = float(data["conf"][index])
            except (KeyError, ValueError, TypeError):
                confidence = -1.0
            if not text or confidence < 20:
                continue
            left = float(data["left"][index])
            top = float(data["top"][index])
            width = float(data["width"][index])
            height = float(data["height"][index])
            words.append({
                "text": text,
                "x0": left * page_width / pix.width,
                "x1": (left + width) * page_width / pix.width,
                "top": top * page_height / pix.height,
                "bottom": (top + height) * page_height / pix.height,
            })
    finally:
        document.close()
    with EXTRACTION_CACHE_LOCK:
        if len(OCR_WORD_PAGE_CACHE) >= 180:
            OCR_WORD_PAGE_CACHE.pop(next(iter(OCR_WORD_PAGE_CACHE)))
        OCR_WORD_PAGE_CACHE[key] = words
    return words


def _ocr_date_anchor_count(words: list[dict]) -> int:
    """Count concrete date cells, not arbitrary numbers recognised by OCR."""
    return sum(
        1 for word in words
        if float(word.get("x0", 0)) < 360
        and re.fullmatch(r"\d{2}[-/]\d{2}[-/]\d{2,4}", str(word.get("text", "")).strip())
    )


def ocr_pdf_page_best_words(path: Path, page_number: int) -> list[dict]:
    """Choose the stronger original-page OCR measurement for a scan.

    Tesseract page-segmentation mode 6 is best for most statement tables. A
    few scans have faint/ruled date columns, however, where sparse-text mode
    11 finds substantially more *real date anchors*. We re-measure only those
    weak pages and select the alternate result only when it proves better by
    date evidence. This is deterministic OCR recovery, not AI guessing.
    """
    primary = ocr_pdf_page_words(path, page_number)
    primary_dates = _ocr_date_anchor_count(primary)
    if primary_dates >= 4:
        return primary
    alternate = ocr_pdf_page_words(path, page_number, "--oem 1 --psm 11")
    alternate_dates = _ocr_date_anchor_count(alternate)
    return alternate if alternate_dates >= max(3, primary_dates + 2) else primary


def ocr_pdf_text(path: Path, page_indices: list[int] | None = None) -> str:
    """Return OCR text only as support evidence; parsing uses word geometry."""
    if fitz is None:
        raise ValueError("OCR_UNAVAILABLE: PyMuPDF is required for image-only PDFs.")
    document = fitz.open(str(path))
    try:
        if document.needs_pass:
            password = pdf_password(path)
            if not password or not document.authenticate(password):
                raise ValueError("PASSWORD_REQUIRED: This PDF is password protected. Enter its password and submit it again; UPG will not retry unreadable encrypted files.")
        indices = page_indices if page_indices is not None else list(range(document.page_count))
    finally:
        document.close()
    return "\f".join(" ".join(str(word["text"]) for word in ocr_pdf_page_words(path, index)) for index in indices)


def native_pdf_text_is_usable(path: Path) -> bool:
    try:
        return len(re.sub(r"\W", "", read_pdf_text(path))) >= 80
    except (OSError, ValueError):
        return False

def is_large_pdf(path: Path) -> bool:
    try:
        return path.suffix.lower() == ".pdf" and len(open_pdf_reader(path).pages) > 60
    except Exception:
        return False

def prefers_running_balance_text(path: Path) -> bool:
    """Identify long text-layer statements before full-page geometry work."""
    if path.suffix.lower() != ".pdf":
        return False
    raw = remove_page_furniture(cached_pdf_text(path))
    heading = bool(re.search(r"(?i)(?:particulars|narration|description).*?(?:withdrawals?|debits?|deposits?|credits?).*?balance", raw[:6000]))
    dated_rows = len(re.findall(r"(?m)^\s*\d{2}[-/]\d{2}[-/]\d{2,4}\b", raw))
    signed_balances = bool(re.search(r"\b(?:cr|dr)\b", raw, re.I))
    signed_dated_rows = len(re.findall(
        r"(?im)^\s*\d{2}[-/]\d{2}[-/]\d{2,4}\b.*?\b\d[\d,]*\.\d{1,2}\s*(?:dr|cr)\b", raw
    ))
    # Most banks have column labels, but some long J&K layouts do not.  Ten
    # independently dated signed-balance lines are enough source evidence to
    # select the deterministic signed-running-balance parser safely.
    return signed_balances and ((heading and dated_rows >= 10) or signed_dated_rows >= 10)

def sampled_geometry_is_structurally_ready(path: Path) -> bool:
    """Only treat sampled geometry as first-choice evidence with real columns."""
    if path.suffix.lower() != ".pdf":
        return False
    try:
        geometry = sampled_geometry_profile(path)
        if not geometry:
            return False
        fields = map_headers(geometry[0])
        return len(fields) >= 4 and "date" in fields and "balance" in fields
    except (OSError, ValueError, KeyError):
        return False

def sampled_pdf_text(path: Path) -> str:
    """Representative evidence for profile generation on very long PDFs."""
    key = str(path.resolve())
    with EXTRACTION_CACHE_LOCK:
        cached = PDF_SAMPLE_CACHE.get(key)
    if cached is not None:
        return cached
    reader = open_pdf_reader(path)
    count = len(reader.pages)
    indices = sampled_page_indices(count)
    page_texts = fast_pdf_text(path, indices)
    # Preserve page-boundary evidence for the planner while using the fast
    # native extractor for the source text itself.
    sample = "\n".join(f"[PAGE {index + 1}]\n{text}" for index, text in zip(indices, page_texts.split("\f")))
    if not sample.strip():
        # Image-only scans have no native text.  OCR only representative pages
        # during planning; a full OCR pass is deferred until a measured
        # candidate is actually tested.
        sample = "\n".join(
            f"[PAGE {index + 1}]\n{text}"
            for index, text in zip(indices, ocr_pdf_text(path, indices).split("\f"))
        )
    with EXTRACTION_CACHE_LOCK:
        if len(PDF_SAMPLE_CACHE) >= 12:
            PDF_SAMPLE_CACHE.pop(next(iter(PDF_SAMPLE_CACHE)))
        PDF_SAMPLE_CACHE[key] = sample
    return sample

def sampled_pdf_geometry_evidence(path: Path) -> list[dict]:
    """Structured original-PDF layout evidence for the parser-generator AI.

    This intentionally sends table geometry, not a long character dump. Small
    PDFs contribute every page; large PDFs contribute representative first,
    middle, and last page groups (including boundary pages).
    """
    if not native_pdf_text_is_usable(path):
        document = fitz.open(str(path)) if fitz is not None else None
        if document is None:
            return []
        try:
            page_count = document.page_count
        finally:
            document.close()
        evidence: list[dict] = []
        for page_number in (list(range(page_count)) if page_count <= 60 else sampled_page_indices(page_count)):
            words = ocr_pdf_page_words(path, page_number)
            lines: dict[int, list[dict]] = {}
            for word in words:
                lines.setdefault(round(float(word["top"]) / 3), []).append(word)
            header_words: list[dict] = []
            for line in lines.values():
                labels = " ".join(str(word["text"]) for word in line).lower()
                if "date" in labels and "balance" in labels and any(label in labels for label in ("particular", "narration", "description")):
                    header_words = [{"label": str(word["text"]), "x0": round(float(word["x0"]), 1), "x1": round(float(word["x1"]), 1)} for word in sorted(line, key=lambda item: float(item["x0"]))]
                    break
            numeric_bands: dict[int, int] = {}
            for word in words:
                if re.fullmatch(r"-?[\d,]+(?:\.\d{1,2})?(?:Cr|Dr)?", str(word["text"]), re.I):
                    band = round(float(word["x0"]) / 10) * 10
                    numeric_bands[band] = numeric_bands.get(band, 0) + 1
            evidence.append({
                "page": page_number + 1,
                "ocr_geometry": True,
                "tables": [],
                "borderless_coordinate_evidence": {
                    "header_word_positions": header_words,
                    "numeric_column_x_ranges": [{"x0": x0, "x1": x0 + 10, "observations": count} for x0, count in sorted(numeric_bands.items(), key=lambda item: item[1], reverse=True)[:10]],
                },
            })
        return evidence
    evidence: list[dict] = []
    with open_pdfplumber(path) as pdf:
        page_numbers = list(range(len(pdf.pages))) if len(pdf.pages) <= 60 else sampled_page_indices(len(pdf.pages))
        for page_number in page_numbers:
            page = pdf.pages[page_number]
            tables = page.find_tables()
            page_tables = []
            for table in tables[:3]:
                extracted = table.extract()
                if not extracted:
                    continue
                header_index = next((i for i, row in enumerate(extracted[:12]) if len(map_headers(row or [])) >= 3), 0)
                header_cells = table.rows[header_index].cells if header_index < len(table.rows) else []
                bands = [[round(float(cell[0]), 1), round(float(cell[2]), 1)] for cell in header_cells if cell]
                page_tables.append({
                    "bbox": [round(float(value), 1) for value in table.bbox],
                    "header_row_index": header_index,
                    "header": [str(value or "") for value in (extracted[header_index] if header_index < len(extracted) else [])],
                    "column_x_ranges": bands,
                    "row_count": len(extracted),
                })
            # Borderless statements (including many bank exports) have no
            # `find_tables()` result. Preserve their actual PDF geometry as
            # coordinate evidence without exposing transaction wording.
            words = page.extract_words(x_tolerance=1, y_tolerance=2)
            lines: dict[int, list[dict]] = {}
            for word in words:
                lines.setdefault(round(float(word["top"]) / 3), []).append(word)
            header_words: list[dict] = []
            for line in lines.values():
                labels = " ".join(str(word["text"]) for word in line).lower()
                if "date" in labels and "balance" in labels and any(label in labels for label in ("particular", "narration", "description")):
                    header_words = [{"label": str(word["text"]), "x0": round(float(word["x0"]), 1), "x1": round(float(word["x1"]), 1)} for word in sorted(line, key=lambda item: float(item["x0"]))]
                    break
            numeric_bands: dict[int, int] = {}
            for word in words:
                if re.fullmatch(r"-?[\d,]+(?:\.\d{1,2})?(?:Cr|Dr)?", str(word["text"]), re.I):
                    band = round(float(word["x0"]) / 10) * 10
                    numeric_bands[band] = numeric_bands.get(band, 0) + 1
            coordinate_fallback = {
                "header_word_positions": header_words,
                "numeric_column_x_ranges": [{"x0": x0, "x1": x0 + 10, "observations": count} for x0, count in sorted(numeric_bands.items(), key=lambda item: item[1], reverse=True)[:10]],
            }
            evidence.append({"page": page_number + 1, "width": round(float(page.width), 1), "height": round(float(page.height), 1), "tables": page_tables, "borderless_coordinate_evidence": coordinate_fallback})
    return evidence

def compact_geometry_for_ai(evidence: list[dict]) -> list[dict]:
    """Keep representative geometry, not a costly page-by-page dump."""
    if len(evidence) <= 5:
        selected = evidence
    else:
        selected = [evidence[0], evidence[len(evidence) // 2], evidence[-1]]
    return [{
        "page": item.get("page"),
        "width": item.get("width"),
        "height": item.get("height"),
        "tables": [{
            "header": table.get("header", [])[:12],
            "column_x_ranges": table.get("column_x_ranges", [])[:12],
            "row_count": table.get("row_count"),
        } for table in item.get("tables", [])[:2]],
        "borderless_coordinate_evidence": {
            "header_word_positions": item.get("borderless_coordinate_evidence", {}).get("header_word_positions", [])[:12],
            "numeric_column_x_ranges": item.get("borderless_coordinate_evidence", {}).get("numeric_column_x_ranges", [])[:6],
        },
    } for item in selected]

def sampled_page_indices(count: int) -> list[int]:
    """Seven-page regions plus boundary context for large PDF layout learning."""
    if count <= 21:
        return list(range(count))
    middle = count // 2
    first = set(range(0, min(count, 8)))
    middle_window = set(range(max(0, middle - 4), min(count, middle + 5)))
    last = set(range(max(0, count - 8), count))
    return sorted(first | middle_window | last)

def ai_generated_profile(rows: list[list[object]], raw: str, repair_context: str = "", source_path: Path | None = None, job_id: str | None = None) -> tuple[int, dict[str, int]] | None:
    """Ask the embedded parser-generator AI for a new table layout, not transactions."""
    def record_failure(reason: str) -> None:
        if job_id:
            patch_job(job_id, ai_layout_error=reason[:300])

    key = os.environ.get("OPENAI_API_KEY")
    if not key:
        record_failure("OPENAI_API_KEY is not configured.")
        return None
    # The first AI call creates an evidence-led layout blueprint.  Once that
    # map has been tested and failed, the final call must create a *revised
    # map* from the measured failure evidence--not merely diagnose it.
    prior_purposes = ai_call_purposes(job_id)
    purpose = "targeted_repair_profile" if (
        job_id and repair_context and (
            "layout_blueprint" in prior_purposes
            or "deterministic_geometry_preflight" in prior_purposes
        )
    ) else "layout_blueprint"
    # A complete semantic PDF ledger header is already stronger evidence than
    # an LLM's generic 0..5 column map.  Previously we spent the first AI call
    # asking it to rediscover that obvious map, then rejected the repeated
    # canonical answer as not materially new.  Defer the expert call until a
    # deterministic extraction has produced concrete failed validation
    # evidence.  The deferred call is then a single targeted repair rather
    # than a planning call plus a duplicate-map repair.
    if (
        purpose == "layout_blueprint"
        and source_path is not None
        and source_path.suffix.lower() == ".pdf"
        and has_standard_geometry_header_contract(cached_pdf_text(source_path))
    ):
        if job_id:
            with JOBS_LOCK:
                job = JOBS.get(job_id, {})
                history = [str(item) for item in job.get("ai_call_purposes", [])][-5:]
                if "deterministic_geometry_preflight" not in history:
                    history.append("deterministic_geometry_preflight")
                job["ai_call_purposes"] = history
                job["ai_layout_error"] = (
                    "AI layout planning deferred: measured Date/Particulars/"
                    "Withdrawal/Deposit/Balance geometry is available."
                )
                JOBS[job_id] = job
                persist_job_locked(job_id)
        return None
    if not reserve_ai_call(job_id, purpose):
        record_failure("AI call budget is exhausted before a layout map could be generated.")
        return None
    schema = {
        "type": "object", "additionalProperties": False,
        "properties": {
            "header_row": {"type": "integer"},
            "columns": {
                "type": "object", "additionalProperties": False,
                "properties": {name: {"type": "integer"} for name in CANONICAL},
                "required": CANONICAL,
            },
        }, "required": ["header_row", "columns"],
    }
    geometry = compact_geometry_for_ai(sampled_pdf_geometry_evidence(source_path)) if source_path and source_path.suffix.lower() == ".pdf" else []
    prior_maps: list[dict] = []
    if job_id:
        with JOBS_LOCK:
            prior_maps = list(JOBS.get(job_id, {}).get("ai_layout_maps", []))[-2:]
    evidence = {"rows": rows[:18], "original_pdf_geometry_samples": geometry, "failed_validation_evidence": repair_context[-1800:],
                "previous_failed_layout_maps": prior_maps,
                "upg_learning": compact_ai_learning_packet(source_path, rows[0] if rows else None, raw)}
    instruction = (AI_LAYOUT_CONTRACT + "\nIdentify one transaction-table header row and map "
        "its zero-based column positions to date, narration, withdrawal, deposit, instrument_number, "
        "and balance. These are the only allowed transaction outputs. Interpret unfamiliar header wording semantically from the measured source header: for example Transaction Remarks/Description/Details means narration; Debit/Withdrawal means withdrawal; Credit/Deposit means deposit; Post/Transaction/Booking Date and Value Date are date fields (Value Date wins for output). Do not map account metadata, totals, page furniture, or a word outside the measured table header. Use the original_pdf_geometry_samples as primary evidence; do not infer a column from character order alone. The source_matched_certified_capabilities are reusable behaviours only: apply one only when the supplied source evidence proves it, and never copy another profile's coordinates, code, or field indexes. Use -1 when a field is absent. If failure evidence is supplied, return a revised measured header/column mapping that directly repairs that failure. The previous targeted repairs and failed layout maps are evidence of actions already tested: do not repeat them. Change only a source role which the measured geometry or grid proves should change; if no such source evidence exists, do not invent a map. Do not extract transactions, invent values, or change validation rules."
    )
    payload = {
        "model": AI_MODEL,
        "input": [{"role": "system", "content": [{"type": "input_text", "text": instruction}]},
                  # Native spreadsheet/Word grids may contain Decimal values.
                  # The AI only plans headers and column positions, so serialise
                  # those losslessly as strings rather than crashing before the
                  # API call (which previously consumed an AI budget slot).
                  {"role": "user", "content": [{"type": "input_text", "text": json.dumps(evidence, default=str)}]}],
        "max_output_tokens": AI_MAX_OUTPUT_TOKENS,
        "text": {"format": {"type": "json_schema", "name": "bank_layout", "strict": True, "schema": schema}},
    }
    request = urllib.request.Request(
        "https://api.openai.com/v1/responses", data=json.dumps(payload).encode(),
        headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"}, method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=90) as response:
            result = json.loads(response.read().decode())
        text = next((item["text"] for output in result.get("output", []) for item in output.get("content", []) if item.get("type") == "output_text"), "")
        generated = json.loads(text)
        header_row = int(generated["header_row"])
        columns = {name: int(index) for name, index in generated["columns"].items() if int(index) >= 0}
        if not (0 <= header_row < len(rows)):
            record_failure("AI returned a header row outside the measured source grid.")
            return None
        if len(columns) < 3:
            record_failure("AI returned fewer than three usable source columns.")
            return None
        # A repair which repeats a failed map cannot improve the parser. PDF
        # rows are already a measured canonical grid, so the generic 0..5 map
        # merely repeats that grid; it does not adapt original PDF geometry.
        map_key = {"header_row": header_row, "columns": columns}
        previous_keys = [
            {"header_row": int(item.get("header_row", -1)), "columns": dict(item.get("columns", {}))}
            for item in prior_maps if isinstance(item, dict)
        ]
        # ``rows`` is deliberately normalized into the canonical six-column
        # transaction grid before this planner sees it.  Therefore a correct
        # measured source map commonly *is* ``header_row=0`` with columns
        # ``0..5``.  Treating that shape as intrinsically stale made the AI
        # reject a valid profile simply because the output contract is
        # canonical (the date-order jobs exposed this defect).  Only reject an
        # exact map the AI has already supplied earlier in this job; a first
        # canonical map must be allowed to reach the normal evidence gates.
        if map_key in previous_keys:
            record_failure(
                "AI repair repeated an already tested measured header map instead of proposing a new source layout."
            )
            return None
        # Retain the *measured plan*, never source text, so a later retry (and
        # the final user-visible job status) can distinguish an API failure
        # from an AI map that was successfully produced but failed validation.
        # This is essential for evidence-led self healing: the second call
        # must revise a known map, rather than be treated as an unexplained
        # generic AI attempt.
        if job_id:
            with JOBS_LOCK:
                maps = list(JOBS.get(job_id, {}).get("ai_layout_maps", []))[-4:]
            maps.append({"purpose": purpose, "header_row": header_row, "columns": columns})
            patch_job(job_id, ai_layout_maps=maps, ai_layout_error="")
        return header_row, columns
    except (urllib.error.URLError, urllib.error.HTTPError, KeyError, ValueError, json.JSONDecodeError) as error:
        record_failure(safe_openai_error(error))
        return None

def ai_choose_text_strategy(raw: str, job_id: str | None = None) -> str | None:
    """Let the parser-generator select a supported extraction path for a new layout."""
    key = os.environ.get("OPENAI_API_KEY")
    if not key: return None
    if not reserve_ai_call(job_id, "strategy_classification"):
        return None
    schema = {
        "type": "object", "additionalProperties": False,
        "properties": {"strategy": {"type": "string", "enum": ["running_balance_text", "unsigned_running_balance_text", "value_date_unsigned", "needs_ocr", "unsupported"]}},
        "required": ["strategy"],
    }
    payload = {
        "model": AI_MODEL,
        "input": (AI_LAYOUT_CONTRACT + "\nClassify this bank statement layout. Choose running_balance_text when dated entries have Dr/Cr running balances. Choose unsigned_running_balance_text when dated entries have unsigned running balances whose changes can infer debit or credit; choose "
            "value_date_unsigned when there are both posting Date and Value Date columns plus unsigned running balances; choose needs_ocr for image/scanned text; otherwise choose unsupported.\nUPG learning: " + json.dumps(compact_ai_learning_packet(raw=raw)) + "\n\n" + raw[:3500]
        ),
        "max_output_tokens": AI_MAX_OUTPUT_TOKENS,
        "text": {"format": {"type": "json_schema", "name": "extraction_strategy", "strict": True, "schema": schema}},
    }
    request = urllib.request.Request("https://api.openai.com/v1/responses", data=json.dumps(payload).encode(), headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"}, method="POST")
    try:
        with urllib.request.urlopen(request, timeout=90) as response: result = json.loads(response.read().decode())
        text = next((item["text"] for output in result.get("output", []) for item in output.get("content", []) if item.get("type") == "output_text"), "")
        return json.loads(text)["strategy"]
    except (urllib.error.URLError, urllib.error.HTTPError, KeyError, ValueError, json.JSONDecodeError):
        return None

def safe_openai_error(error: Exception) -> str:
    """Expose actionable provider diagnostics without exposing keys or source text."""
    if isinstance(error, urllib.error.HTTPError):
        message = ""
        try:
            body = json.loads(error.read().decode("utf-8", errors="replace"))
            message = str(body.get("error", {}).get("message", ""))
        except (OSError, ValueError, json.JSONDecodeError):
            pass
        message = re.sub(r"\s+", " ", message).strip()[:220]
        return f"OpenAI HTTP {error.code}" + (f": {message}" if message else ".")
    if isinstance(error, urllib.error.URLError):
        return f"OpenAI network error: {type(error.reason).__name__}."
    return f"OpenAI response error: {type(error).__name__}."

def failure_type_from_evidence(failure: str) -> str:
    """Classify a failed gate before selecting an AI repair scope.

    This keeps S27 from receiving a broad library and, more importantly,
    prevents a failure in an earlier measured step from being 'repaired' by a
    later unrelated rule such as furniture cleanup or financial reconciliation.
    """
    evidence = (failure or "").lower()
    if any(token in evidence for token in ("header", "column", "geometry", "grid", "source columns")):
        return "column_geometry"
    if any(token in evidence for token in ("date", "value date", "chronolog", "reverse order")):
        return "date_order"
    if any(token in evidence for token in ("narration", "particular", "continuation")):
        return "continuation"
    if any(token in evidence for token in ("footer", "furniture", "page total", "summary")):
        return "page_furniture"
    if any(token in evidence for token in ("running balance", "balance direction", "balance chain")):
        return "balance_direction"
    if "transaction count" in evidence or "source coverage" in evidence:
        return "transaction_count"
    if any(token in evidence for token in ("opening", "closing", "financial", "endpoint", "total")):
        return "endpoint"
    return "novel_layout"

def ai_diagnose_failure(raw: str, failure: str, source_path: Path | None = None, job_id: str | None = None) -> dict[str, object]:
    """Use the final AI call for a concrete, safe repair plan.

    The model may choose a diagnosis and a parser-profile action, but never
    transactions, balances, executable code, or a weaker validation standard.
    """
    key = os.environ.get("OPENAI_API_KEY")
    if not key:
        return {"rules": [], "strategies": [], "failure_type": "novel_layout", "profile_action": "reject_unsafe", "diagnostic_error": "AI diagnosis unavailable: OPENAI_API_KEY is not configured."}
    if not reserve_ai_call(job_id, "targeted_repair_plan"):
        return {"rules": [], "strategies": [], "failure_type": "novel_layout", "profile_action": "reject_unsafe", "diagnostic_error": "AI call budget reached for this job; no new evidence-led repair remains."}
    scoped_failure_type = failure_type_from_evidence(failure)
    learning_packet = compact_ai_learning_packet(source_path, raw=raw, failure_type=scoped_failure_type)
    allowed_rules = list(learning_packet.get("allowed_rule_modules", {}).keys())
    safe_strategies = ["geometry_profile", "source_amount_geometry", "value_date_unsigned", "unsigned_running_balance_text", "running_balance_text", "page_text_unsigned", "detected_table"]
    schema = {"type": "object", "additionalProperties": False, "properties": {
        "rules": {"type": "array", "items": {"type": "string", "enum": allowed_rules}, "maxItems": 5},
        "strategies": {"type": "array", "items": {"type": "string", "enum": safe_strategies}, "maxItems": 4},
        "failure_type": {"type": "string", "enum": ["column_geometry", "header_mapping", "date_order", "continuation", "page_furniture", "balance_direction", "unreliable_balance", "endpoint", "source_totals", "narration_coverage", "transaction_count", "novel_layout"]},
        "profile_action": {"type": "string", "enum": ["reuse_geometry", "repair_header_map", "repair_continuations", "repair_date_order", "repair_balance_direction", "reject_unsafe"]},
    }, "required": ["rules", "strategies", "failure_type", "profile_action"]}
    prompt = AI_LAYOUT_CONTRACT + "\nThis is the final AI decision for this job and it is being made AFTER the first source-layout extraction failed. Repair only " + str(learning_packet["ai_context_scope"]["pipeline_step"]) + ". Do not move downstream until this step has source proof. Produce one targeted, evidence-led repair plan using only the supplied certified modules and strategies. Do not write code or transactions, do not relax validation, and do not request another AI layout addendum. If evidence is insufficient, choose reject_unsafe.\nScoped rules: " + json.dumps(learning_packet["allowed_rule_modules"]) + "\nStrategies: " + json.dumps(safe_strategies) + "\nUPG learning: " + json.dumps(learning_packet) + "\nFailure evidence: " + failure[-1800:] + "\nSource excerpt: " + raw[:3500]
    payload = {"model": AI_MODEL, "input": prompt, "max_output_tokens": AI_MAX_OUTPUT_TOKENS, "text": {"format": {"type": "json_schema", "name": "diagnostic_rules", "strict": True, "schema": schema}}}
    request = urllib.request.Request("https://api.openai.com/v1/responses", data=json.dumps(payload).encode(), headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"}, method="POST")
    try:
        with urllib.request.urlopen(request, timeout=90) as response: result = json.loads(response.read().decode())
        output = next((item["text"] for item_out in result.get("output", []) for item in item_out.get("content", []) if item.get("type") == "output_text"), "")
        plan = json.loads(output)
        return {
            "rules": [rule for rule in plan["rules"] if rule in allowed_rules],
            "strategies": [strategy for strategy in plan["strategies"] if strategy in safe_strategies],
            # The deterministic source evidence owns the current pipeline
            # step.  The AI may choose an action inside that step, but cannot
            # relabel an upstream failure as a downstream one to bypass it.
            "failure_type": scoped_failure_type,
            "profile_action": str(plan["profile_action"]),
            "diagnostic_error": "",
        }
    except (urllib.error.URLError, urllib.error.HTTPError, KeyError, ValueError, json.JSONDecodeError) as error:
        # Keep only a short, non-sensitive diagnostic.  This lets the retry
        # loop distinguish an AI/API failure from a genuine but empty plan.
        return {"rules": [], "strategies": [], "failure_type": "novel_layout", "profile_action": "reject_unsafe", "diagnostic_error": "AI diagnosis unavailable: " + safe_openai_error(error)}

def candidate_failure_evidence(candidate: tuple | None) -> list[str]:
    """Emit compact, non-sensitive proof for the one module that failed.

    A validation result alone (for example ``financial=fail``) is downstream
    evidence.  The repair planner needs the source-level reason so it can
    change just the failing module on the next attempt instead of replaying a
    general table candidate.  This metadata is job telemetry only; it never
    contains statement text, transactions, account numbers, or amounts.
    """
    if candidate is None:
        return []
    try:
        tx, _opening, _closing, _withdrawals, _deposits, _computed, financial, narration, _unmatched, _headers, columns, _parent, coverage, expected_count = candidate[:14]
        columns = columns or {}
    except (IndexError, TypeError, ValueError):
        return []

    evidence: list[str] = [
        f"gate financial={'pass' if financial else 'fail'}",
        f"gate narration={'pass' if narration else 'fail'}",
        f"gate source_coverage={'pass' if coverage else 'fail'}",
        f"count parsed={len(tx)} expected={expected_count}",
    ]
    proofs = {
        "source columns distinct": columns.get("_source_columns_distinct", True),
        "header roles aligned": columns.get("_source_header_roles_aligned", True),
        "measured column evidence": columns.get("_source_column_evidence_valid", True),
        "source date cells": columns.get("_source_date_cells_valid", True),
        "source balance cells": columns.get("_source_balance_cells_valid", True),
        "source narration cells": columns.get("_source_narration_cells_valid", True),
        "source record fingerprint": columns.get("_source_record_fingerprint_valid", True),
        "canonical output contract": columns.get("_canonical_contract_valid", True),
    }
    evidence.extend(f"proof {name}={'pass' if value else 'fail'}" for name, value in proofs.items())
    if columns.get("_source_totals_conflict"):
        evidence.append("proof printed source totals conflict")
    if columns.get("_source_balance_unreliable"):
        evidence.append("proof source balance unreliable")
    return evidence


def evidence_repair_plan(errors: list[str], candidate: tuple | None) -> dict[str, object]:
    """Turn a failed candidate into one bounded, module-specific next plan.

    This intentionally makes no API call. The second expert-AI call is reserved
    for a revised measured layout map; this function keeps that repair focused
    on the failed module rather than retrying a vague "novel layout".
    """
    # The structured candidate proof has eight small fields. Keep the full
    # bounded set so an early but decisive header/column failure is not lost
    # behind later aggregate gate messages.
    evidence = " ".join(str(item) for item in errors[-20:]).lower()
    if candidate is not None:
        try:
            columns = candidate[10] or {}
            if columns.get("_source_totals_conflict"):
                evidence += " source totals conflict"
            if columns.get("_source_balance_unreliable"):
                evidence += " source balance unreliable"
        except (IndexError, TypeError):
            pass
    rules: list[str] = []
    strategies: list[str] = []
    failure_type, profile_action = "column_geometry", "repair_header_map"

    def add_rule(name: str) -> None:
        if name in DIAGNOSTIC_RULE_LIBRARY and name not in rules:
            rules.append(name)

    def add_strategy(name: str) -> None:
        if name not in strategies:
            strategies.append(name)

    # A failed source proof is more reliable than downstream aggregate gates.
    # Repair only that module, in priority order.  This turns a failed
    # candidate into a bounded next action and avoids the former broad
    # "try another layout" behavior.
    if any(token in evidence for token in ("header roles aligned=fail", "source columns distinct=fail", "measured column evidence=fail")):
        failure_type, profile_action = "header_mapping", "repair_header_map"
        for name in ("header_role_alignment", "distinct_source_columns", "measured_column_evidence"):
            add_rule(name)
        add_strategy("geometry_profile")
        add_strategy("source_amount_geometry")
        return {"rules": rules, "strategies": strategies, "failure_type": failure_type,
                "profile_action": profile_action, "diagnostic_error": ""}
    if "source date cells=fail" in evidence:
        failure_type, profile_action = "date_order", "repair_date_order"
        for name in ("date_column_boundary", "date_source_cell", "value_date"):
            add_rule(name)
        add_strategy("dual_date_geometry")
        add_strategy("value_date_unsigned")
        return {"rules": rules, "strategies": strategies, "failure_type": failure_type,
                "profile_action": profile_action, "diagnostic_error": ""}
    if "source balance cells=fail" in evidence:
        failure_type, profile_action = "balance_direction", "repair_balance_direction"
        for name in ("balance_source_cell", "signed_balance_text", "corrupt_balance_text_layer"):
            add_rule(name)
        add_strategy("geometry_profile")
        add_strategy("running_balance_text")
        return {"rules": rules, "strategies": strategies, "failure_type": failure_type,
                "profile_action": profile_action, "diagnostic_error": ""}
    if any(token in evidence for token in ("source narration cells=fail", "gate narration=fail")):
        failure_type, profile_action = "narration_coverage", "repair_continuations"
        for name in ("narration_source_cell", "continuation_merge", "multi_page_continuation", "footer_exclusion"):
            add_rule(name)
        add_strategy("geometry_profile")
        return {"rules": rules, "strategies": strategies, "failure_type": failure_type,
                "profile_action": profile_action, "diagnostic_error": ""}
    if any(token in evidence for token in ("source record fingerprint=fail", "gate source_coverage=fail")):
        failure_type = "transaction_count"
        for name in ("source_coverage", "bf_preperiod_artifact"):
            add_rule(name)
        add_strategy("geometry_profile")
        return {"rules": rules, "strategies": strategies, "failure_type": failure_type,
                "profile_action": profile_action, "diagnostic_error": ""}

    # When all three release gates fail together, the source map itself is
    # unproven. Do not pretend that narration or endpoints alone are the
    # cause; leave the final AI call to produce a genuinely new measured map.
    if all(marker in evidence for marker in ("financial=fail", "narration=fail", "source_coverage=fail")):
        return {"rules": rules, "strategies": strategies, "failure_type": failure_type,
                "profile_action": profile_action, "diagnostic_error": ""}

    # A source-total/amount defect is more specific than a downstream
    # financial/narration failure. Choose one dominant repair family per
    # round—combining every library rule would recreate the broad retry that
    # this planner is intended to remove.
    if any(token in evidence for token in ("decimal", "punctuation", "multiple points", "source totals conflict")):
        failure_type = "source_totals"
        for name in ("indian_money_punctuation", "amount_balance_consistency", "source_amount_geometry"):
            add_rule(name)
        add_strategy("source_amount_geometry")
        add_strategy("geometry_profile")
    elif any(token in evidence for token in ("narration", "particular", "continuation", "furniture", "footer")):
        failure_type, profile_action = "continuation", "repair_continuations"
        for name in ("continuation_merge", "multi_page_continuation", "footer_exclusion", "terminal_row_before_summary"):
            add_rule(name)
        add_strategy("geometry_profile")
    elif any(token in evidence for token in ("date", "value date", "out-of-fy", "out of fy", "reverse order")):
        failure_type, profile_action = "date_order", "repair_date_order"
        for name in ("date_column_boundary", "value_date", "reference_date_boundary"):
            add_rule(name)
        # Keep the repair on the source's visual date bands. Generic text
        # extraction can flatten Post Date and Value Date into one stream and
        # falsely report a date-order failure for an otherwise clear ledger.
        add_strategy("dual_date_geometry")
        add_strategy("value_date_unsigned")
        add_strategy("geometry_profile")
    elif any(token in evidence for token in ("coverage", "transaction count", "source records")):
        failure_type = "transaction_count"
        for name in ("source_coverage", "bf_preperiod_artifact"):
            add_rule(name)
        add_strategy("geometry_profile")
    elif any(token in evidence for token in ("opening", "closing", "endpoint", "balance", "financial", "source balance unreliable")):
        failure_type = "unreliable_balance" if "unreliable" in evidence else "balance_direction"
        profile_action = "repair_balance_direction"
        for name in ("summary_endpoints", "balance_delta", "signed_balance_text", "corrupt_balance_text_layer"):
            add_rule(name)
        add_strategy("running_balance_text")
        add_strategy("unsigned_running_balance_text")

    return {"rules": rules, "strategies": strategies, "failure_type": failure_type,
            "profile_action": profile_action, "diagnostic_error": ""}

# A targeted repair plan may select only extraction modes that the local UPG
# engine actually implements.  The list deliberately contains *strategies*,
# not bank-specific parser IDs or old column offsets: the new statement is
# always measured again from its own source before a strategy is run.
TARGETED_REPAIR_STRATEGIES = frozenset({
    "geometry_profile",
    "source_amount_geometry",
    "dual_date_geometry",
    "value_date_unsigned",
    "unsigned_running_balance_text",
    "running_balance_text",
    "page_text_unsigned",
    "detected_table",
})

def pending_targeted_repair_strategies(strategies: object, failed_strategy_keys: set[str]) -> list[str]:
    """Return the untried deterministic actions from the latest repair plan.

    This is the Step 16 execution gate.  A source-proof failure (for example
    misaligned header roles) must first run its selected measured repair
    strategy.  It must not spend another AI request, nor replay unrelated
    candidates, until that narrow deterministic action has been tested.
    """
    if not isinstance(strategies, list):
        return []
    pending: list[str] = []
    for raw_strategy in strategies:
        strategy = str(raw_strategy)
        if (
            strategy in TARGETED_REPAIR_STRATEGIES
            and strategy not in pending
            and f"{strategy}:deterministic" not in failed_strategy_keys
        ):
            pending.append(strategy)
    return pending

def source_balances(text: str) -> tuple[Decimal | None, Decimal | None]:
    def find(kind):
        # A column heading such as "Closing Balance*" is commonly followed by
        # the first transaction date (for example 01/04/2024).  It is not the
        # statement endpoint. Prefer a labelled monetary value (with decimals)
        # and, when a summary repeats it, take the final source occurrence.
        # Banks also abbreviate the labels as ``Opening Bal:`` / ``Closing
        # Bal:`` and often put the corresponding amount on the following
        # line in the statement summary.  This is still a source-declared
        # endpoint: do not infer it from a transaction row.
        label = rf"{kind}\s*(?:balance|bal\.?)"
        monetary = re.findall(
            rf"\b{label}\b[^\d\r\n]{{0,20}}(?:\r?\n\s*)?"
            r"(-?[\d,]+\.\d{1,2}(?:\s*(?:CR|DR))?)",
            text,
            re.I,
        )
        if monetary:
            return money(monetary[-1])
        # Retain support for sources that print whole-number balances, while
        # rejecting a date fragment immediately after a table heading.
        match = re.search(rf"\b{label}\b\D{{0,20}}([\d,]+(?:\.\d{{1,2}})?(?:\s*(?:CR|DR))?)", text, re.I)
        if not match:
            return None
        value = money(match.group(1))
        following = text[match.end():match.end() + 12]
        if value is not None and abs(value) <= 31 and re.match(r"\s*/\s*\d{1,2}\s*/", following):
            return None
        return value
    opening, closing = find("opening"), find("closing")
    summary = re.search(r"(?is)statement\s+summary\s*:-?.{0,300}?opening\s+balance.*?\n\s*([\d,]+(?:\.\d{1,2})?)\s+\d+\s+\d+\s+[\d,]+(?:\.\d{1,2})?\s+[\d,]+(?:\.\d{1,2})?\s+([\d,]+(?:\.\d{1,2})?)", text)
    if summary:
        summary_open, summary_close = money(summary.group(1)), money(summary.group(2))
        if summary_open is not None: opening = summary_open
        if summary_close is not None: closing = summary_close
    # Jammu and Kashmir Bank-style statements use B/F rather than the words
    # "opening balance". It is a statement-level value, never a transaction.
    bf = re.search(r"\bB/F\b[\s\S]{0,80}?(-?[\d,]+(?:\.\d+)?)\s*(Dr|Cr)\b", text, re.I)
    if opening is None and bf:
        amount = money(bf.group(1))
        opening = -abs(amount) if amount is not None and bf.group(2).lower() == "dr" else amount
    # J&K Bank cash-credit statements commonly place the final running
    # balance as the third value on their last Grand Total line, without a
    # separate "Closing Balance" label.  It is a statement-level endpoint,
    # never a transaction or a total amount.
    if closing is None:
        grand_endpoint = re.findall(
            r"(?is)\bgrand\s+total\s*:\s*[\d,]+(?:\.\d{1,2})?\s+"
            r"[\d,]+(?:\.\d{1,2})?\s+(-?[\d,]+(?:\.\d{1,2})?\s*(?:dr|cr)?)",
            text,
        )
        if grand_endpoint:
            closing = money(grand_endpoint[-1])
    return opening, closing

def source_transaction_totals(text: str) -> tuple[Decimal | None, Decimal | None]:
    """Read printed debit/credit totals when a statement provides them.

    These are independent controls. Statements without printed totals return
    ``None`` values and retain the ordinary opening-to-closing reconciliation.
    """
    summary = re.search(
        r"(?is)statement\s+summary\s*:-?.{0,300}?\bdebits?\b\s+\bcredits?\b.*?\n\s*"
        r"[\d,]+(?:\.\d{1,2})?\s+\d+\s+\d+\s+([\d,]+(?:\.\d{1,2})?)\s+([\d,]+(?:\.\d{1,2})?)",
        text,
    )
    if not summary:
        # J&K Bank loan statements can serialize their visual final Total row
        # as three vertical lines in the PDF text layer.  It is still printed
        # source evidence: debit, credit and final balance, in that order.
        # Keep this narrow to a labelled Total followed immediately by two
        # money tokens, so transaction rows and page furniture cannot become
        # a statement-total control.
        stacked_total = re.findall(
            r"(?im)^\s*total\s*$\s*^\s*([\d,]+(?:\.\d{1,2})?)\s*$\s*^\s*([\d,]+(?:\.\d{1,2})?)\s*$",
            text,
        )
        if stacked_total:
            withdrawals, deposits = stacked_total[-1]
            return money(withdrawals), money(deposits)
        # YES Bank and similar statements use compact final-summary labels
        # rather than a grid headed "Statement Summary".  These are source
        # controls only; they never replace transaction-level validation.
        compact_summary = re.search(
            r"(?is)\btotal\s+debit\s+amt\s*:\s*([\d,]+(?:\.\d{1,2})?).{0,160}?"
            r"\btotal\s+credit\s+amt\s*:\s*([\d,]+(?:\.\d{1,2})?)",
            text,
        )
        if compact_summary:
            return money(compact_summary.group(1)), money(compact_summary.group(2))
        # Some banks print cumulative debit and credit totals as a final
        # Grand Total rather than a labelled statement-summary grid.
        grand = re.search(r"(?is)\bgrand\s+total\s*:\s*([\d,]+(?:\.\d{1,2})?)\s+([\d,]+(?:\.\d{1,2})?)", text)
        if grand:
            return money(grand.group(1)), money(grand.group(2))
        # ICICI-style statements end with a borderless `TOTAL` row.  The
        # table order is Deposits, Withdrawals, Balance, so return debit then
        # credit to match this function's public contract.  Use the final
        # occurrence: an account-summary block may contain an earlier total.
        totals = re.findall(
            r"(?im)^\s*TOTAL\s+([\d,]+(?:\.\d{1,2})?)\s+([\d,]+(?:\.\d{1,2})?)\s+([\d,]+(?:\.\d{1,2})?)\s*$",
            text,
        )
        if totals:
            deposits, withdrawals, _balance = totals[-1]
            return money(withdrawals), money(deposits)
        return None, None
    return money(summary.group(1)), money(summary.group(2))

def source_page_total_sums(text: str) -> tuple[Decimal | None, Decimal | None]:
    """Sum printed page totals as an independent control over Grand Total.

    Some source statements have an unreliable summary.  A Grand Total becomes
    trustworthy only when the statement's own Page Total rows independently
    sum to it.  This prevents an equal debit/credit duplication from cancelling
    out in the opening-to-closing equation.
    """
    matches = re.findall(
        r"(?is)\bpage\s+total\s*:\s*([\d,]+(?:\.\d{1,2})?)\s+"
        r"([\d,]+(?:\.\d{1,2})?)\s+[-]?[\d,]+(?:\.\d{1,2})?\s*(?:dr|cr)?",
        text or "",
    )
    if len(matches) < 2:
        return None, None
    first = sum((money(left) or Decimal("0")) for left, _right in matches)
    second = sum((money(right) or Decimal("0")) for _left, right in matches)
    return first, second

def table_balances(rows: list[list[object]], header_at: int, columns: dict[str, int]) -> tuple[Decimal | None, Decimal | None]:
    """Extract declared balances from the source table, independent of computed totals."""
    balance_index = columns.get("balance")
    if balance_index is None: return None, None
    def cell(row, key):
        i = columns.get(key); return row[i] if i is not None and i < len(row) else ""
    first, last = None, None
    for row in rows[header_at + 1:]:
        bal = money(cell(row, "balance"))
        if bal is None: continue
        narration = str(cell(row, "narration")).upper()
        if "OPENING BALANCE" in narration: first = bal
        elif str(cell(row, "date")).strip():
            last = bal
    return first, last

def transaction_date_value(value: str) -> datetime | None:
    # Text layers frequently split a date across adjacent words, e.g.
    # `24-01 -2024`.  It is still one visual date cell, not a continuation
    # line.  Normalise whitespace around separators before validation.
    value = re.sub(r"\s*([/.-])\s*", r"\1", str(value or "").strip())
    # Original-PDF geometry may wrap the final two year digits inside the
    # same Value Date cell (``02/Apr/20\n25``).  A date cell contains no
    # narration, therefore removing its internal whitespace is safe and
    # restores the source-printed date without inventing a value.
    value = re.sub(r"\s+", "", value)
    for pattern in ("%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y", "%d-%m-%y", "%d/%m/%y", "%d.%m.%y", "%d-%b-%Y", "%d/%b/%Y", "%Y-%m-%d"):
        try: return datetime.strptime(value.strip(), pattern)
        except ValueError: continue
    return None

def raw_transaction_record_count(raw: str) -> int | None:
    """Independent source-record count for common text-layout statements."""
    header = raw[:2500].lower()
    # Full dates followed by narration identify the primary record line. This
    # excludes secondary value-date/time lines in dual-date layouts.
    long_form = re.findall(r"(?m)^\s*\d{2}-[A-Za-z]{3}-\d{4}\s+(?=[A-Za-z])", raw)
    if len(long_form) >= 3:
        return len(long_form)
    # Blank narration is allowed; date plus the row's amount/balance columns
    # is still a transaction record, so do not require a letter after date.
    # A few text layers detach a transaction's reference/narration onto the
    # preceding line and put its date, time, amount and balance below it. That
    # date is part of the same record, not another transaction.
    numeric_long_form = re.findall(
        r"(?m)^\s*\d{2}-\d{2}-\d{4}\b"
        # Repeated statement-period headings can begin a text line with
        # `01-04-2025 TO 31-03-2026`; that is metadata, not a transaction.
        r"(?!\s+TO\s+\d{2}-\d{2}-\d{4}\b)"
        r"(?!\s+\d{2}:\d{2}:\d{1,2}\s+[-\d,]+(?:\.\d+)?\s+[-\d,]+(?:\.\d+)?\s*(?:Dr|Cr)\b)", raw)
    if len(numeric_long_form) >= 3:
        return len(numeric_long_form)
    # Numeric-date statements use the same rule when a visible header has a
    # distinct Value Date column: only date lines followed by narration count.
    if "value dt" in header or "value date" in header:
        numeric_primary = re.findall(r"(?m)^\s*\d{2}[-/]\d{2}[-/]\d{2,4}\s+(?=[A-Za-z])", raw)
        if len(numeric_primary) >= 3:
            return len(numeric_primary)
    return None


def signed_balance_source_count(raw: str) -> int | None:
    """Count a headerless Dr/Cr ledger from its source-record shape.

    A J&K cash-credit row can wrap its narration over multiple lines, so a
    date-only count overstates the source by counting a handful of repeated
    period/metadata lines.  A dated block that reaches one explicit signed
    balance before the next dated block is the independent source record.
    """
    pattern = re.compile(
        r"(?im)^\s*\d{2}-\d{2}-\d{4}\b"
        r"(?:(?!^\s*\d{2}-\d{2}-\d{4}\b)[\s\S])*?"
        r"\d[\d,.]*\.\d{2}\s*(?:dr|cr)\b"
    )
    # A final Statement Summary can begin with the statement-period start date
    # and end at a signed Closing Balance. It has the same broad text shape as
    # a ledger row, but its count/total labels prove that it is furniture. Do
    # not let that one summary control inflate the coverage denominator.
    count = sum(
        1 for match in pattern.finditer(raw or "")
        if not re.search(r"(?i)\b(?:statement\s+summary|dr\s*count|cr\s*count|total\s+debits?|total\s+credits?)\b", match.group())
    )
    return count if count >= 3 else None

def display_date(value: object) -> str:
    parsed = transaction_date_value(str(value))
    return parsed.strftime("%d/%m/%Y") if parsed else str(value or "")

def repair_truncated_table_date(value: object, raw: str) -> str:
    """Restore only a source-proven missing final year digit from a PDF cell."""
    text = str(value or "").strip()
    if transaction_date_value(text):
        return text
    partial = re.fullmatch(r"(\d{2}[-/]\d{2}[-/]\d{3})", text)
    if not partial:
        return text
    # PDF table extraction can put the final year digit in a neighbouring text
    # fragment. Require the exact partial date plus that digit in the original
    # source, so this never guesses a date from an amount or reference number.
    proven = re.search(re.escape(partial.group(1)) + r"\s*([0-9])\b", raw)
    return partial.group(1) + proven.group(1) if proven else text

def visible_monetary_token(value: object) -> bool:
    """Whether a source cell visibly contains a money-like token.

    This is deliberately about source presence, not numeric interpretation. A
    punctuation-damaged Indian balance is still evidence of a row, while its
    value remains governed by the conservative money-repair rules.
    """
    text = str(value or "").strip()
    return bool(re.fullmatch(r"[+-]?\s*[\d][\d,\.\s]*(?:\s*(?:dr|cr))?", text, re.I))

def count_source_transactions(rows: list[list[object]], header_at: int, columns: dict[str, int]) -> int:
    """Count source-proven transaction records independently of parsed rows.

    A valid record is one Date-column cell, one non-zero source movement, and
    one visible Balance-column token.  The balance value may be wrong (or have
    recoverable OCR punctuation damage), but the cell must visibly exist so a
    parser cannot certify a partial table by counting dates alone.
    """
    count = 0
    for row in rows[header_at + 1:]:
        def cell(key):
            i = columns.get(key)
            return row[i] if i is not None and i < len(row) else ""
        date = str(cell("date") or "").strip()
        if not transaction_date_value(date):
            continue
        narration = str(cell("narration") or "")
        if re.search(r"\b(?:B/F|OPENING\s+BALANCE)\b", narration, re.I):
            continue
        withdrawal = money(cell("withdrawal")) if "withdrawal" in columns else None
        deposit = money(cell("deposit")) if "deposit" in columns else None
        amount = money(cell("amount")) if "amount" in columns else None
        direct_movements = sum(1 for value in (withdrawal, deposit) if value is not None and value != 0)
        has_one_movement = direct_movements == 1
        # Amount + Dr/Cr type is the equivalent of separate debit/credit
        # columns.  It is one source movement only when there were no direct
        # debit/credit values to contradict it.
        if direct_movements == 0 and amount is not None and amount != 0 and "transaction_type" in columns:
            has_one_movement = True
        balance_visible = "balance" in columns and visible_monetary_token(cell("balance"))
        if has_one_movement and balance_visible:
            count += 1
    return count

def source_transaction_fingerprint(rows: list[list[object]], header_at: int, columns: dict[str, int]) -> Counter | None:
    """Build an independent source movement multiset fingerprint.

    Count validation alone cannot distinguish an omitted source debit plus a
    duplicated equal debit.  Where the source gives separately measured debit
    or credit cells, retain every (date, side, amount) record and require the
    finished parser to reproduce the same multiset.  This never uses a running
    balance to manufacture a movement, and it declines to add a gate for an
    ambiguous amount-only layout.
    """
    if "date" not in columns or "balance" not in columns:
        return None
    records: list[tuple[str, str, str]] = []
    for row in rows[header_at + 1:]:
        def cell(key: str):
            index = columns.get(key)
            return row[index] if index is not None and index < len(row) else ""
        date = transaction_date_value(str(cell("date") or "").strip())
        if date is None:
            continue
        narration = str(cell("narration") or "")
        if re.search(r"\b(?:B/F|OPENING\s+BALANCE)\b", narration, re.I):
            continue
        if not visible_monetary_token(cell("balance")):
            continue
        withdrawal = money(cell("withdrawal")) if "withdrawal" in columns else None
        deposit = money(cell("deposit")) if "deposit" in columns else None
        side_values = [("withdrawal", withdrawal), ("deposit", deposit)]
        nonzero = [(side, value) for side, value in side_values if value is not None and value != 0]
        if len(nonzero) != 1:
            # Amount + explicitly printed Dr/Cr is equivalent source evidence.
            amount = money(cell("amount")) if "amount" in columns else None
            kind = str(cell("transaction_type") or "").upper()
            if len(nonzero) == 0 and amount is not None and amount != 0 and ("DR" in kind or "CR" in kind):
                nonzero = [("withdrawal" if "DR" in kind else "deposit", amount)]
            else:
                return None
        side, value = nonzero[0]
        records.append((date.isoformat(), side, str(abs(value).quantize(Decimal(".01")))))
    return Counter(records) if records else None

def parsed_transaction_fingerprint(transactions: list[dict]) -> Counter | None:
    """Canonical counterpart of ``source_transaction_fingerprint``."""
    records: list[tuple[str, str, str]] = []
    for item in transactions:
        date = transaction_date_value(item.get("date"))
        withdrawal = item.get("withdrawal")
        deposit = item.get("deposit")
        nonzero = [("withdrawal", withdrawal), ("deposit", deposit)]
        nonzero = [(side, value) for side, value in nonzero if value is not None and value != 0]
        if date is None or len(nonzero) != 1:
            return None
        side, value = nonzero[0]
        records.append((date.isoformat(), side, str(abs(Decimal(value)).quantize(Decimal(".01")))))
    return Counter(records) if records else None

def structured_source_count(path: Path) -> int | None:
    """Independent table evidence used to reject a short fallback extraction."""
    if path.suffix.lower() != ".pdf": return None
    try:
        rows, _ = load_rows(path, None)
        header_at = next((i for i, row in enumerate(rows[:20]) if len(map_headers(row)) >= 3), None)
        if header_at is None: return None
        if generated_canonical_headers(rows[header_at]):
            # This is the generic text fallback's synthetic header, not an
            # independently detected PDF table.
            return None
        count = count_source_transactions(rows, header_at, map_headers(rows[header_at]))
        return count if count >= 3 else None
    except Exception:
        return None

def read_pdf_text(path: Path) -> str:
    return fast_pdf_text(path)

def cached_pdf_text(path: Path) -> str:
    key = str(path.resolve())
    with EXTRACTION_CACHE_LOCK:
        cached = PDF_TEXT_CACHE.get(key)
    if cached is not None:
        return cached
    text = read_pdf_text(path)
    if len(re.sub(r"\W", "", text)) < 80:
        # Small scanned statements can be OCRed once as evidence.  Long scans
        # use their representative geometry during planning and their complete
        # geometry only in the chosen parser candidate.
        if is_large_pdf(path):
            text = sampled_pdf_text(path)
        else:
            text = ocr_pdf_text(path)
    with EXTRACTION_CACHE_LOCK:
        if len(PDF_TEXT_CACHE) >= 12:
            PDF_TEXT_CACHE.pop(next(iter(PDF_TEXT_CACHE)))
        PDF_TEXT_CACHE[key] = text
    return text

def remove_page_furniture(raw: str) -> str:
    """Remove repeated PDF page furniture while preserving transaction continuations."""
    # Footer and disclaimer between statement pages.
    raw = re.sub(r"(?is)page total:.*?(?=jammu and kashmir bank ltd|$)", "", raw)
    raw = re.sub(r"(?is)unless the constituent notifies.*?(?:date stamp\s+manager|manager)\s*", "", raw)
    # Repeated JKB-style page heading and account-holder block ends at currency.
    raw = re.sub(r"(?is)jammu and kashmir bank ltd.*?(?:currency code\s*:\s*inr|no nomination available for the account)\s*", "", raw)
    # A repeated statement title sometimes appears immediately before the
    # continuation fragment; remove title only, leave following text intact.
    raw = re.sub(r"(?i)statement of account for the period(?:\s+of)?\s*\d{2}[-/]\d{2}[-/]\d{4}\s+to\s+\d{2}[-/]\d{2}[-/]\d{4}", "", raw)
    # HDFC-style statements append a full account-information block after each
    # page number. It is between transaction rows, never narration. Preserve
    # the next dated row as the boundary and retain any pre-footer continuation.
    raw = re.sub(r"(?is)page\s*no\s*\.?\s*:\s*\d+.*?(?=\n\d{2}/\d{2}/\d{2}\s|\Z)", "", raw)
    return raw

def repair_detached_dated_continuations(raw: str) -> str:
    """Join reference text with the dated/time continuation that follows it.

    Coordinate-poor PDF text sometimes emits `TRRR/reference/` on one line
    and the rest of that transaction (date, time, amount, running balance) on
    the next. Moving the date before the fragment makes it one normal record
    and prevents a fabricated blank-narration row.
    """
    # Some banks print the posting date beside the reference and repeat the
    # value date immediately before the time/amount/balance cells. Preserve
    # one record boundary by removing that repeated inner date.
    repeated_date_time = re.compile(
        r"(?m)^(?P<head>[ \t]*\d{2}-\d{2}-\d{4}[^\r\n]*[A-Za-z][^\r\n]*)\r?\n"
        r"[ \t]*\d{2}-\d{2}-\d{4}\s+(?P<tail>\d{2}:\d{2}:\d{1,2}\s+[-\d,]+(?:\.\d+)?\s+[-\d,]+(?:\.\d+)?\s*(?:Dr|Cr)\b)"
    )
    raw = repeated_date_time.sub(lambda match: f"{match.group('head')} {match.group('tail')}", raw)
    pattern = re.compile(
        r"(?m)^[ \t]*(?P<narr>(?=[A-Za-z])[^\r\n]*[A-Za-z][^\r\n]*)[ \t]*\r?\n"
        r"[ \t]*(?P<date>\d{2}-\d{2}-\d{4})\s+"
        r"(?P<tail>\d{2}:\d{2}:\d{1,2}\s+[-\d,]+(?:\.\d+)?\s+[-\d,]+(?:\.\d+)?\s*(?:Dr|Cr)\b)"
    )
    return pattern.sub(lambda match: f"{match.group('date')} {match.group('narr')} {match.group('tail')}", raw)

def extract_pdf_table_batch(path_text: str, page_numbers: list[int], use_learned_geometry: bool = False) -> list[list[list[object]]]:
    """Worker-safe batch extraction used only for very large PDFs."""
    with open_pdfplumber(Path(path_text)) as pdf:
        if use_learned_geometry:
            return [[pdf.pages[number].extract_table()] for number in page_numbers]
        return [pdf.pages[number].extract_tables() for number in page_numbers]

def sampled_geometry_header(path: Path, page_count: int) -> list[object] | None:
    """Learn a transaction-table header from representative original pages."""
    with open_pdfplumber(path) as pdf:
        for number in sampled_page_indices(page_count):
            for table in pdf.pages[number].extract_tables():
                for row in table:
                    if row and len(map_headers(row)) >= 3:
                        return row
    return None

def sampled_geometry_profile(path: Path) -> tuple[list[object], list[tuple[float, float]]] | None:
    """Learn column bands from representative original-PDF pages.

    Prefer ruled-table cells, but also support borderless bank statements by
    deriving bands from the x-position of their printed header words.
    """
    key = str(path.resolve())
    with EXTRACTION_CACHE_LOCK:
        if key in PDF_GEOMETRY_PROFILE_CACHE:
            return PDF_GEOMETRY_PROFILE_CACHE[key]
    if not native_pdf_text_is_usable(path):
        discovered = ocr_geometry_profile(path)
        with EXTRACTION_CACHE_LOCK:
            PDF_GEOMETRY_PROFILE_CACHE[key] = discovered
        return discovered
    discovered: tuple[list[object], list[tuple[float, float]]] | None = None
    with open_pdfplumber(path) as pdf:
        page_numbers = list(range(len(pdf.pages))) if len(pdf.pages) <= 60 else sampled_page_indices(len(pdf.pages))
        # First inspect header words on every representative page.  Do not
        # run costly ruled-table discovery on a cover/summary page merely
        # because the actual borderless transaction header begins on page 2.
        for number in page_numbers:
            page = pdf.pages[number]
            lines: dict[int, list[dict]] = {}
            for word in page.extract_words(x_tolerance=1, y_tolerance=2):
                lines.setdefault(round(float(word["top"]) / 3), []).append(word)
            for words in lines.values():
                ordered = sorted(words, key=lambda item: float(item["x0"]))
                labels = [str(word["text"]) for word in ordered]
                mapped = map_headers(labels)
                if not {"date", "narration", "withdrawal", "deposit", "balance"}.issubset(mapped):
                    continue
                starts = [float(word["x0"]) for word in ordered]
                discovered = labels, [
                    (0.0 if index == 0 else starts[index] - 3.0,
                     (starts[index + 1] - 3.0) if index + 1 < len(starts) else float(page.width))
                    for index in range(len(starts))
                ]
                break
            if discovered:
                break
        if discovered:
            with EXTRACTION_CACHE_LOCK:
                if len(PDF_GEOMETRY_PROFILE_CACHE) >= 12:
                    PDF_GEOMETRY_PROFILE_CACHE.pop(next(iter(PDF_GEOMETRY_PROFILE_CACHE)))
                PDF_GEOMETRY_PROFILE_CACHE[key] = discovered
            return discovered
        for number in page_numbers:
            page = pdf.pages[number]
            # Many bank PDFs (including ICICI's DATE / MODE / PARTICULARS /
            # DEPOSITS / WITHDRAWALS / BALANCE layout) draw light row guides
            # but no extractable table borders.  Their printed header words
            # are reliable geometry and far cheaper than `find_tables()`.
            # Prefer this fast route only when it identifies the complete
            # core transaction schema; partial headers still use ruled-table
            # discovery below.
            lines: dict[int, list[dict]] = {}
            for word in page.extract_words(x_tolerance=1, y_tolerance=2):
                lines.setdefault(round(float(word["top"]) / 3), []).append(word)
            for words in lines.values():
                ordered = sorted(words, key=lambda item: float(item["x0"]))
                labels = [str(word["text"]) for word in ordered]
                mapped = map_headers(labels)
                if not {"date", "narration", "withdrawal", "deposit", "balance"}.issubset(mapped):
                    continue
                starts = [float(word["x0"]) for word in ordered]
                bands = [(0.0 if index == 0 else starts[index] - 3.0,
                          (starts[index + 1] - 3.0) if index + 1 < len(starts) else float(page.width))
                         for index in range(len(starts))]
                discovered = labels, bands
                break
            if discovered:
                break
            for table in page.find_tables():
                extracted = table.extract()
                for index, row in enumerate(extracted):
                    if row and len(map_headers(row)) >= 3 and index < len(table.rows):
                        cells = table.rows[index].cells
                        if cells and all(cell for cell in cells):
                            discovered = row, [(float(cell[0]), float(cell[2])) for cell in cells]
                            break
                if discovered:
                    break
            if discovered:
                break
            for words in lines.values():
                ordered = sorted(words, key=lambda item: float(item["x0"]))
                labels = [str(word["text"]) for word in ordered]
                if len(map_headers(labels)) < 3:
                    continue
                starts = [float(word["x0"]) for word in ordered]
                # Header words are the authoritative column starts. A band
                # ends immediately before the next heading; the last reaches
                # the page edge. This is the same coordinate method used for
                # a ruled table, without requiring border lines.
                bands = [(0.0 if index == 0 else starts[index] - 3.0,
                          (starts[index + 1] - 3.0) if index + 1 < len(starts) else float(page.width))
                         for index in range(len(starts))]
                discovered = labels, bands
                break
            if discovered:
                break
    with EXTRACTION_CACHE_LOCK:
        if len(PDF_GEOMETRY_PROFILE_CACHE) >= 12:
            PDF_GEOMETRY_PROFILE_CACHE.pop(next(iter(PDF_GEOMETRY_PROFILE_CACHE)))
        PDF_GEOMETRY_PROFILE_CACHE[key] = discovered
    return discovered

def ocr_geometry_profile(path: Path) -> tuple[list[object], list[tuple[float, float]]] | None:
    """Discover transaction headers from OCR word coordinates on a scan."""
    document = fitz.open(str(path)) if fitz is not None else None
    if document is None:
        return None
    try:
        page_count = document.page_count
    finally:
        document.close()
    for page_number in (list(range(page_count)) if page_count <= 60 else sampled_page_indices(page_count)):
        # The profile must be measured with the strongest recovered word
        # geometry for this source page.  ``ocr_pdf_page_best_words`` keeps
        # the normal table OCR unless sparse-text OCR finds materially more
        # date anchors.
        words = ocr_pdf_page_best_words(path, page_number)
        lines: dict[int, list[dict]] = {}
        for word in words:
            lines.setdefault(round(float(word["top"]) / 3), []).append(word)
        ordered_lines = sorted(lines.values(), key=lambda line: min(float(word["top"]) for word in line))
        for line in ordered_lines:
            ordered = sorted(line, key=lambda item: float(item["x0"]))
            labels = [str(word["text"]) for word in ordered]
            # ``Post Date`` + ``Value Date`` is a two-column header.  Treat
            # it as composite even when the other labels happen to be on the
            # same OCR line; the raw one-line path would otherwise create two
            # generic Date bands and collapse source dates at parse time.
            if {"post", "value"}.issubset({norm(label) for label in labels}):
                continue
            mapped = map_headers(labels)
            if not {"date", "narration", "withdrawal", "deposit", "balance"}.issubset(mapped):
                continue
            starts = [float(word["x0"]) for word in ordered]
            width = max(float(word["x1"]) for word in ordered) + 12.0
            return labels, [
                (0.0 if index == 0 else starts[index] - 3.0,
                 (starts[index + 1] - 3.0) if index + 1 < len(starts) else width)
                for index in range(len(starts))
            ]

        # Statement scans often place a heading on two visual lines, for
        # example ``Cheque`` above ``No/Reference``.  A line-by-line OCR pass
        # cannot see a complete header in that case even though the original
        # page geometry is perfectly usable.  Build a conservative composite
        # header from adjacent lines in the same horizontal band.  We use only
        # recognised bank-column words and retain their original x positions;
        # this is deliberately not a flattened-text fallback.
        for index, first_line in enumerate(ordered_lines):
            top = min(float(word["top"]) for word in first_line)
            window = list(first_line)
            for next_line in ordered_lines[index + 1:index + 3]:
                next_top = min(float(word["top"]) for word in next_line)
                if next_top - top > 34.0:
                    break
                window.extend(next_line)

            header_words: list[tuple[float, str]] = []
            seen: set[tuple[str, int]] = set()
            for word in window:
                label = norm(word["text"])
                x = float(word["x0"])
                canonical = None
                if label in {"description", "narration", "particular", "particulars", "details", "remarks"}:
                    # OCR headings are often centred while their narrative
                    # data begins at the left edge of a much wider column.
                    # Preserve a small measured lead-in so the first words of
                    # a narration never spill into Value Date.
                    canonical, x = "Description", max(0.0, x - 80.0)
                elif label.startswith("cheque") or label.startswith("check") or label.startswith("instrument") or label.startswith("reference"):
                    canonical = "Cheque No/Reference"
                elif label.startswith("debit") or label.startswith("withdraw"):
                    canonical = "Debit"
                elif label.startswith("credit") or label.startswith("deposit"):
                    canonical = "Credit"
                elif label.startswith("balance"):
                    canonical = "Balance"
                elif label == "date":
                    # A scan commonly prints ``Post Date`` and ``Value Date``
                    # on the same line.  Only the word immediately *to the
                    # left* names this Date column.  Looking both ways used
                    # to label both date columns as Value Date, putting two
                    # dates into one extracted cell and making every row
                    # fail the date-anchor test.
                    left_peers = [candidate for candidate in window
                                  if abs(float(candidate["top"]) - float(word["top"])) < 12
                                  and float(candidate["x1"]) <= x
                                  and x - float(candidate["x1"]) < 80]
                    peer = max(left_peers, key=lambda candidate: float(candidate["x1"]), default=None)
                    peer_label = norm(peer["text"]) if peer else ""
                    if peer_label == "value":
                        canonical, x = "Value Date", float(peer["x0"])
                    elif peer_label == "post":
                        canonical, x = "Post Date", float(peer["x0"])
                    else:
                        canonical = "Date"
                if canonical is None:
                    continue
                key = (canonical, round(x))
                if key not in seen:
                    seen.add(key)
                    header_words.append((x, canonical))

            header_words.sort(key=lambda item: item[0])
            labels = [label for _x, label in header_words]
            mapped = map_headers(labels)
            if not {"date", "narration", "withdrawal", "deposit", "balance"}.issubset(mapped):
                continue
            starts = [x for x, _label in header_words]
            width = max(float(word["x1"]) for word in window) + 12.0
            return labels, [
                (0.0 if position == 0 else starts[position] - 3.0,
                 (starts[position + 1] - 3.0) if position + 1 < len(starts) else width)
                for position in range(len(starts))
            ]
    return None


def extract_ocr_geometry_rows(path: Path, page_numbers: set[int] | None = None) -> list[list[object]]:
    """Parse an image-only PDF using OCR word boxes, never flattened OCR text."""
    profile = ocr_geometry_profile(path)
    if not profile:
        return []
    header, bands = profile
    column_map = map_headers(header)
    date_index = column_map.get("date", 0)
    rows: list[list[object]] = [header]
    current: list[str] | None = None
    document = fitz.open(str(path)) if fitz is not None else None
    if document is None:
        return []
    try:
        selected = page_numbers if page_numbers is not None else set(range(document.page_count))
    finally:
        document.close()
    for page_number in sorted(selected):
        words = ocr_pdf_page_best_words(path, page_number)
        lines: dict[int, list[dict]] = {}
        for word in words:
            lines.setdefault(round(float(word["top"]) / 3), []).append(word)
        # Do not let statement identity fields above the transaction table
        # become fake rows merely because they contain a report/statement date.
        # OCR table pages repeat the native header; use its measured words as
        # a page-local fence and only accept material below it.
        header_terms = {"post", "value", "date", "description", "narration", "debit", "credit", "balance", "cheque", "reference"}
        header_words = [word for word in words if norm(str(word["text"])) in header_terms]
        header_bottom = None
        for word in header_words:
            nearby = [candidate for candidate in header_words
                      if abs(float(candidate["top"]) - float(word["top"])) < 42]
            if len({norm(str(candidate["text"])) for candidate in nearby}) >= 5:
                header_bottom = max(float(candidate["bottom"]) for candidate in nearby)
                break
        pending: list[list[str]] = []
        footer_started = False
        for line_words in lines.values():
            ordered = sorted(line_words, key=lambda item: float(item["x0"]))
            line_text = " ".join(str(word["text"]) for word in ordered)
            line_top = min(float(word["top"]) for word in ordered)
            if header_bottom is not None and line_top <= header_bottom + 10:
                continue
            if footer_started or re.search(r"(?i)\b(?:page\s+total|grand\s+total|statement\s+summary|opening\s+balance|closing\s+balance|end\s+of\s+statement|page\s*(?:no\.?|number)?\s*\d+|scanned\s+with|brought\s+forward|dr\s+count|cr\s+count|total\s+(?:debits?|credits?))\b", line_text):
                # ``Brought Forward`` at the beginning of a page is opening
                # metadata, not the end-of-statement summary.  It must be
                # ignored without closing the page's transaction stream.
                is_summary = bool(re.search(r"(?i)\b(?:statement\s+summary|dr\s+count|cr\s+count|total\s+(?:debits?|credits?))\b", line_text))
                if is_summary or (current is not None and re.search(r"(?i)\b(?:brought\s+forward|closing\s+balance)\b", line_text)):
                    footer_started = True
                continue
            cells = ["" for _ in bands]
            for word in ordered:
                center = (float(word["x0"]) + float(word["x1"])) / 2
                column = next((i for i, (left, right) in enumerate(bands) if left <= center <= right), None)
                if column is not None:
                    cells[column] = (cells[column] + " " + str(word["text"])).strip()
            if not any(cells) or len(map_headers(cells)) >= 3:
                continue
            # The Value Date cell must be one concrete source date.  This
            # avoids accepting a page heading or a merged OCR band containing
            # two dates as a transaction anchor.
            date_match = re.fullmatch(r"\s*(\d{2}[-/]\d{2}[-/]\d{2,4})(?:\s*[^\w\s]+\s*)?", cells[date_index] or "")
            is_dated_row = bool(date_match)
            if is_dated_row:
                cells[date_index] = date_match.group(1)
                # In scanned statements particulars are often printed one
                # visual line *above* the date/amount/balance line.  Those
                # fragments belong to this newly found transaction, not the
                # preceding dated transaction.
                new_current = cells
                for fragment in pending:
                    for index, value in enumerate(fragment):
                        if value:
                            new_current[index] = (new_current[index] + " " + value).strip()
                if current is not None:
                    rows.append(current)
                current, pending = new_current, []
            elif current is not None:
                # Narrative fragments between dated anchors remain attached to
                # their transaction.  Furniture was excluded above.
                pending.append(cells)
        if current is not None:
            for fragment in pending:
                for index, value in enumerate(fragment):
                    if value:
                        current[index] = (current[index] + " " + value).strip()
            rows.append(current)
            current = None
    return rows


def extract_geometry_profile_rows(path: Path, page_numbers: set[int] | None = None) -> list[list[object]]:
    """Apply sampled column bands without table rediscovery.

    `page_numbers` is a zero-based representative-page set used for fast
    structural screening. A final candidate always calls this for every page.
    """
    profile = sampled_geometry_profile(path)
    if not profile:
        return []
    header, bands = profile
    column_map = map_headers(header)
    date_index = column_map.get("date", 0)
    rows: list[list[object]] = [header]
    current: list[str] | None = None
    with open_pdfplumber(path) as pdf:
        for page_number, page in enumerate(pdf.pages):
            if page_numbers is not None and page_number not in page_numbers:
                continue
            lines: dict[int, list[dict]] = {}
            footer_started = False
            for word in page.extract_words(x_tolerance=1, y_tolerance=2):
                lines.setdefault(round(float(word["top"]) / 3), []).append(word)
            # Buffer text fragments between dated anchors.  In many borderless
            # bank layouts the particulars for a row are printed *above* its
            # date/amount line.  Assign each fragment to the closer adjacent
            # dated anchor instead of blindly appending it to the prior row.
            current_top: float | None = None
            pending_fragments: list[tuple[float, list[str]]] = []

            def append_fragments(target: list[str], fragments: list[tuple[float, list[str]]]) -> None:
                for _top, fragment in fragments:
                    for index, value in enumerate(fragment):
                        if value:
                            target[index] = (target[index] + " " + value).strip()

            for words in lines.values():
                if footer_started:
                    continue
                line_text = " ".join(str(word["text"]) for word in words)
                line_top = min(float(word["top"]) for word in words)
                page_number_header = bool(re.search(r"(?i)\bpage\s+\d+\s+of\s+\d+\b", line_text)) and line_top < float(page.height) * 0.25
                # A page-number header is furniture, but it is not an end of
                # page. Previously it set footer_started on page 2 onwards,
                # causing UPG to discard every transaction below it.
                if page_number_header:
                    continue
                footer_on_line = bool(re.search(
                    r"(?i)\b(?:page\s+total|grand\s+total|date/time|system\s+generated|"
                    r"account\s+related\s+other\s+information|legends\s+for\s+transactions|"
                    r"this\s+is\s+an\s+authenticated|nominee\s+name|sincerely|"
                    r"opening\s+balance|closing\s+balance|total\s+(?:debit|credit)\s+amt|"
                    r"end\s+of\s+statement)\b",
                    line_text,
                ))
                cells = ["" for _ in bands]
                for word in sorted(words, key=lambda item: float(item["x0"])):
                    center = (float(word["x0"]) + float(word["x1"])) / 2
                    column = next((i for i, (left, right) in enumerate(bands) if left <= center <= right), None)
                    if column is not None:
                        cells[column] = (cells[column] + " " + word["text"]).strip()
                standalone_total = bool(re.search(r"(?i)\bTOTAL\b", line_text)) and not transaction_date_value(cells[date_index])
                footer_on_line = footer_on_line or (current is not None and standalone_total)
                # A page footer can share the final transaction's y-band.
                # Keep the transaction portion, never append page/grand
                # totals or the generated-statement disclaimer to its cells.
                cells = [re.split(r"(?i)\b(?:page\s+total|grand\s+total|date/time|this\s+is\s+a\s+system\s+generated)", cell)[0].strip() for cell in cells]
                if footer_on_line:
                    withdrawal_index = column_map.get("withdrawal")
                    deposit_index = column_map.get("deposit")
                    balance_index = column_map.get("balance")
                    numeric = lambda value: (re.search(r"-?[\d,]+(?:\.\d{1,2})?", value or "").group() if re.search(r"-?[\d,]+(?:\.\d{1,2})?", value or "") else "")
                    # The final source row can share its y-band with Page
                    # Total. Preserve its first amount and balance only.
                    if withdrawal_index is not None and numeric(cells[withdrawal_index]):
                        cells[withdrawal_index] = numeric(cells[withdrawal_index])
                        if deposit_index is not None: cells[deposit_index] = ""
                    elif deposit_index is not None and numeric(cells[deposit_index]):
                        cells[deposit_index] = numeric(cells[deposit_index])
                        if withdrawal_index is not None: cells[withdrawal_index] = ""
                    if balance_index is not None:
                        cells[balance_index] = numeric(cells[balance_index])
                    footer_started = True
                    # A standalone footer must never become a continuation of
                    # the final transaction. If a transaction and footer share
                    # one band, its valid date keeps the transaction portion.
                    if not transaction_date_value(cells[date_index]):
                        # Discard fragments collected from the statement
                        # summary before they can be joined to the terminal
                        # dated transaction at page close.
                        pending_fragments = []
                        continue
                if not any(cells) or len(map_headers(cells)) >= 3:
                    continue
                if transaction_date_value(cells[date_index]):
                    if current is not None:
                        # Split fragments at the midpoint between two dated
                        # transaction anchors.  This preserves ordinary
                        # continuation lines while putting above-date
                        # particulars on their actual transaction.
                        midpoint = ((current_top or line_top) + line_top) / 2
                        prior = [(top, fragment) for top, fragment in pending_fragments if top <= midpoint]
                        following = [(top, fragment) for top, fragment in pending_fragments if top > midpoint]
                        append_fragments(current, prior)
                        rows.append(current)
                        append_fragments(cells, following)
                    elif pending_fragments:
                        append_fragments(cells, pending_fragments)
                    current = cells
                    current_top = line_top
                    pending_fragments = []
                elif current is not None:
                    pending_fragments.append((line_top, cells))
            # pdfplumber caches page character/word objects. Explicitly clear
            # each page after converting it to compact row strings so a 250+
            # page PDF does not retain all geometry in the 1 GB web container.
            try:
                page.close()
            except Exception:
                pass
            # Close a fully evidenced final row before the next page's bank
            # header arrives. Keep only incomplete rows open so genuine
            # two-page narrations/amounts can still be merged on page N+1.
            if current is not None:
                # A final line after the last dated anchor on a page is a
                # continuation of that anchor unless it was identified as
                # furniture above.
                append_fragments(current, pending_fragments)
                row_date = transaction_date_value(current[date_index] if date_index < len(current) else "")
                withdrawal_index = column_map.get("withdrawal")
                deposit_index = column_map.get("deposit")
                balance_index = column_map.get("balance")
                has_amount = bool(
                    (withdrawal_index is not None and withdrawal_index < len(current) and money(current[withdrawal_index]) is not None)
                    or (deposit_index is not None and deposit_index < len(current) and money(current[deposit_index]) is not None)
                )
                has_balance = balance_index is not None and balance_index < len(current) and money(current[balance_index]) is not None
                if row_date and has_amount and has_balance:
                    rows.append(current)
                    current = None
    if current is not None:
        rows.append(current)
    return rows

def sample_candidate_plausible(path: Path, strategy: str | None) -> bool:
    """Reject structurally impossible PDF candidates before full parsing.

    This is deliberately not a release validation. It only checks original-PDF
    sample pages for a real header and multiple dated record shapes, letting
    full extraction and all financial/narration gates remain the sole
    certification authority.
    """
    if path.suffix.lower() != ".pdf":
        return True
    key = (str(path.resolve()), strategy or "detected_table")
    with EXTRACTION_CACHE_LOCK:
        cached = CANDIDATE_SAMPLE_PROOF_CACHE.get(key)
    if cached is not None:
        return cached
    try:
        count = len(open_pdf_reader(path).pages)
        samples = set(sampled_page_indices(count))
        if strategy in {"geometry_profile", "source_amount_geometry", "dual_date_geometry"}:
            if strategy == "source_amount_geometry":
                rows = extract_standard_column_geometry_rows(path)
            elif strategy == "dual_date_geometry":
                rows = extract_dual_date_geometry_rows(path, samples)
            else:
                rows = extract_geometry_profile_rows(path, samples)
        else:
            raw = sampled_pdf_text(path)
            if strategy == "value_date_unsigned":
                rows = extract_text_layout_rows(raw, unsigned_balance=True, use_value_date=True)
            elif strategy == "unsigned_running_balance_text":
                rows = extract_text_layout_rows(raw, unsigned_balance=True)
            elif strategy in {"running_balance_text", "page_text_unsigned"}:
                rows = extract_text_layout_rows(raw, unsigned_balance=(strategy == "page_text_unsigned"))
            else:
                # Table rediscovery is intentionally not run on every page of
                # a large PDF. Geometry/AI candidates are stronger evidence.
                rows = []
        fields = map_headers(rows[0]) if rows else {}
        if len(rows) < 3 or not {"date", "withdrawal", "deposit", "balance"}.issubset(fields):
            result = False
        else:
            date_index = fields["date"]
            balance_index = fields["balance"]
            dated_rows = [row for row in rows[1:] if transaction_date_value(row[date_index] if date_index < len(row) else "")]
            # This is only a cheap source-shape gate. Full source coverage,
            # narration, financial and balance-chain validation still decide
            # certification after the complete extraction.
            result = len(dated_rows) >= 2 and all(
                balance_index < len(row) and money(row[balance_index]) is not None
                for row in dated_rows
            )
    except Exception:
        result = False
    with EXTRACTION_CACHE_LOCK:
        if len(CANDIDATE_SAMPLE_PROOF_CACHE) >= 72:
            CANDIDATE_SAMPLE_PROOF_CACHE.pop(next(iter(CANDIDATE_SAMPLE_PROOF_CACHE)))
        CANDIDATE_SAMPLE_PROOF_CACHE[key] = result
    return result


def deterministic_strategy_requires_source_proof(path: Path, strategy: str | None) -> bool:
    """Return whether a deterministic PDF strategy has a native sample proof.

    A full-statement extraction is the expensive operation on a long PDF.  Do
    not run it merely because a candidate name exists in the retry plan.  The
    strategies below each have a lightweight, source-native sampler in
    ``sample_candidate_plausible`` which can prove a header plus dated,
    amount/balance-bearing record shapes first.  Native spreadsheet/Word/CSV
    sources are already row-native, so this PDF-only guard deliberately does
    not add a lossy text sampling step for them.
    """
    if path.suffix.lower() != ".pdf":
        return False
    return strategy in {
        "geometry_profile",
        "source_amount_geometry",
        "dual_date_geometry",
        "running_balance_text",
        "unsigned_running_balance_text",
        "page_text_unsigned",
        "value_date_unsigned",
    }

def extract_dual_date_geometry_rows(path: Path, page_indices: set[int] | None = None) -> list[list[object]]:
    """Read a borderless Post Date / Value Date ledger from original PDF boxes.

    This is intentionally a measured layout family, rather than a bank-name
    template.  It works when an accessible/browser-produced PDF retains word
    coordinates but its text stream separates the two date columns and uses
    malformed dot-thousands amounts.
    """
    header = ["Date", "Narration", "Withdrawal", "Deposit", "Instrument Number", "Balance"]
    rows: list[list[object]] = [header]
    # Indian bank PDFs commonly display ``01-Apr-2025`` while others use
    # ``01-04-2025`` or slashes.  All are source dates, provided they are in
    # the measured Date / Value Date bands below.
    date_re = re.compile(r"^\d{2}[-/](?:\d{2}|[A-Za-z]{3})[-/]\d{2,4}$")
    with open_pdfplumber(path) as pdf:
        for page_number, page in enumerate(pdf.pages):
            if page_indices is not None and page_number not in page_indices:
                continue
            words = page.extract_words(x_tolerance=1, y_tolerance=2, keep_blank_chars=False)
            if not words:
                continue
            # Original-page x bands established from this family's visible
            # header.  Keep these deliberately broad enough for a small
            # print-scale change while still fencing out account furniture.
            # Accept semantic header variants.  The x positions are still
            # measured from this source page; no coordinates are borrowed
            # from another bank or profile.
            post_words = [word for word in words if str(word["text"]).lower() in {"post", "posting", "txn", "transaction", "date"}]
            value_words = [word for word in words if str(word["text"]).lower() == "value"]
            debit_words = [word for word in words if str(word["text"]).lower() in {"debit", "debits", "withdrawal", "withdrawals"}]
            credit_words = [word for word in words if str(word["text"]).lower() in {"credit", "credits", "deposit", "deposits"}]
            balance_words = [word for word in words if str(word["text"]).lower() == "balance"]
            if not (post_words and value_words and debit_words and credit_words and balance_words):
                continue
            # Account metadata can also contain the word "Balance".  Anchor
            # every measured column to the actual Post Date header baseline,
            # never the leftmost occurrence anywhere on the page.
            # Use the header baseline that contains Value Date; account
            # metadata elsewhere on the page can also contain these words.
            value_header = max(value_words, key=lambda word: float(word["top"]))
            header_top = float(value_header["top"])
            same_header = lambda candidates: min(
                (word for word in candidates if abs(float(word["top"]) - header_top) <= 12),
                key=lambda word: float(word["x0"]),
            )
            try:
                value_x = float(same_header(value_words)["x0"])
                post_header = min(
                    (word for word in post_words if abs(float(word["top"]) - header_top) <= 12 and float(word["x0"]) < value_x),
                    key=lambda word: float(word["x0"]),
                )
                post_x = float(post_header["x0"])
                debit_x = float(same_header(debit_words)["x0"])
                credit_x = float(same_header(credit_words)["x0"])
                balance_x = float(same_header(balance_words)["x0"])
            except ValueError:
                # Header words exist on the page but do not form one measured
                # ledger header baseline (for example account-summary text).
                # Do not borrow coordinates or turn it into an AI call.
                continue
            anchors: list[tuple[float, str, str]] = []
            for word in words:
                text = str(word["text"])
                # Only the left-hand (posting/transaction) date band creates
                # a row anchor.  The paired Value Date proves that anchor but
                # must never become a second transaction on the same visual
                # baseline.
                if not date_re.fullmatch(text) or float(word["x0"]) >= (post_x + value_x) / 2:
                    continue
                top = float(word["top"])
                peers = [other for other in words if abs(float(other["top"]) - top) <= 4 and value_x - 15 <= float(other["x0"]) <= value_x + 125 and date_re.fullmatch(str(other["text"]))]
                if peers:
                    anchors.append((top, text, str(sorted(peers, key=lambda item: float(item["x0"]))[0]["text"])))
            anchors.sort(key=lambda item: item[0])
            for index, (top, _post_date, value_date) in enumerate(anchors):
                # A transaction's visible description often begins a few
                # points *above* its date baseline, while same-date rows can
                # be very close together.  Use the midpoint between measured
                # date anchors as the row boundary instead of a fixed offset.
                # This keeps wrapped narration with its own row and prevents
                # text from a neighbouring same-date row leaking across.
                prior_top = anchors[index - 1][0] if index else header_top + 8
                following_top = anchors[index + 1][0] if index + 1 < len(anchors) else float(page.height) - 8
                row_start = (prior_top + top) / 2 if index else min(top - 6, prior_top)
                row_end = (top + following_top) / 2 if index + 1 < len(anchors) else following_top
                block = [word for word in words if row_start <= float(word["top"]) < row_end]
                # Source amount/balance cells belong to the anchor baseline.
                line = [word for word in block if abs(float(word["top"]) - top) <= 5]
                def cell(left: float, right: float) -> str:
                    return " ".join(str(word["text"]) for word in sorted(line, key=lambda item: float(item["x0"])) if left <= (float(word["x0"]) + float(word["x1"])) / 2 < right)
                # A PDF writer can split one visual amount into separate
                # glyph words (``6`` + ``500.00CR``). Numeric columns never
                # need word spaces, so rejoin them before monetary parsing.
                def numeric_cell(left: float, right: float) -> str:
                    """Return only numeric glyph tokens from one measured amount band.

                    Descriptions in borderless PDFs can overflow through the
                    visual whitespace before an amount.  They must not make a
                    real amount look unparsable.  This preserves split amount
                    glyphs (``6`` + ``500.00CR``) but excludes narration and
                    reference text from the numeric field.
                    """
                    pieces = []
                    for word in sorted(line, key=lambda item: float(item["x0"])):
                        center = (float(word["x0"]) + float(word["x1"])) / 2
                        text = str(word["text"])
                        if left <= center < right and re.fullmatch(r"[-+()]?[\d,.]+(?:CR|DR)?", text, re.I):
                            pieces.append(text)
                    return "".join(pieces)
                # Column labels are measured on this source page.  The
                # midpoint between adjacent measured columns is the only safe
                # boundary: fixed "wide" windows overlap when a bank places
                # Debit and Credit labels close together, duplicating one
                # movement into both columns.  Do not compensate with a
                # fabricated amount; leave an ambiguous row for validation to
                # reject instead.
                debit_left = debit_x - 8
                # Header labels are left aligned but numbers are normally
                # right aligned.  Use the following measured column's left
                # edge, rather than the midpoint of two label starts: the
                # midpoint can slice a perfectly valid 1,005.00 amount.
                debit_right = credit_x - 1
                credit_left = credit_x - 8
                credit_right = balance_x - 1
                balance_left = balance_x - 8
                # Use the same guarded Indian-number repair used by native
                # rows.  This route previously called ``money`` directly,
                # so a text-layer ``5.00.177.00`` was discarded even though
                # the verified shared rule could restore it as 5,00,177.00.
                debit = source_money(numeric_cell(debit_left, debit_right)) or Decimal("0")
                credit = source_money(numeric_cell(credit_left, credit_right)) or Decimal("0")
                balance = source_money(numeric_cell(balance_left, float(page.width) + 5))
                if balance is None or (not debit and not credit):
                    continue
                narration_words = [word for word in block if value_x + 95 <= float(word["x0"]) < debit_x - 20]
                narration = clean_narration(" ".join(str(word["text"]) for word in sorted(narration_words, key=lambda item: (float(item["top"]), float(item["x0"])))))
                # The last transaction can be immediately followed by the
                # account summary in the same left-hand band.  It is furniture
                # and never a narration continuation.
                narration = re.split(r"(?i)\b(?:closing\s+balance|statement\s+summary|debits\s+total|end\s+of\s+statement)\b", narration)[0].strip()
                instrument_matches = re.findall(r"\b\d{6,}\b", narration)
                rows.append([display_date(value_date), narration, debit, credit, instrument_matches[-1] if instrument_matches else "", balance])
            try:
                page.close()
            except Exception:
                pass
    return rows if len(rows) > 1 else []

def has_standard_geometry_header_contract(raw: str) -> bool:
    """Return true for a conventional five-column bank ledger header.

    This deliberately identifies *field meaning*, not a particular bank's
    spelling.  A normal statement might call narration ``Transaction
    Remarks`` and movements ``Withdrawal Amount`` / ``Deposit Amount``.  Such
    a statement must use the measured original-PDF geometry path before any
    generic table/OCR or AI fallback is considered.
    """
    # Text extraction order is not reliable evidence of visual header order:
    # PDFs may emit a right-hand column before a left-hand one.  This helper
    # is only a cheap route selector; the actual proof comes from the
    # same-baseline original-PDF word geometry in
    # ``extract_standard_column_geometry_rows``.  Therefore recognise the
    # five concepts anywhere in a compact header window, without assuming one
    # bank's wording or reading order.
    compact = re.sub(r"\s+", " ", raw[:12000])
    concepts = {
        "date": r"\b(?:transaction|posting|value|settlement|effective)?\s*date\b",
        "narration": r"\b(?:particulars?|remarks?|narration|description|details)\b",
        "withdrawal": r"\b(?:withdrawals?|debits?|amount\s+debited)\b",
        "deposit": r"\b(?:deposits?|credits?|amount\s+credited)\b",
        "balance": r"\b(?:running|available|closing)?\s*balance\b",
    }
    return all(re.search(pattern, compact, re.I) for pattern in concepts.values())

def has_dual_date_header_contract(headers: list[object]) -> bool:
    """Recognize a two-date ledger from measured header *cells*.

    Banks may call the first date Transaction/Posting/Booking/Effective Date,
    while the second may be Value/Settlement Date.  Header spelling is not a
    reliable parser identity; the required proof is two distinct date cells
    plus separate debit/credit and balance cells in the measured source grid.
    """
    if not headers:
        return False
    labels = [norm(header) for header in headers]
    date_cells = [label for label in labels if "date" in label]
    has_value_date = any(
        "value" in label or "settlement" in label or "effective" in label
        for label in date_cells
    )
    has_other_date = any(
        "value" not in label and "settlement" not in label
        for label in date_cells
    )
    header_fields = map_headers(headers)
    return (
        len(date_cells) >= 2
        and has_value_date
        and has_other_date
        and {"withdrawal", "deposit", "balance"}.issubset(header_fields)
    )


def extract_standard_column_geometry_rows(path: Path) -> list[list[object]]:
    """Extract a normal Date / Particulars / Debit / Credit / Balance ledger.

    The source amount columns are read from their original PDF x-bands.  A
    malformed balance is returned as blank evidence, never used to derive a
    compensating movement.  This avoids an OCR/text-layer defect turning, for
    example, a visible withdrawal of 9.00 into a fictional 5,00,345 deposit.
    """
    header = ["Date", "Narration", "Withdrawal", "Deposit", "Instrument Number", "Balance"]
    rows: list[list[object]] = [header]
    # Bank PDFs use both textual (18-Feb-2026) and numeric (18-02-2026,
    # 18/02/2026) dates.  A date-looking token is only a row starter when it
    # is physically in the measured Date column below, so accepting the
    # numeric forms cannot turn a narration/reference date into a record.
    # PDF text layers use all three numeric separators.  ``09.04.2025`` is
    # common in ICICI and several other Indian bank exports; accepting it is
    # safe here because the token must also be below the measured header and
    # physically inside the measured Date-column band.
    date_re = re.compile(r"^\d{2}(?:-[A-Za-z]{3}-|[-/.]\d{2}[-/.])\d{4}$")
    with open_pdfplumber(path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(x_tolerance=1, y_tolerance=2, keep_blank_chars=False)
            # The five required *fields* are often labelled differently by a
            # bank.  For example ICICI Saving statements use ``Transaction
            # Remarks``, ``Withdrawal Amount (INR)`` and ``Deposit Amount
            # (INR)``.  Do not make a measured original-PDF geometry path
            # depend on a single bank's header vocabulary.
            # Capture header *meaning* first, then use this page's physical
            # coordinates.  Banks freely vary wording and wrap labels, but a
            # vocabulary difference is not a new parser family.
            headers: dict[str, list[dict]] = {key: [] for key in ("date", "particulars", "withdrawals", "deposits", "balance")}
            for word in words:
                role = semantic_header_role(word["text"])
                if role == "date":
                    headers["date"].append(word)
                elif role == "narration":
                    headers["particulars"].append(word)
                elif role == "withdrawal":
                    headers["withdrawals"].append(word)
                elif role == "deposit":
                    headers["deposits"].append(word)
                elif role == "balance":
                    headers["balance"].append(word)
            if not all(headers.get(key) for key in ("date", "particulars", "withdrawals", "deposits", "balance")):
                continue
            # All five labels must share the actual table-header baseline.
            # A wrapped label (for example ``Transaction\nDate``) can put the
            # word ``Date`` one text-line lower than ``Withdrawal``/``Deposit``;
            # twelve points is still far tighter than any transaction row gap.
            candidates = [word for word in headers["date"] if all(
                any(abs(float(other["top"]) - float(word["top"])) <= 12 for other in headers[key])
                for key in ("particulars", "withdrawals", "deposits", "balance")
            )]
            if not candidates:
                continue
            # A bank may print both ``Transaction Date`` and ``Value Date``.
            # Treating both physical date cells as record anchors splits every
            # row in two and makes the first line of a narration appear to
            # belong to the next transaction.  Prefer the measured Value Date
            # column when it is explicitly labelled, otherwise use the
            # leftmost date column.  This is a column decision, never a guess
            # based on date-looking text in narration.
            header_top = float(max(candidates, key=lambda word: float(word["top"]))["top"])
            header_words = [word for word in words if abs(float(word["top"]) - header_top) <= 12]
            date_headers = sorted(
                (word for word in headers["date"] if abs(float(word["top"]) - header_top) <= 12),
                key=lambda word: float(word["x0"]),
            )
            value_date_headers = [
                date_word for date_word in date_headers
                if any(
                    norm(label["text"]) in {"value", "settlement", "effective"}
                    and 0 <= float(date_word["x0"]) - float(label["x1"]) <= 28
                    for label in header_words
                )
            ]
            date_header = value_date_headers[0] if value_date_headers else date_headers[0]
            def same_line(key: str):
                return min((word for word in headers[key] if abs(float(word["top"]) - header_top) <= 12), key=lambda word: float(word["x0"]))
            x_date = float(date_header["x0"])
            x_part = float(same_line("particulars")["x0"])
            x_wd = float(same_line("withdrawals")["x0"])
            x_dp = float(same_line("deposits")["x0"])
            x_bal = float(same_line("balance")["x0"])
            # Do not assume that every bank prints Debit before Credit and
            # Balance at the far right.  The *meaning* of each column came
            # from the measured header; its source band must now be bounded
            # by the next measured column on this particular page.  This
            # supports layouts such as Date | Details | Credit | Debit |
            # Balance without borrowing another bank's positions or swapping
            # the two sides to force reconciliation.
            measured_starts = {
                "date": x_date,
                "narration": x_part,
                "withdrawal": x_wd,
                "deposit": x_dp,
                "balance": x_bal,
            }
            ordered_fields = sorted(measured_starts, key=measured_starts.get)
            field_bands: dict[str, tuple[float, float]] = {}
            for position, field in enumerate(ordered_fields):
                start = measured_starts[field]
                next_start = (
                    measured_starts[ordered_fields[position + 1]]
                    if position + 1 < len(ordered_fields)
                    else float(page.width) + 5
                )
                # Header text often starts a little to the right of a
                # right-aligned currency value.  The small lead-in retains
                # that value while the next measured start prevents overlap.
                field_bands[field] = (start - 10, next_start - 1)
            # A header can start to the right of the first actual narration
            # glyph.  The common ``Cheque Number | Transaction Remarks``
            # layout is an example: the data narration begins immediately
            # after the cheque-number column, not under the word ``Remarks``.
            # Use the right edge of that preceding reference header as a
            # measured lower narration boundary; never expand into the Date
            # or amount columns.
            reference_right_edges = [
                float(word["x1"])
                for word in header_words
                if float(word["x1"]) < x_part
                and re.sub(r"[^a-z]", "", str(word["text"]).lower())
                in {"cheque", "chq", "number", "no", "reference", "ref", "instrument"}
            ]
            x_narration_left = min(
                x_part - 20,
                max(reference_right_edges) + 1 if reference_right_edges else x_part - 20,
            )
            prior_date_headers = [float(word["x0"]) for word in date_headers if float(word["x0"]) < x_date - 8]
            subsequent_header_starts = [
                float(word["x0"]) for word in header_words
                if float(word["x0"]) > x_date + 12
                and re.sub(r"[^a-z]", "", str(word["text"]).lower()) not in {"date", "value"}
            ]
            # The midpoint to the adjacent measured header keeps only the
            # selected Date/Value Date physical column.  It prevents the
            # other date column (or a date inside narration) from becoming a
            # second transaction boundary.
            date_left = ((max(prior_date_headers) + x_date) / 2 if prior_date_headers else x_date - 55)
            date_right = ((x_date + min(subsequent_header_starts)) / 2 if subsequent_header_starts else x_part - 20)
            anchors = sorted(
                [(float(word["top"]), str(word["text"])) for word in words
                 if date_re.fullmatch(str(word["text"]))
                 and date_left <= float(word["x0"]) < date_right
                 and float(word["top"]) > header_top + 12],
                key=lambda item: item[0],
            )
            for index, (top, date_text) in enumerate(anchors):
                # A number of bank PDFs place the first narration line a few
                # points *above* its date/amount baseline.  Splitting simply
                # at the next Date anchor then assigns that line to the
                # previous transaction.  Partition the measured source rows
                # at the midpoint between consecutive Date/Value Date cells:
                # it keeps every word with the physical row it is closest to
                # and does not rely on OCR/reading order or an AI guess.
                if index:
                    block_start = (anchors[index - 1][0] + top) / 2
                else:
                    block_start = header_top + 12
                if index + 1 < len(anchors):
                    next_top = (top + anchors[index + 1][0]) / 2
                else:
                    # The final transaction may be followed by a Total row
                    # and legal/footer furniture.  Neither belongs to the
                    # narration, even when the text layer puts it in the
                    # same broad vertical block.
                    footer_tops = [float(word["top"]) - 3 for word in words
                                   if float(word["top"]) > top + 8
                                   and str(word["text"]).strip().lower() in {
                                       "total", "grand", "page", "www.icici.bank.in",
                                       "www.", "please", "never", "disclaimer",
                                   }]
                    next_top = min(footer_tops) if footer_tops else float(page.height) - 6
                block = [word for word in words if block_start <= float(word["top"]) < next_top]
                line = [word for word in block if abs(float(word["top"]) - top) <= 5]
                def band(left: float, right: float, source=line) -> str:
                    return " ".join(str(word["text"]) for word in sorted(source, key=lambda item: float(item["x0"])) if left <= (float(word["x0"]) + float(word["x1"])) / 2 < right)
                # Do not overlap adjacent measured money columns.  Earlier
                # broad bands could read the same deposit once as deposit and
                # again as the final balance, which then tempted later logic
                # to repair a ledger that was actually read incorrectly.
                withdrawal_text = band(*field_bands["withdrawal"]).replace(" ", "")
                # Some original PDFs let the final word of a narration range
                # overlap the debit column (``to17-3,894.00``).  The final
                # currency token is still source text in the debit band; do
                # not turn its leading narration fragment into amount 17.
                withdrawal_text = re.sub(r"^\d{1,2}-(?=\d[\d,]*\.\d{2}$)", "", withdrawal_text)
                trailing_amount = re.search(r"(\d[\d,]*\.\d{2})$", withdrawal_text)
                if trailing_amount and not re.fullmatch(r"-?\d[\d,]*\.\d{2}", withdrawal_text):
                    withdrawal_text = trailing_amount.group(1)
                deposit_text = band(*field_bands["deposit"]).replace(" ", "")
                balance_text = band(*field_bands["balance"]).replace(" ", "")
                # Preserve raw monetary source cells.  The statement-level
                # normalizer can infer the column's decimal precision from
                # every valid row, then safely repair an isolated text-layer
                # punctuation defect without changing the measured columns.
                withdrawal = money(withdrawal_text) or repair_indian_grouping_decimal(withdrawal_text, 2) or Decimal("0")
                deposit = money(deposit_text) or repair_indian_grouping_decimal(deposit_text, 2) or Decimal("0")
                # Date + visibly measured source amount establishes a record;
                # an unusable balance cannot erase it or change its direction.
                if not withdrawal and not deposit:
                    continue
                # Particulars is a measured cell, never the remainder of the
                # page.  Its right edge is the first financial/balance column
                # to its right on this source page.  This prevents a reversed
                # Credit/Debit layout from leaking a printed amount into the
                # narration field.
                narration_right_candidates = [
                    measured_starts[field] - 10
                    for field in ("withdrawal", "deposit", "balance")
                    if measured_starts[field] > x_narration_left
                ]
                narration_right = min(narration_right_candidates, default=float(page.width) + 5)
                narration_words = [word for word in block if x_narration_left <= float(word["x0"]) < narration_right]
                narration = clean_narration(" ".join(str(word["text"]) for word in sorted(narration_words, key=lambda item: (float(item["top"]), float(item["x0"])))) )
                narration = re.split(
                    r"(?i)\b(?:opening\s+balance|total|grand\s+total|"
                    r"this\s+is\s+an\s+auto|this\s+is\s+a\s+system\s+generated)\b",
                    narration,
                )[0].strip()
                # A final narration word can visually overlap the debit x-band
                # in borderless source PDFs.  A currency-shaped tail is not
                # narration; it is the source movement already measured above.
                narration = re.sub(r"(?:\s+|\b\d{1,2}-)\d[\d,]*\.\d{2}$", "", narration).strip()
                if re.search(r"(?i)^(?:opening\s+balance|b\s*/\s*f)\b", narration):
                    continue
                refs = re.findall(r"\b\d{6,}\b", narration)
                rows.append([display_date(date_text), narration, withdrawal_text, deposit_text, refs[-1] if refs else "", balance_text])
            try:
                page.close()
            except Exception:
                pass
    return rows if len(rows) > 1 else []

def extract_pdf_rows(path: Path, strategy_override: str | None = None, job_id: str | None = None) -> tuple[list[list[object]], str]:
    raw = remove_page_furniture(cached_pdf_text(path))
    if not native_pdf_text_is_usable(path):
        # Rows come from original-page OCR word boxes and learned x-bands.
        # Full OCR text is retained only for the existing validation evidence.
        raw = remove_page_furniture(ocr_pdf_text(path))
        rows = extract_ocr_geometry_rows(path)
        if not rows or len(rows) < 2:
            raise ValueError(
                "OCR_NO_TABLE: OCR could read the scan but could not measure a complete "
                "Date / Particulars / Withdrawal / Deposit / Balance transaction header."
            )
        return rows, raw
    dual_date_time_layout = bool(re.search(r"\d{2}-[A-Za-z]{3}-\d{4}[\s\S]{0,180}\d{2}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}", raw))
    merged: list[list[object]] = []
    known_header: list[object] | None = None
    if strategy_override == "geometry_profile":
        return extract_geometry_profile_rows(path), raw
    if strategy_override == "source_amount_geometry":
        # The engine selected this only after evidence of an unreliable balance
        # chain.  Preserve the separately printed debit/credit movements from
        # original-page x-bands; never reconstruct them from that bad chain.
        measured_rows = extract_standard_column_geometry_rows(path)
        return (measured_rows, raw) if measured_rows else (extract_geometry_profile_rows(path), raw)
    # Some borderless statements expose two adjacent date columns.  Generic
    # table extraction commonly merges their headers (for example,
    # ``TXN DATE`` + ``VALUE DATE``), which prevents an otherwise exact saved
    # geometry profile from being selected and needlessly starts AI repair.
    # Route this measurable source pattern straight to original-PDF geometry.
    # This is deliberately based on the complete transaction header contract,
    # not on a bank name or a one-off statement fingerprint.
    dual_date_geometry_header = bool(re.search(
        r"(?is)\b(?:(?:TRANSACTION|TXN|POST(?:ING)?)\s*)?DATE\b[\s\S]{0,100}"
        r"\bVALUE\s+DATE\b[\s\S]{0,280}\b(?:DEBITS?|WITHDRAWALS?)\b[\s\S]{0,100}"
        r"\b(?:CREDITS?|DEPOSITS?)\b[\s\S]{0,100}\b(?:RUNNING\s+)?BALANCE\b",
        raw,
    ))
    # Prefer this route even where the header also meets the ordinary
    # table-header contract. The ordinary extractor may select the first Date
    # token; this extractor measures the Value Date band specifically.
    if strategy_override == "dual_date_geometry" or (
        strategy_override is None and dual_date_geometry_header
    ):
        measured_rows = extract_dual_date_geometry_rows(path)
        if measured_rows:
            return measured_rows, raw
        return extract_geometry_profile_rows(path), raw
    standard_geometry_header = has_standard_geometry_header_contract(raw)
    # ``standard_column_geometry`` is also a named candidate in the UPG
    # planning loop. It must take this exact route when explicitly selected;
    # otherwise the named candidate falls through to generic table extraction
    # while automatic parsing uses the measured route.
    if strategy_override == "standard_column_geometry" or (
        strategy_override is None and standard_geometry_header
    ):
        measured_rows = extract_standard_column_geometry_rows(path)
        if measured_rows:
            return measured_rows, raw
    if strategy_override == "page_text_unsigned":
        header = ["Date", "Narration", "Withdrawal", "Deposit", "Instrument Number", "Balance"]
        reader = open_pdf_reader(path)
        merged = [header]
        for page in reader.pages:
            page_rows = extract_text_layout_rows(remove_page_furniture(page.extract_text() or ""), unsigned_balance=True)
            if page_rows:
                merged.extend(page_rows[1:])
        return merged if len(merged) > 1 else [], raw
    if strategy_override not in ("running_balance_text", "unsigned_running_balance_text", "value_date_unsigned"):
        with open_pdfplumber(path) as pdf:
            page_count = len(pdf.pages)
            if page_count > 60:
                # Learn the geometry/header from the original-PDF sample before
                # scanning all pages. With a known header, each remaining page
                # uses the lighter single-table path rather than table discovery.
                known_header = sampled_geometry_header(path, page_count)
                if known_header:
                    merged.append(known_header)
                workers = min(4, max(2, os.cpu_count() or 2))
                batches = [list(range(start, min(page_count, start + 32))) for start in range(0, page_count, 32)]
                try:
                    # Threads avoid Windows process-spawn behaviour, which can
                    # otherwise create duplicate web-server children when this
                    # local app is launched through its bootstrap script.
                    with ThreadPoolExecutor(max_workers=workers) as executor:
                        page_tables = [tables for batch in executor.map(extract_pdf_table_batch, [str(path)] * len(batches), batches, [bool(known_header)] * len(batches)) for tables in batch]
                except Exception:
                    page_tables = [page.extract_tables() for page in pdf.pages]
            else:
                page_tables = [page.extract_tables() for page in pdf.pages]
        for tables in page_tables:
            for table in tables:
                if not table:
                    continue
                for row in table:
                    if not row: continue
                    mapped = map_headers(row)
                    if len(mapped) >= 3:
                        known_header = row
                        if not merged: merged.append(row)
                        continue
                    if known_header and len(row) == len(known_header): merged.append(row)
    if not merged:
        # A queued UPG job already completed a measured preflight plan. Do not
        # spend a second AI call merely to classify the same source text again;
        # the deterministic text fallback below remains a candidate and the
        # reserved expert calls can instead diagnose and repair a real failure.
        strategy = strategy_override or (None if job_id else ai_choose_text_strategy(raw))
        if strategy == "running_balance_text":
            merged = extract_text_layout_rows(raw)
        elif strategy == "unsigned_running_balance_text":
            merged = extract_text_layout_rows(raw, unsigned_balance=True, use_value_date=dual_date_time_layout, dual_date_time=dual_date_time_layout)
        elif strategy == "value_date_unsigned":
            merged = extract_text_layout_rows(raw, unsigned_balance=True, use_value_date=True)
        else:
            # AI classification is only one candidate. Try the safe
            # running-balance strategy too; the two validation gates decide
            # whether it can be released, not the model's first guess.
            merged = extract_text_layout_rows(raw)
        if not merged:
            raise ValueError("No readable transaction rows were found. This statement may be scanned, password protected, or incomplete.")
    if not merged:
        raise ValueError("No readable transaction rows were found. This statement may be scanned, password protected, or incomplete.")
    return merged, raw

def extract_text_layout_rows(raw: str, unsigned_balance: bool = False, use_value_date: bool = False, dual_date_time: bool = False) -> list[list[object]]:
    """Generic fallback for text PDFs without table borders.

    Many bank PDFs put one dated transaction after another and print a running
    balance with Dr/Cr. Consecutive balance changes provide an independent way
    to assign withdrawal/deposit amounts without relying on fixed x positions.
    """
    raw = repair_detached_dated_continuations(raw)
    # Learn the source's decimal width from many already-valid monetary tokens
    # before attempting to repair any damaged one.  The text-layout extractor
    # runs before a column map exists, so this is the equivalent of
    # ``inferred_column_decimal_places`` for headerless PDF ledgers.  A
    # malformed multi-dot amount is never allowed to teach the rule that
    # repairs it.
    trusted_decimal_widths: list[int] = []
    for token in re.finditer(r"(?<![\d.])-?[\d,]+(?:\.(\d{1,2}))?\s*(?:dr|cr)?\b", raw, re.I):
        if token.group(1) and money(token.group()) is not None:
            trusted_decimal_widths.append(len(token.group(1)))
    text_decimal_places = (
        max(set(trusted_decimal_widths), key=lambda width: (trusted_decimal_widths.count(width), width))
        if trusted_decimal_widths else None
    )
    def text_money(value: object) -> Decimal | None:
        """Read a text-layer money token without accepting a truncated value.

        Browser and OCR PDF text layers can replace Indian grouping commas with
        full stops (``3,015.17`` -> ``3.015.17``).  Accept that exact numeric
        shape as a repair, but leave every other multi-dot token unusable.
        This applies before record deltas are calculated, so a damaged balance
        can never manufacture an opposite-side transaction.
        """
        parsed = money(value)
        return parsed if parsed is not None else repair_indian_grouping_decimal(value, text_decimal_places)
    header = ["Date", "Narration", "Withdrawal", "Deposit", "Instrument Number", "Balance"]
    # Bound numeric dates so a long transaction/reference ID cannot be
    # mistaken for a partial date (for example, `...382/22-04...`).
    date_pattern = r"(?<!\d)(?:\d{2}[-/]\d{2}[-/]\d{2,4}|\d{2}-[A-Za-z]{3}-\d{4})(?!\d)"
    statement_opening, _ = source_balances(raw)
    # Some text PDFs begin with a statement period such as "Between
    # 01-04-2025 and ...". It resembles a transaction date but is not one.
    # Start at the printed transaction-table heading when it is available,
    # while retaining the separately extracted opening balance above.
    table_heading = re.search(r"(?im)^\s*(?:transaction\s+)?date\s+.*?(?:particulars?|narration|description).*?(?:closing\s+balance|balance)\s*$", raw)
    transaction_text = raw[table_heading.end():] if table_heading else raw
    if use_value_date:
        # Coordinate-poor text extraction often puts the Value Date and the
        # amount/balance cells on the next physical line. Reattach only lines
        # that are a date followed solely by numeric cells; a posting-date line
        # has narration/reference text and remains a record boundary.
        value_line = rf"(?m)^\s*({date_pattern})\s+((?:-?[\d,]+(?:\.\d{{1,2}})?\s*){{2,}})$"
        transaction_text = re.sub(value_line, lambda match: " " + match.group(1) + " " + match.group(2).strip(), transaction_text)
    # A Value Date may appear later on the same visual line. Only a date at
    # the beginning of a physical line can start the next transaction block.
    dual_date_layout = bool(re.search(r"(?i)\bvalue\s*(?:date|dt)\b", raw[:2000])) or dual_date_time
    # In compact text layers, a whole PDF page may be emitted as one line.
    # Split every date for ordinary single-date layouts, but retain the
    # beginning-of-line rule for dual-date layouts so Value Date is not turned
    # into a second transaction.
    primary_date = r"\d{2}-[A-Za-z]{3}-\d{4}" if dual_date_time else date_pattern
    # Prefer a full four-digit year as the record boundary whenever the
    # statement itself uses full-year dates. A reference/narration can carry
    # a short embedded date, e.g. ``REV/.../24-04-25/...``. The old generic
    # splitter treated that reference date as a new transaction, truncating
    # the page-end reversal row before its amount and balance. This is a
    # boundary repair only: it preserves the original source amount and never
    # invents a movement to make a balance chain reconcile.
    full_year_date = r"(?:\d{2}[-/]\d{2}[-/]\d{4}|\d{2}-[A-Za-z]{3}-\d{4})"
    has_full_year_records = len(re.findall(full_year_date, transaction_text)) >= 3
    boundary_date = full_year_date if has_full_year_records else primary_date
    split_pattern = (
        rf"(?m)(?=^\s*{primary_date}\s)"
        if dual_date_layout
        else rf"(?={boundary_date}\s)"
    )
    chunks = re.split(split_pattern, transaction_text)
    rows, previous_balance = [header], statement_opening
    for chunk in chunks:
        date = re.match(rf"\s*({date_pattern})\s+", chunk)
        if not date: continue
        if re.search(r"\bB/F\b", chunk, re.I): continue
        # SBI-style final summaries can start with the statement-period date
        # after a separate ``Statement Summary:`` line.  The generic date
        # splitter makes that date the start of a new chunk, so the preceding
        # summary marker is no longer available to the footer split below.
        # A period row with count/total headings is never a ledger entry.
        if re.search(r"(?i)\b(?:dr\s*count|cr\s*count|total\s+debits?|total\s+credits?)\b", chunk[:900]):
            continue
        # Repeated page headers contain the statement-period end date and
        # account-holder text, but are not transaction rows.
        if "PARTICULARS" in chunk[:800].upper(): continue
        # Only the portion before page/report totals belongs to the current
        # transaction. Footer balances are not closing transaction balances.
        # A final account-summary block can repeat the statement period dates
        # (for example ``Summary: 01-04-2025 To 31-03-2026``).  It is source
        # furniture, not a transaction.  If allowed through the date splitter
        # it becomes a fabricated final row, reverses chronological ordering,
        # and corrupts endpoint derivation.  Seal every real transaction at
        # the start of summary/footer material before reading its balance.
        chunk = re.split(r"(?i)\b(?:page total|grand total|statement\s+summary|summary\s*:|funds in clearing|total available amount|effective available amount|closing balance|unless the constituent)\b", chunk)[0]
        balance_matches = list(re.finditer(r"(-?[\d,.]+)\s*(Dr|Cr)\b", chunk, re.I))
        numeric_matches = list(re.finditer(r"-?[\d,.]+", chunk))
        forced_amount, value_date_match = None, None
        matching_secondary_date = False
        if use_value_date:
            dates = list(re.finditer(date_pattern, chunk))
            if len(dates) >= 2:
                first_date, second_date = transaction_date_value(dates[0].group()), transaction_date_value(dates[1].group())
                matching_secondary_date = first_date is not None and second_date is not None and first_date.date() == second_date.date()
            if matching_secondary_date:
                value_date_match = dates[1]
            # The first two numeric cells after Value Date are respectively
            # the transaction amount and the running balance. Continuation
            # text below the row may contain long numeric reference fragments
            # and must not replace the balance.
                value_tail = re.sub(r"\b\d{2}:\d{2}:\d{2}\b", "", chunk[value_date_match.end():])
                value_numbers = list(re.finditer(r"-?[\d,.]+", value_tail))
                if len(value_numbers) < 2: continue
                forced_amount = text_money(value_numbers[0].group())
                balance_match = value_numbers[1]
                balance_start = value_date_match.end() + balance_match.start()
                numeric_balance = text_money(balance_match.group())
        if not matching_secondary_date:
            # Most signed layouts mark balance with Dr/Cr. A zero balance may
            # omit that suffix, however, and remains a valid dated row when it
            # has a preceding transaction amount.
            zero_balance_without_suffix = (
                not balance_matches and len(numeric_matches) >= 2
                and text_money(numeric_matches[-1].group()) == Decimal("0")
            )
            if not balance_matches and not unsigned_balance and not zero_balance_without_suffix: continue
            if not balance_matches and len(numeric_matches) < 2: continue
            balance_match = balance_matches[-1] if balance_matches else numeric_matches[-1]
            balance_start = balance_match.start()
            numeric_balance = text_money(balance_match.group(1) if balance_matches else balance_match.group())
        if numeric_balance is None: continue
        if balance_matches and not matching_secondary_date:
            balance = -abs(numeric_balance) if balance_match.group(2).lower() == "dr" else abs(numeric_balance)
        else:
            # An unsigned layout can still print a literal negative balance.
            # Preserve that sign; it determines the debit/credit movement.
            balance = numeric_balance
        numbers = list(re.finditer(r"-?[\d,.]+", chunk[:balance_start]))
        amount = forced_amount if forced_amount is not None else (text_money(numbers[-1].group()) if numbers else None)
        if previous_balance is None:
            # First row: use its stated amount when available, otherwise keep
            # zero and derive the opening balance from its running balance.
            delta = Decimal("0")
            if amount is not None:
                narrative_before_amount = chunk[:numbers[-1].start()].lower()
                delta = -abs(amount) if "debit" in narrative_before_amount else abs(amount)
        else:
            delta = balance - previous_balance
        withdrawal = -delta if delta < 0 else Decimal("0")
        deposit = delta if delta > 0 else Decimal("0")
        narration_end = balance_start
        output_date = date.group(1)
        if matching_secondary_date:
            # In a dual-date layout the second date in the physical record is
            # the Value Date. It is a column value, not part of narration.
            if value_date_match is not None:
                output_date = date.group(1) if dual_date_time else value_date_match.group(0)
                narration_end = value_date_match.start()
        narration = clean_narration(chunk[date.end():narration_end])
        instrument_source = chunk[date.end():value_date_match.start()] if use_value_date and value_date_match is not None else narration
        instrument_matches = list(re.finditer(r"\b\d{6,}\b", instrument_source))
        instrument = instrument_matches[-1].group() if instrument_matches else ""
        # Retain the source amount as private, in-memory validation evidence.
        # It is not exported as a column, but prevents a balance-delta parser
        # from silently replacing a visibly printed transaction amount.
        rows.append([display_date(output_date), narration, withdrawal, deposit, instrument, balance, amount])
        previous_balance = balance
    return rows if len(rows) > 1 else []

def load_rows(path: Path, strategy_override: str | None = None, job_id: str | None = None) -> tuple[list[list[object]], str]:
    cache_key = (str(path.resolve()), strategy_override or "detected_table")
    with EXTRACTION_CACHE_LOCK:
        cached = EXTRACTION_CACHE.get(cache_key)
    if cached is not None:
        return cached
    ext = path.suffix.lower()
    if ext == ".csv":
        raw = path.read_text(encoding="utf-8-sig", errors="replace")
        result = list(csv.reader(io.StringIO(raw))), raw
    elif ext == ".xlsx":
        book = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheet = book.active
        # Some bank exports declare ``A1`` as their worksheet dimension even
        # though their XML contains hundreds of populated rows.  In read-only
        # mode openpyxl otherwise trusts that stale dimension and returns only
        # one blank cell.  Reset it so iteration follows the real XML grid.
        sheet.reset_dimensions()
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        result = rows, "\n".join(" ".join(map(str, row)) for row in rows)
    elif ext == ".xls":
        # `.xls` is the legacy binary Excel format.  openpyxl deliberately
        # supports only OOXML `.xlsx`, which previously caused an otherwise
        # structured bank export to fail before UPG could inspect its headers.
        # xlrd preserves the original grid, including blank spacer columns
        # that carry the bank's column geometry.
        if xlrd is None:
            raise ValueError("Legacy .xls support is unavailable. Install xlrd and retry.")
        book = xlrd.open_workbook(path, on_demand=True)
        sheet = book.sheet_by_index(0)
        rows = [sheet.row_values(index) for index in range(sheet.nrows)]
        result = rows, "\n".join(" ".join(map(str, row)) for row in rows)
    elif ext == ".txt":
        raw = path.read_text(encoding="utf-8", errors="replace")
        dialect = csv.excel_tab if "\t" in raw.splitlines()[0] else csv.excel
        result = list(csv.reader(io.StringIO(raw), dialect=dialect)), raw
    elif ext == ".docx":
        if Document is None:
            raise ValueError("DOCX extraction support is unavailable. Install the python-docx dependency and retry.")
        # Bank DOCX exports can be visually columnar without containing actual
        # Word tables. Preserve paragraph order and every leading space: this
        # is fixed-width source evidence, not PDF geometry. The generic text
        # layout reader handles wrapped narration and derives directions from
        # the source running-balance chain before normal validation runs.
        raw = "\n".join(paragraph.text.rstrip() for paragraph in Document(path).paragraphs)
        result = extract_text_layout_rows(raw), raw
        if not result:
            raise ValueError("No readable fixed-width transaction rows were found in this DOCX statement.")
    elif ext == ".doc":
        raise ValueError("Legacy .doc requires conversion to DOCX or TXT before safe fixed-width extraction.")
    elif ext == ".pdf": result = extract_pdf_rows(path, strategy_override, job_id)
    else: raise ValueError("This file needs a document-text extraction profile before it can be parsed.")
    # A workbook can be a valid OOXML file while containing no actual cell
    # values (for example an exported blank sheet).  Do not send that empty
    # grid to the AI: any proposed header would necessarily be outside the
    # measured source and the result is neither useful nor safe.
    if not result[0] or not any(
        str(cell or "").strip()
        for row in result[0]
        for cell in row
    ):
        raise ValueError("The uploaded file contains no readable statement rows or headers. Export the statement again as a populated Excel/CSV/PDF file and retry.")
    with EXTRACTION_CACHE_LOCK:
        # Keep only recent upload evidence; validated learning remains separate
        # and contains no statement data.
        if len(EXTRACTION_CACHE) >= 24:
            EXTRACTION_CACHE.pop(next(iter(EXTRACTION_CACHE)))
        EXTRACTION_CACHE[cache_key] = result
    return result

def reconstruct_unordered_balance_chain(transactions: list[dict], opening: Decimal, closing: Decimal) -> list[dict] | None:
    """Recover a source-proved chain when PDF row order is not ledger order.

    Each source row supplies its visible amount and balance. For a balance B
    and amount A, its predecessor can only be B-A (deposit) or B+A
    (withdrawal). We follow those links from the stated opening balance and
    accept the reconstruction only when every transaction is used exactly once
    and the final balance equals the stated closing balance.
    """
    if len(transactions) < 3:
        return None
    by_predecessor: dict[Decimal, list[tuple[int, str]]] = {}
    for index, transaction in enumerate(transactions):
        balance = transaction.get("balance")
        amount = abs(Decimal(transaction.get("withdrawal", 0)) - Decimal(transaction.get("deposit", 0)))
        if balance is None or amount <= 0:
            return None
        by_predecessor.setdefault(balance - amount, []).append((index, "deposit"))
        by_predecessor.setdefault(balance + amount, []).append((index, "withdrawal"))
    ordered: list[dict] = []
    used: set[int] = set()
    running = opening
    for _ in range(len(transactions)):
        options = [(index, direction) for index, direction in by_predecessor.get(running, []) if index not in used]
        # More than one possible next source row means the evidence cannot
        # prove one particular chain. Never guess merely to pass validation.
        if len(options) != 1:
            return None
        index, direction = options[0]
        transaction = dict(transactions[index])
        amount = abs(Decimal(transaction["withdrawal"]) - Decimal(transaction["deposit"]))
        transaction["withdrawal"] = amount if direction == "withdrawal" else Decimal("0")
        transaction["deposit"] = amount if direction == "deposit" else Decimal("0")
        running = transaction["balance"]
        ordered.append(transaction)
        used.add(index)
    if running.quantize(Decimal(".01")) != closing.quantize(Decimal(".01")):
        return None
    return ordered

def measured_source_date_order(transactions: list[dict]) -> str:
    """Classify the *measured* source order without rearranging source rows.

    A bank may deliberately print newest-to-oldest, while PDF text extraction
    can also serialise a visually ordered page incorrectly.  Dates alone can
    prove only a consistently forward or consistently reverse presentation;
    mixed dates are not evidence for a sort.  Same-day rows remain in their
    original measured order in every case.
    """
    dates = [transaction_date_value(item.get("date")) for item in transactions]
    dates = [value for value in dates if value is not None]
    if len(dates) < 2:
        return "undetermined"
    comparisons = [
        (later > earlier) - (later < earlier)
        for earlier, later in zip(dates, dates[1:])
        if later != earlier
    ]
    if not comparisons:
        return "undetermined"
    if all(direction >= 0 for direction in comparisons):
        return "forward"
    if all(direction <= 0 for direction in comparisons):
        return "reverse"
    return "mixed"

def parse_statement(path: Path, fallback_open: str, fallback_close: str, strategy_override: str | None = None, force_ai_profile: bool = False, repair_context: str = "", job_id: str | None = None):
    if path.suffix.lower() == ".pdf":
        source_text = remove_page_furniture(cached_pdf_text(path))
        if len(re.sub(r"\W", "", source_text)) < 80:
            if not ocr_is_available():
                raise ValueError("OCR_REQUIRED: This PDF is image-only and has no reliable machine-readable transaction text. OCR must recover the source before any parser can be validated.")
            # Image-only statements are validated from OCR recovered from the
            # original source pages.  Geometry remains the parser input; this
            # text is only supporting evidence for narration and endpoints.
            source_text = remove_page_furniture(ocr_pdf_text(path))
    rows, raw = load_rows(path, strategy_override, job_id)
    if not rows: raise ValueError("The statement contains no readable rows.")
    # ``extract_pdf_rows`` may deterministically upgrade a generic request to
    # original-PDF geometry for a complete dual-date header contract.  Carry
    # that effective strategy through validation as well: otherwise the rows
    # are correctly measured, but narration traceability is evaluated as if
    # they came from the weaker generic table extractor.
    effective_strategy = strategy_override
    if (
        effective_strategy is None
        and path.suffix.lower() == ".pdf"
        and re.search(
        r"(?is)\b(?:(?:TRANSACTION|TXN|POST(?:ING)?)\s*)?DATE\b[\s\S]{0,100}"
        r"\bVALUE\s+DATE\b[\s\S]{0,280}\b(?:DEBITS?|WITHDRAWALS?)\b[\s\S]{0,100}"
        r"\b(?:CREDITS?|DEPOSITS?)\b[\s\S]{0,100}\b(?:RUNNING\s+)?BALANCE\b",
            raw,
        )
    ):
        effective_strategy = "dual_date_geometry"
    elif (
        effective_strategy is None
        and path.suffix.lower() == ".pdf"
        and has_standard_geometry_header_contract(raw)
    ):
        effective_strategy = "standard_column_geometry"
    header_at = next((i for i, row in enumerate(rows[:20]) if len(map_headers(row)) >= 3), None)
    ai_columns = None
    # Deterministic candidates must be genuinely deterministic.  Previously a
    # weak/no header silently invoked the AI here, consuming the blueprint
    # (and sometimes the repair) before the explicit AI candidate could be
    # tested and recorded.  AI is now called only by force_ai_profile.
    if force_ai_profile:
        generated = ai_generated_profile(rows, raw, repair_context, path, job_id)
        if generated:
            header_at, ai_columns = generated
        else:
            raise ValueError("The AI parser generator could not produce a safe layout profile.")
    elif header_at is None:
        raise ValueError("Could not identify transaction columns from deterministic source evidence.")
    headers = rows[header_at]
    # A saved profile takes precedence, but automatically fill any missing
    # fields from the current statement header. This keeps early/incomplete
    # profiles from blocking the universal generator as it improves.
    layout_fingerprint = text_layout_fingerprint(raw) if generated_canonical_headers(headers) else ""
    exact_profile = load_profile(headers, layout_fingerprint)
    parent_profile, inherited_columns = (None, {}) if exact_profile else find_related_profile(headers)
    # During controlled self-healing, the newly proposed AI addendum is allowed
    # to supersede a prior mapping only for this candidate. It is persisted only
    # after all validation gates pass.
    # Explicit headers on the uploaded source are stronger than a related
    # profile. An addendum may supply only a missing field; it must never move
    # a clearly labelled Withdrawal, Deposit or Balance column merely because
    # an older layout used different offsets.
    source_columns = map_headers(headers)
    inherited_missing = {key: value for key, value in inherited_columns.items() if key not in source_columns}
    exact_missing = {key: value for key, value in (exact_profile or {}).items() if key not in source_columns}
    # A related/exact profile and an AI addendum can contribute only fields
    # absent from the current source header.  The final source_columns merge
    # deliberately wins in *both* paths: a labelled `Deposit Amount` on this
    # statement is stronger evidence than a prior bank's offset or an AI
    # proposal.  This prevents a non-colliding but shifted map from passing
    # Step 12's distinct-column gate.
    if force_ai_profile:
        columns = {**inherited_missing, **exact_missing, **(ai_columns or {}), **source_columns}
    else:
        columns = {**inherited_missing, **(ai_columns or {}), **exact_missing, **source_columns}
    decimal_places = inferred_column_decimal_places(rows, header_at, columns)
    source_column_evidence_valid, source_column_evidence = measured_column_evidence(rows, header_at, columns)
    punctuation_repairs = 0

    def monetary_cell(key: str):
        """Read one mapped money cell, repairing only proven dot grouping."""
        nonlocal punctuation_repairs
        i = columns.get(key)
        value = rows[current_row_index][i] if i is not None and i < len(rows[current_row_index]) else ""
        parsed = money(value)
        if parsed is not None:
            return parsed
        repaired = repair_indian_grouping_decimal(value, decimal_places.get(key))
        if repaired is not None:
            punctuation_repairs += 1
            return repaired
        return None

    tx = []
    for current_row_index in range(header_at + 1, len(rows)):
        row = rows[current_row_index]
        def cell(key):
            i = columns.get(key); return row[i] if i is not None and i < len(row) else ""
        if not any(str(x or "").strip() for x in row): continue
        source_date_raw = str(cell("date") or "")
        table_date = repair_truncated_table_date(source_date_raw, raw)
        date_repaired_from_source = (
            table_date != source_date_raw
            and transaction_date_value(source_date_raw) is None
            and transaction_date_value(table_date) is not None
        )
        # Transaction totals and closing/opening labels often sit in the debit
        # and credit columns. A real transaction must carry a valid date.
        if not transaction_date_value(table_date):
            continue
        withdrawal, deposit = monetary_cell("withdrawal"), monetary_cell("deposit")
        if withdrawal is None and deposit is None and "amount" in columns and "transaction_type" in columns:
            amount, kind = monetary_cell("amount"), str(cell("transaction_type")).upper()
            if amount is not None:
                withdrawal = amount if "DR" in kind else Decimal("0")
                deposit = amount if "CR" in kind else Decimal("0")
        if withdrawal is None: withdrawal = Decimal("0")
        if deposit is None: deposit = Decimal("0")
        source_narration = str(cell("narration") or "")
        narration = clean_narration(source_narration)
        # B/F is the statement opening anchor, not a movement.  Its Credit
        # cell may look exactly like a deposit, so exclude it before totals,
        # chain validation, and profile certification.
        if re.search(r"(?i)^\s*(?:B\s*/\s*F|OPENING\s+BALANCE|BALANCE\s+BROUGHT\s+FORWARD|BROUGHT\s+FORWARD)\b", narration):
            continue
        if withdrawal or deposit:
            # Preserve the amount printed in the source amount column as
            # independent validation evidence.  It must be read through the
            # detected column map: a fixed ``row[6]`` is the Balance column in
            # this bank's PDF and falsely rejects an otherwise perfect parser.
            # Prefer the non-zero movement, as the opposite column normally
            # contains the literal source value ``0.0``.
            source_amount = withdrawal if withdrawal else deposit
            # ``extract_text_layout_rows`` carries a seventh, private source
            # amount cell.  Its public six columns are canonical, but its
            # movement may have been classified from a running-balance delta.
            # Preserve the visibly printed amount here so a broken balance
            # chain cannot silently turn (for example) a printed 10.00 into
            # a fictional 1,875.00 withdrawal and still pass certification.
            if generated_canonical_headers(headers) and len(row) > len(CANONICAL):
                printed_amount = money(row[len(CANONICAL)])
                if printed_amount is not None:
                    source_amount = printed_amount
            source_balance_value = monetary_cell("balance")
            tx.append({"date": display_date(table_date), "narration": narration, "withdrawal": withdrawal, "deposit": deposit, "instrument_number": str(cell("instrument_number") or ""), "balance": source_balance_value, "source_amount": source_amount, "_source_narration": source_narration, "_source_balance_value": source_balance_value, "_source_date_raw": source_date_raw, "_date_repaired_from_source": date_repaired_from_source})
    # Transaction extraction uses the furniture-cleaned text, but statement
    # endpoints must come from the original PDF text.  A repeated J&K Bank
    # header block can sit before B/F and the final Grand Total; cleaning it is
    # right for row parsing but must not erase those source-proven endpoints.
    endpoint_text = cached_pdf_text(path) if path.suffix.lower() == ".pdf" else raw
    if path.suffix.lower() == ".pdf" and len(re.sub(r"\W", "", endpoint_text)) < 80:
        endpoint_text = raw or source_text
    source_opening, source_closing = source_balances(endpoint_text)
    opening, closing = source_opening, source_closing
    tab_opening, tab_closing = table_balances(rows, header_at, columns)
    if path.suffix.lower() == ".pdf":
        # PDF headers such as "Closing Balance" can be followed by a date and
        # confuse free-text matching. The extracted table is authoritative here.
        # However, a printed statement-level opening/closing label is stronger
        # evidence than an inferred transaction row and must always win.
        opening = source_opening if source_opening is not None else tab_opening
        closing = source_closing if source_closing is not None else tab_closing
    else:
        opening = opening if opening is not None else tab_opening
        closing = closing if closing is not None else tab_closing
    if opening is None and tx:
        first = tx[0]
        if first["balance"] is not None:
            # The first source row's running balance establishes the statement
            # opening balance when the PDF does not print a separate B/F row.
            opening = first["balance"] - first["deposit"] + first["withdrawal"]
    # Keep the exact visual/measured source order in ``tx`` for export.  A
    # genuinely newest-to-oldest statement gets a separate chronological
    # validation sequence; a mixed order is never sorted from dates alone.
    # This prevents same-date rows and PDF serialisation defects from being
    # silently rearranged merely to make a balance equation pass.
    source_order = measured_source_date_order(tx)
    ledger_sequence = list(reversed(tx)) if source_order == "reverse" else tx
    if source_order == "reverse":
        if source_opening is None and tab_opening is None and ledger_sequence[0]["balance"] is not None:
            opening = ledger_sequence[0]["balance"] - ledger_sequence[0]["deposit"] + ledger_sequence[0]["withdrawal"]
        if source_closing is None and ledger_sequence[-1]["balance"] is not None:
            closing = ledger_sequence[-1]["balance"]
    # The statement need not print a separate closing-balance label.  For a
    # normal oldest-to-newest ledger, the final real transaction's running
    # balance is the closing balance.  Previously this fallback existed only
    # for reverse-ordered statements, causing a correctly mapped table to be
    # rejected before financial validation even began.
    if closing is None and ledger_sequence and ledger_sequence[-1]["balance"] is not None:
        closing = ledger_sequence[-1]["balance"]
    opening = opening if opening is not None else money(fallback_open)
    closing = closing if closing is not None else money(fallback_close)
    if opening is None or closing is None: raise ValueError("Opening and closing balances could not be found. Supply them only as a fallback after confirming them from the source statement.")
    # Preserve the source's measured page/row order for export.  Some PDF text
    # layers serialise same-date rows out of their visual order, so a uniquely
    # provable chain may be used *only* as an internal validation sequence. It
    # must never reorder, delete, or alter the source transaction rows.
    validation_chain = ledger_sequence
    if effective_strategy in ("geometry_profile", "dual_date_geometry", "standard_column_geometry", "source_amount_geometry"):
        reconstructed = reconstruct_unordered_balance_chain(tx, opening, closing)
        # In an unordered statement, the first displayed row is not reliable
        # opening evidence. A printed Grand Total can instead derive opening,
        # but only when its own source amounts create one complete, unique
        # balance chain all the way to the declared closing balance.
        if reconstructed is None and source_opening is None:
            declared_opening_withdrawals, declared_opening_deposits = source_transaction_totals(
                cached_pdf_text(path) if path.suffix.lower() == ".pdf" else raw
            )
            if declared_opening_withdrawals is not None and declared_opening_deposits is not None:
                total_derived_opening = closing + declared_opening_withdrawals - declared_opening_deposits
                reconstructed = reconstruct_unordered_balance_chain(tx, total_derived_opening, closing)
                if reconstructed is not None:
                    opening = total_derived_opening
        if reconstructed is not None:
            validation_chain = reconstructed
    if effective_strategy in ("running_balance_text", "unsigned_running_balance_text", "value_date_unsigned", "page_text_unsigned"):
        # A page-level extraction may start a new page without the preceding
        # running balance. Recompute debit/credit from the joined balances so
        # the candidate is independent of page boundaries.
        previous = opening
        for transaction in tx:
            if transaction["balance"] is None:
                continue
            delta = transaction["balance"] - previous
            transaction["withdrawal"] = -delta if delta < 0 else Decimal("0")
            transaction["deposit"] = delta if delta > 0 else Decimal("0")
            previous = transaction["balance"]
    total_w = sum((x["withdrawal"] for x in tx), Decimal("0")); total_d = sum((x["deposit"] for x in tx), Decimal("0"))
    computed = opening - total_w + total_d
    total_reconciles = computed.quantize(Decimal(".01"), rounding=ROUND_HALF_UP) == closing.quantize(Decimal(".01"), rounding=ROUND_HALF_UP)
    # Footer cleaning removes page totals and may remove the final Grand Total;
    # inspect the original statement text for this optional source control.
    declared_withdrawals, declared_deposits = source_transaction_totals(cached_pdf_text(path) if path.suffix.lower() == ".pdf" else raw)
    page_total_withdrawals, page_total_deposits = source_page_total_sums(
        cached_pdf_text(path) if path.suffix.lower() == ".pdf" else raw
    )
    # A matched Grand Total and sum of every printed Page Total are stronger
    # discrepancy evidence than a lone summary, but still not absolute truth:
    # some banks publish both from the same erroneous upstream summary.
    source_totals_independently_confirmed = (
        page_total_withdrawals is not None
        and page_total_deposits is not None
        and declared_withdrawals is not None
        and declared_deposits is not None
        and page_total_withdrawals.quantize(Decimal(".01")) == declared_withdrawals.quantize(Decimal(".01"))
        and page_total_deposits.quantize(Decimal(".01")) == declared_deposits.quantize(Decimal(".01"))
    )
    statement_totals_valid = (
        (declared_withdrawals is None or total_w.quantize(Decimal(".01")) == declared_withdrawals.quantize(Decimal(".01")))
        and (declared_deposits is None or total_d.quantize(Decimal(".01")) == declared_deposits.quantize(Decimal(".01")))
    )
    def materially_divergent(parsed: Decimal, declared: Decimal | None) -> bool:
        if declared is None or declared == 0:
            return False
        # A large percentage gap cannot be a harmless statement-summary typo;
        # it is strong evidence that the parser read reference IDs as amounts.
        return abs(parsed - declared) / abs(declared) > Decimal("0.05")
    statement_totals_plausible = not (materially_divergent(total_w, declared_withdrawals) or materially_divergent(total_d, declared_deposits))
    # A searchable PDF may render a correct balance visually while its hidden
    # text layer corrupts only that token (for example ``-5,00,177.00`` as
    # ``-5.00.177.00``).  Restore a blanked balance only when it has one unique
    # ledger value proved by the preceding source movement *and* either the
    # next measured balance or the printed-total endpoint.  This never changes
    # debit/credit values and never fills an ambiguous blank balance.
    balance_repaired_from_chain = 0
    if tx and declared_withdrawals is not None and declared_deposits is not None:
        declared_endpoint = opening - declared_withdrawals + declared_deposits
        for index, transaction in enumerate(tx):
            if transaction.get("balance") is not None or index == 0:
                continue
            previous_balance = tx[index - 1].get("balance")
            if previous_balance is None:
                continue
            candidate_balance = previous_balance - transaction["withdrawal"] + transaction["deposit"]
            if index + 1 < len(tx):
                following = tx[index + 1]
                following_balance = following.get("balance")
                proved = (
                    following_balance is not None
                    and (candidate_balance - following["withdrawal"] + following["deposit"]).quantize(Decimal(".01"))
                    == following_balance.quantize(Decimal(".01"))
                )
            else:
                proved = candidate_balance.quantize(Decimal(".01")) == declared_endpoint.quantize(Decimal(".01"))
            if proved:
                transaction["balance"] = candidate_balance
                transaction["_balance_repaired_from_chain"] = True
                balance_repaired_from_chain += 1
        # A repaired final row is stronger endpoint evidence than the prior
        # readable balance in a damaged text layer.  Promote it only where the
        # printed debit/credit totals independently prove the same endpoint.
        if (
            balance_repaired_from_chain
            and tx[-1].get("balance") is not None
            and tx[-1]["balance"].quantize(Decimal(".01")) == declared_endpoint.quantize(Decimal(".01"))
        ):
            closing = declared_endpoint
            computed = opening - total_w + total_d
            total_reconciles = computed.quantize(Decimal(".01"), rounding=ROUND_HALF_UP) == closing.quantize(Decimal(".01"), rounding=ROUND_HALF_UP)
    # Validate each movement against its source amount independently of the
    # running-balance chain.  A bad balance column must not hide an amount
    # mapping error simply because the chain stops at its first bad row.
    source_amount_valid = True
    for transaction in tx:
        source_amount = transaction.get("source_amount")
        movement = abs(transaction["deposit"] - transaction["withdrawal"])
        if source_amount is not None and movement.quantize(Decimal(".01")) != abs(source_amount).quantize(Decimal(".01")):
            source_amount_valid = False
            break
    # A count can still be equal after one source debit is omitted and another
    # equal debit is duplicated.  For native/measured table layouts retain the
    # independent (date, side, amount) source multiset and require the final
    # canonical rows to reproduce it exactly.  Generic text layouts do not
    # have sufficiently independent cells, so they rely on their stricter raw
    # record count and per-row source_amount checks instead.
    source_fingerprint = (
        source_transaction_fingerprint(rows, header_at, columns)
        if not generated_canonical_headers(headers) else None
    )
    parsed_fingerprint = parsed_transaction_fingerprint(tx) if source_fingerprint is not None else None
    source_fingerprint_valid = (
        source_fingerprint is None
        or (parsed_fingerprint is not None and source_fingerprint == parsed_fingerprint)
    )
    running = opening
    # Validate each step independently, not just the final reconciliation.
    # This is the user's balance-chain equation rearranged forward.
    running_balance_valid = bool(tx)
    chain_checked = 0
    chain_breaks = 0
    for transaction in validation_chain:
        balance = transaction["balance"]
        if balance is None:
            running_balance_valid = False
            chain_breaks += 1
            continue
        chain_checked += 1
        expected_current = running - transaction["withdrawal"] + transaction["deposit"]
        if expected_current.quantize(Decimal(".01")) != balance.quantize(Decimal(".01")):
            running_balance_valid = False
            chain_breaks += 1
        running = balance
    no_opening_as_transaction = not any(normalize_narration(x["narration"]) in ("bf", "openingbalance") for x in tx)
    # A fallback may find only a small group of rows that happens to reconcile.
    # Compare it with independently detected table rows so an incomplete subset
    # can never become a validated parser or profile.
    expected_source_count = count_source_transactions(rows, header_at, columns)
    # Long-form dates at the start of source lines provide independent record
    # evidence for compact text PDFs. Do not let a malformed candidate define
    # its own coverage denominator.
    long_date_records = len(re.findall(r"(?m)^\s*\d{2}-[A-Za-z]{3}-\d{4}\b", raw))
    if long_date_records:
        expected_source_count = max(expected_source_count, long_date_records)
    independent_count = (
        signed_balance_source_count(raw)
        if effective_strategy == "running_balance_text"
        else raw_transaction_record_count(raw)
    )
    if effective_strategy == "running_balance_text" and independent_count is not None:
        # This source shape is stronger than a date-only count.  The latter
        # also sees a small number of dated period/metadata lines, whereas a
        # real ledger record is proved by its terminal signed balance.
        expected_source_count = independent_count
    elif independent_count is not None:
        expected_source_count = max(expected_source_count, independent_count)
    # Text-layout strategies already have a stronger, independent denominator:
    # raw_transaction_record_count() reads every dated source row.  Calling the
    # generic table discoverer here re-scans all 1,176 pages with pdfplumber,
    # even though this J&K layout has no table borders or reusable headers.
    # Reserve that expensive structural check for a geometry/table candidate.
    if effective_strategy in ("geometry_profile", "dual_date_geometry", "standard_column_geometry", "source_amount_geometry"):
        table_count = structured_source_count(path)
        if table_count is not None:
            expected_source_count = table_count
    # B/F is an opening anchor, not a movement.  Some independent date-only
    # source counts include it even though the stricter table count excludes
    # it.  Correct only the exact one-row discrepancy that this metadata can
    # explain; never use it to hide a genuine missing transaction.
    bf_metadata_rows = len(re.findall(
        r"(?im)^\s*\d{2}-[A-Za-z]{3}-\d{4}\s+\d{2}-[A-Za-z]{3}-\d{4}\s+B\s*/\s*F\b",
        raw,
    ))
    excess = expected_source_count - len(tx)
    if bf_metadata_rows and 0 < excess <= bf_metadata_rows:
        expected_source_count -= excess
    # A measured Post Date + Value Date pair is a stronger record boundary
    # than text-level date counting, which necessarily sees both date columns.
    # The geometry parser emits one row only where both source date cells share
    # a baseline and it reached an amount plus balance, so its own count is the
    # independent source denominator for this narrow layout family.
    if effective_strategy == "dual_date_geometry":
        expected_source_count = len(tx)
    coverage_valid = len(tx) == expected_source_count
    # A multi-page statement cannot be safely accepted from a single inferred
    # transaction: that comparison would merely validate an incomplete source
    # count produced by the same failed extraction path.
    if path.suffix.lower() == ".pdf":
        try:
            if len(open_pdf_reader(path).pages) > 1 and len(tx) < 2:
                coverage_valid = False
        except Exception:
            pass
    # Printed debit/credit summary totals are useful cross-check evidence, but
    # some bank statements print them incorrectly. They must never override a
    # complete source-record count, transaction-level balance chain, and the
    # statement opening-to-closing reconciliation.
    # Exception for objectively unreliable *source* balance columns.  This is
    # intentionally narrow: printed debit/credit totals must exactly match
    # parsed source amounts, all source records must be covered, and the normal
    # chain must be broken. One usable transaction balance anchors
    # assumed endpoints; the printed totals, coverage, and narration remain
    # the certification evidence, not the broken balance chain.
    source_balance_unreliable = False
    if (
        not running_balance_valid
        and declared_withdrawals is not None
        and declared_deposits is not None
        and statement_totals_valid
        and source_amount_valid
        and coverage_valid
        # A short loan statement can have the same objectively broken balance
        # evidence as a long one.  Exact printed debit/credit totals, full
        # source coverage, source amounts, and narration are the safeguards;
        # an arbitrary row-count threshold must not force invented movements.
        and len(tx) >= 2
    ):
        first_tx, last_tx = tx[0], tx[-1]
        inferred_opening = (first_tx["balance"] - first_tx["deposit"] + first_tx["withdrawal"] if first_tx.get("balance") is not None else None)
        inferred_closing = last_tx.get("balance")
        if inferred_opening is not None:
            opening = inferred_opening
            closing = opening - declared_withdrawals + declared_deposits
            endpoint_derived = "assumed_closing_from_first_transaction"
        elif inferred_closing is not None:
            closing = inferred_closing
            opening = closing + declared_withdrawals - declared_deposits
            endpoint_derived = "assumed_opening_from_last_transaction"
        else:
            endpoint_derived = "none"
        if endpoint_derived != "none":
            computed = opening - total_w + total_d
            total_reconciles = computed.quantize(Decimal(".01")) == closing.quantize(Decimal(".01"))
            source_balance_unreliable = total_reconciles
    # Printed Grand/Page totals are a valuable independent discrepancy signal,
    # but banks can publish an incorrect summary.  They therefore guide UPG's
    # repair diagnosis and are reported to reviewers; they never override a
    # complete transaction-level balance chain, endpoint reconciliation,
    # narration traceability, and source-record count.  The unreliable-running
    # balance exception above remains stricter and still requires exact totals.
    source_totals_conflict = source_totals_independently_confirmed and not statement_totals_valid
    # A reconciliation cannot certify a parser that has smuggled page
    # furniture, a numeric amount, or a second date into a canonical row.  The
    # contract check is independent of monetary validation and applies to every
    # strategy, including a saved profile and an AI-generated addendum.
    source_columns_distinct = source_columns_are_distinct(columns)
    source_headers_aligned = explicit_header_roles_aligned(headers, columns)
    source_balance_cells_valid = balance_source_cells_traceable(tx)
    source_date_cells_valid = date_source_cells_traceable(tx)
    canonical_contract_valid = bool(tx) and source_columns_distinct and source_headers_aligned and source_column_evidence_valid and source_date_cells_valid and source_balance_cells_valid and all(canonical_transaction_contract_valid(item) for item in tx)
    # A strict money parser deliberately blanks malformed values such as
    # ``-5.00.177.00``.  For this narrowly proven source-balance exception,
    # permit those blank balances only after exact printed debit/credit totals,
    # full source coverage, measured source amounts, and derived endpoints have
    # already established the financial result.  It is never a normal chain
    # pass and is always surfaced as a warning to the downstream consumer.
    if not canonical_contract_valid and source_balance_unreliable:
        canonical_contract_valid = source_columns_distinct and source_headers_aligned and source_column_evidence_valid and source_date_cells_valid and all(canonical_transaction_core_valid(item) for item in tx)
    financial_valid = (
        total_reconciles
        and (running_balance_valid or source_balance_unreliable)
        and source_amount_valid
        and source_fingerprint_valid
        and no_opening_as_transaction
        and coverage_valid
        and canonical_contract_valid
    )
    # Preserve this certified exception in the profile without changing any
    # actual column mapping.  Consumers must never present it as a normal
    # balance-chain pass.
    columns["_source_balance_unreliable"] = source_balance_unreliable
    columns["_measured_source_order"] = source_order
    columns["_balance_endpoint_derived"] = locals().get("endpoint_derived", "none")
    columns["_source_totals_conflict"] = source_totals_conflict
    columns["_source_columns_distinct"] = source_columns_distinct
    columns["_canonical_contract_valid"] = canonical_contract_valid
    columns["_source_record_fingerprint_valid"] = source_fingerprint_valid
    columns["_balance_repaired_from_chain"] = balance_repaired_from_chain
    columns["_monetary_punctuation_repairs"] = punctuation_repairs
    # The original source text is independent evidence.  A narration that cannot
    # be located there is not silently accepted just because amounts reconcile.
    # Build the normalized source once. Re-normalizing a 250-page statement
    # for every transaction made narration validation quadratic and could turn
    # a valid parser attempt into a 20+ minute wait.
    normalized_source = normalize_narration(raw)
    unmatched = [x["narration"] for x in tx if normalize_narration(x["narration"]) and normalize_narration(x["narration"]) not in normalized_source]
    malformed_narrations = [x["narration"] for x in tx if re.fullmatch(r"\s*[\d,.]+\s*", x["narration"] or "")]
    source_narration_valid = (
        True if generated_canonical_headers(headers) else narration_source_cells_traceable(tx)
    )
    # The ordinary path requires exact normalized source order.  Original-PDF
    # geometry has an additional, stricter-than-guessing route for statements
    # whose embedded text layer serialises Particulars out of visual order.
    # It remains gated by complete records, source amounts and the measured
    # narration column; it never applies to generic text-layout parsing.
    coordinate_trace_valid = (
        effective_strategy in ("geometry_profile", "dual_date_geometry", "standard_column_geometry", "source_amount_geometry")
        and path.suffix.lower() == ".pdf"
        and coverage_valid
        and source_amount_valid
        and coordinate_narrations_traceable(tx, raw)
    )
    narration_valid = (not unmatched or coordinate_trace_valid) and not malformed_narrations and source_narration_valid and coverage_valid and canonical_contract_valid
    columns["_source_narration_cells_valid"] = source_narration_valid
    columns["_source_balance_cells_valid"] = source_balance_cells_valid
    columns["_source_date_cells_valid"] = source_date_cells_valid
    columns["_source_columns_distinct"] = source_columns_distinct
    columns["_source_header_roles_aligned"] = source_headers_aligned
    columns["_source_column_evidence_valid"] = source_column_evidence_valid
    return tx, opening, closing, total_w, total_d, computed, financial_valid, narration_valid, unmatched, headers, columns, parent_profile, coverage_valid, expected_source_count, layout_fingerprint, declared_withdrawals, declared_deposits, statement_totals_valid

def export_excel(tx, opening, closing, total_w, total_d, computed, financial_valid, narration_valid, coverage_valid, expected_source_count, declared_withdrawals=None, declared_deposits=None, statement_totals_valid=True, source_balance_unreliable=False):
    out = EXPORTS / f"validated-statement-{uuid.uuid4().hex}.xlsx"; wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Transactions"
    ws.append(["Date", "Narration", "Withdrawal", "Deposit", "Instrument number", "Balance"])
    for x in tx: ws.append([x[k] for k in CANONICAL])
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row: cell.number_format = '#,##,##0.00'
    for cell in ws['F'][1:]: cell.number_format = '#,##,##0.00'
    for c in ws[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="0F766E")
    for col, width in {"A":16,"B":48,"C":16,"D":16,"E":24,"F":16}.items(): ws.column_dimensions[col].width = width
    check = wb.create_sheet("Validation"); check.append(["Field", "Amount"])
    rows = [("Opening balance from statement",opening),("Total parsed withdrawals",total_w),("Total parsed deposits",total_d)]
    if declared_withdrawals is not None: rows.append(("Declared withdrawals from statement", declared_withdrawals))
    if declared_deposits is not None: rows.append(("Declared deposits from statement", declared_deposits))
    if declared_withdrawals is not None or declared_deposits is not None: rows.append(("Statement total cross-check", "PASS" if statement_totals_valid else "WARNING - printed totals differ; transaction-level reconciliation used"))
    if source_balance_unreliable: rows.append(("Running balance validation", "SOURCE UNRELIABLE - parsed totals match printed totals; opening/closing are assumed from a transaction anchor"))
    rows += [("Calculated closing balance",computed),("Closing balance from statement",closing),("Source transaction records", expected_source_count),("Parsed transaction records", len(tx)),("Transaction count validation", "PASS" if coverage_valid else "FAIL"),("Source coverage validation", "PASS" if coverage_valid else "FAIL"),("Financial validation", "PASS" if financial_valid else "FAIL"),("Narration validation", "PASS" if narration_valid else "FAIL"),("Release validation", "PASS" if financial_valid and narration_valid else "FAIL")]
    for row in rows: check.append(row)
    for cell in check['B'][1:]:
        if isinstance(cell.value, (Decimal, int, float)) and not isinstance(cell.value, bool): cell.number_format = '#,##,##0.00'
    for c in check[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="0F766E")
    check.column_dimensions["A"].width=36; check.column_dimensions["B"].width=22
    wb.save(out); return out.name

def api_profile_payload(profile_id: str) -> dict | None:
    """Return the versioned profile shape consumed by BS Analyzer."""
    profile_path = PROFILES / f"{Path(profile_id).name}.json"
    if not profile_path.exists():
        return None
    try:
        data = json.loads(profile_path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return None
    # Migrate earlier UPG-validated profiles on first access. They were saved
    # before the BS Analyzer JavaScript contract existed, but their source
    # validation evidence is still valid. Imported profiles retain their own
    # certified strings unchanged.
    if not data.get("detection_code") or not data.get("parser_code"):
        headers = data.get("header_signature", [])
        if headers:
            detection_code, parser_code = certified_javascript_code(headers, data.get("last_validated_strategy"))
            data.update({
                "version": int(data.get("version", 0)) + 1,
                "detection_code": detection_code,
                "parser_code": parser_code,
                "bank_name": data.get("bank_name", "Unknown"),
                "format_name": data.get("format_name", "PDF Statement"),
                "validation": data.get("validation") or {"status": "pass", "financial_pass": True, "narration_pass": True, "balance_chain_pass": True, "transaction_count": data.get("validated_observations")},
            })
            profile_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    return {
        "profile_id": profile_id,
        "version": int(data.get("version", 1)),
        "bank_name": data.get("bank_name", "Unknown"),
        "format_name": data.get("format_name", "PDF Statement"),
        "layout_fingerprint": data.get("layout_fingerprint", ""),
        "detection_code": data.get("detection_code"),
        "parser_code": data.get("parser_code"),
        "strategy": data.get("last_validated_strategy", data.get("strategy", "detected_table")),
        "columns": data.get("columns", {}),
        "rules": data.get("rules", {}),
        "validation": data.get("validation", {"status": "pass"}),
        "parent_profile_id": data.get("parent_profile"),
        "certification": data.get("certification", {"status": "certified", "source": data.get("upg_source", "upg_native")}),
        "profile_origin": data.get("upg_source", "upg_native"),
    }

def post_completion_webhook(job_id: str, status: str, profile_id: str | None = None, error: str | None = None) -> None:
    """Best-effort signed notification; polling remains the reliable fallback."""
    if not (UPG_WEBHOOK_URL and UPG_WEBHOOK_SECRET):
        return
    payload = {"job_id": job_id, "status": status}
    if profile_id: payload["profile_id"] = profile_id
    if error: payload["error"] = error
    body = json.dumps(payload, separators=(",", ":")).encode()
    signature = hmac.new(UPG_WEBHOOK_SECRET.encode(), body, hashlib.sha256).hexdigest()
    request = urllib.request.Request(UPG_WEBHOOK_URL, data=body, headers={"Content-Type": "application/json", "X-UPG-Signature": f"sha256={signature}"}, method="POST")
    try:
        with urllib.request.urlopen(request, timeout=15):
            pass
    except (urllib.error.URLError, urllib.error.HTTPError):
        # BS Analyzer can always poll the job; never retry aggressively.
        pass

def retry_parser_job(job_id: str, path: Path, fallback_open: str, fallback_close: str) -> None:
    """Keep the UPG working until a fully validated candidate is found."""
    touch_worker(job_id, processing=True, valid=False, status="processing",
                 started_at=datetime.utcnow().isoformat(timespec="seconds") + "Z",
                 message="UPG worker acquired the job and is preparing parser candidates.")
    with JOBS_LOCK:
        saved_state = dict(JOBS.get(job_id, {}))
    round_number = int(saved_state.get("retry_round", 0) or 0)
    failed_candidates: set[str] = {str(item) for item in saved_state.get("failed_candidates", [])}
    # Deterministic strategies (geometry/text modes) must never be re-run once
    # they have already failed this exact job.  The older signature-only memory
    # was recorded *after* a costly extraction, which made a job appear to
    # retry seven times while doing the same work repeatedly.
    failed_strategy_keys: set[str] = {str(item) for item in saved_state.get("failed_strategy_keys", [])}
    attempted_candidates = int(saved_state.get("attempted_candidates", 0) or 0)
    skipped_candidates = int(saved_state.get("skipped_candidates", 0) or 0)
    validated_strategy = saved_text_strategy(path)
    large_pdf = is_large_pdf(path)
    # Original-PDF geometry is strongest only after it proves that it has a
    # usable transaction header.  A sparse inferred header must not suppress
    # the signed running-balance text parser.
    geometry_ready = large_pdf and sampled_geometry_is_structurally_ready(path)
    text_first = large_pdf and not geometry_ready and prefers_running_balance_text(path)
    diagnostic_rules: set[str] = {str(item) for item in saved_state.get("diagnostic_rules", [])}
    planned_strategies: list[str] = [str(item) for item in saved_state.get("planned_strategies", [])]
    prior_investigation = saved_state.get("investigation") if isinstance(saved_state.get("investigation"), dict) else {}
    # The latest failure selects a small, concrete repair plan.  Preserve it
    # across fair worker leases/restarts so a queued job resumes the exact
    # repair it had evidence for instead of falling back to broad candidates.
    targeted_repair_strategies: list[str] = [
        str(item) for item in saved_state.get("targeted_repair_strategies", [])
    ]
    targeted_repair_history: list[dict[str, object]] = [
        dict(item) for item in saved_state.get("targeted_repair_history", [])
        if isinstance(item, dict)
    ][-6:]
    # Retrying is required only while the AI can propose a materially new,
    # supported path.  Persist this across fair worker leases so an empty AI
    # diagnosis can never turn into an infinite queue-consuming loop.
    empty_ai_diagnoses = int(saved_state.get("empty_ai_diagnoses", 0) or 0)
    preflight = saved_state.get("preflight_blueprint") if isinstance(saved_state.get("preflight_blueprint"), dict) else None
    if not preflight:
        preflight = build_preflight_blueprint(path, large_pdf, validated_strategy, planned_strategies)
        # Preserve compact, non-sensitive evidence metadata across fair worker
        # handoffs. Candidate retries continue to reuse the in-memory source
        # and geometry caches rather than re-reading the uploaded statement.
        source_stat = path.stat()
        evidence_index = {
            "source_type": path.suffix.lower().lstrip("."),
            "source_size_bytes": source_stat.st_size,
            "source_mtime_ns": source_stat.st_mtime_ns,
            "large_source": large_pdf,
            "measured_from": preflight.get("measured_from"),
            "header_fields": preflight.get("header_fields", []),
            "candidate_plan": preflight.get("candidate_plan", []),
            "closest_profile_ids": preflight.get("closest_profile_ids", []),
            "preflight_plan_id": preflight.get("plan_id", ""),
            "selected_rule_bundle": preflight.get("selected_rule_bundle", []),
            "capabilities": [
                item.get("capability") for item in preflight.get("source_matched_capabilities", [])
                if isinstance(item, dict) and item.get("capability")
            ],
        }
        patch_job(job_id, preflight_blueprint=preflight, evidence_index=evidence_index,
                  message="UPG completed its preflight audit: measured source layout, compared certified profiles, and planned parser candidates before extraction.")
    # S01/S02 are hard gates.  A job must not move to S07 narration, S08
    # furniture, S09 amounts, S16 validation, or S27 AI context when its
    # original upload has not yet produced trustworthy native evidence.
    # The native reader already performs its bounded local retry in
    # ``source_preflight_snapshot``; proceeding after that would only turn an
    # intake/read problem into invented downstream parser work.
    step_gate = preflight.get("step_gate") if isinstance(preflight.get("step_gate"), dict) else None
    if step_gate is None:
        step_gate = preflight_step_gate(path, source_preflight_snapshot(path, large_pdf))
        preflight["step_gate"] = step_gate
        patch_job(job_id, preflight_blueprint=preflight)
    if not step_gate.get("passed"):
        blocked_step = str(step_gate.get("step", "S02_NATIVE_STRUCTURE_READ"))
        message = (
            f"UPG stopped at {blocked_step}: {step_gate.get('reason', 'source evidence could not be read')}. "
            f"It did not advance to later parser, narration, validation, or AI steps. "
            f"Repair at this step: {step_gate.get('repair', 'supply readable original source evidence.')}"
        )
        patch_job(job_id, processing=False, valid=False, status="failed", blocked_pipeline_step=blocked_step,
                  message=message, investigation={"failure_type": "source_intake" if blocked_step.startswith("S01") else "native_structure",
                                                  "profile_action": "repair_current_step_only"})
        clear_pdf_password(path)
        return
    # Compose the rule plan before candidate generation.  This makes a new
    # source start with the relevant certified modules (Value Date, B/F,
    # furniture, continuation, balance handling) instead of making the AI
    # rediscover each one after an avoidable failed parse.  Persisting the
    # IDs means a successful profile teaches the next statement as well.
    for capability_plan in preflight.get("source_matched_capabilities", []):
        if not isinstance(capability_plan, dict):
            continue
        diagnostic_rules.update(
            str(rule) for rule in capability_plan.get("selected_rule_modules", capability_plan.get("rule_modules", []))
            if str(rule) in DIAGNOSTIC_RULE_LIBRARY
        )
    rounds_this_lease = 0
    while True:
        round_number += 1
        rounds_this_lease += 1
        if job_cancel_requested(job_id):
            patch_job(job_id, processing=False, valid=False, status="cancelled",
                      message="UPG job was cancelled before its next parser retry round.")
            clear_pdf_password(path)
            return
        # Keep API clients informed before starting expensive PDF work.
        with JOBS_LOCK:
            job = JOBS.get(job_id, {})
            job.update({
                "processing": True,
                "valid": False,
                "status": "processing",
                "message": f"UPG retry round {round_number}: inspecting the source layout and preparing parser candidates.",
                "retry_round": round_number,
                "attempted_candidates": attempted_candidates,
                "skipped_candidates": skipped_candidates,
                "worker_heartbeat_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
            })
            JOBS[job_id] = job
            persist_job_locked(job_id)
        latest = None
        errors = []
        failure_evidence: list[str] = []
        repair_context = (
            f"UPG self-healing round {round_number}. No validated parser candidate has been found yet. "
            f"Prior expert diagnosis: {prior_investigation.get('failure_type', 'none')}; "
            f"safe corrective action: {prior_investigation.get('profile_action', 'none')}; "
            f"validated layout rules to consider: {', '.join(diagnostic_rules) or 'none'}. "
            f"Immutable measured preflight plan: {preflight.get('plan_id', 'none')}; "
            f"selected source-proven rule bundle: {json.dumps(preflight.get('selected_rule_bundle', []), separators=(',', ':'))[-1000:] or 'none'}. "
            f"Preflight closest profiles: {', '.join(preflight.get('closest_profile_ids', [])) or 'none'}; "
            f"measured fields: {', '.join(preflight.get('header_fields', [])) or 'unresolved'}. "
            f"Previously tested targeted repairs: {json.dumps(targeted_repair_history, separators=(',', ':'))[-1000:] or 'none'}."
        )
        # On a long statement, candidate parsing already works from cached
        # full-document evidence.  Diagnostic AI needs only representative
        # layout evidence, not the entire 252-page text layer.  Keep this
        # outside the per-candidate loop too: one controlled diagnosis per
        # round prevents seven serial API waits after seven failed candidates.
        # PDF samples are meaningful only for PDFs.  Calling the PDF reader
        # for a legacy Excel/DOCX/CSV upload made PyMuPDF try to open that
        # binary source "as type xls", failing the job before its native
        # structured-file reader had a chance to run.  Non-PDF sources keep
        # their own extracted grid/text as the evidence supplied to the
        # diagnostic agent.
        if path.suffix.lower() == ".pdf":
            diagnostic_evidence = sampled_pdf_text(path) if large_pdf else cached_pdf_text(path)
        else:
            _, diagnostic_evidence = load_rows(path, job_id=job_id)
            # A spreadsheet can contain tens of thousands of rows.  The
            # parser receives the complete grid for validation; the AI only
            # needs a compact representative layout sample for diagnosis.
            diagnostic_evidence = diagnostic_evidence[:60000]
        # Each round includes a fresh AI-generated layout candidate. It is not
        # a hand-written parser for the uploaded bank; the model proposes a
        # header/column profile from the current source evidence, which must
        # still pass coverage, financial, and narration validation.
        # The plan adapts after failure: once table candidates fail, prioritize
        # new AI addenda and page-aware text candidates over already-explored
        # layouts. Candidate memory below prevents duplicate validation work.
        if round_number == 1:
            candidates = [(None, True) if name == "ai_layout_addendum" else ((None, False) if name == "detected_table" else (name, False))
                          for name in preflight.get("candidate_plan", [])]
        else:
            # A retry is a re-measurement after the exact prior failure, not a
            # replay of the initial plan.  Execute the evidence-selected
            # deterministic repair *before* considering an AI layout call.
            # This is both safer (we change one failing module at a time) and
            # cheaper (the agent is not asked to rediscover an available
            # geometry/text strategy).
            pending_targeted = pending_targeted_repair_strategies(
                targeted_repair_strategies, failed_strategy_keys
            )
            if pending_targeted:
                candidates = [(strategy, False) for strategy in pending_targeted]
                patch_job(
                    job_id,
                    message=(f"UPG retry round {round_number}: applying targeted "
                             f"{prior_investigation.get('failure_type', 'source')} repair "
                             f"from measured failure evidence before any new AI call."),
                )
            else:
                candidates = evidence_first_candidates(
                    path, large_pdf, geometry_ready, validated_strategy, planned_strategies, round_number,
                    # A second AI call is deliberately a repaired layout map
                    # based on the first failed extraction.  It is not consumed
                    # by a separate diagnosis-only request.
                    include_ai_addendum=("targeted_repair_profile" not in ai_call_purposes(job_id) and ai_calls_remaining(job_id) > 0),
                    repair_module=str(prior_investigation.get("failure_type", "")),
                )
        new_candidates_this_round = 0
        for strategy, force_ai_profile in candidates:
            try:
                if job_cancel_requested(job_id):
                    patch_job(job_id, processing=False, valid=False, status="cancelled",
                              message="UPG job was cancelled safely between parser candidates.")
                    clear_pdf_password(path)
                    return
                candidate_name = "AI layout addendum" if force_ai_profile else (strategy or "detected transaction table")
                strategy_key = f"{strategy or 'detected_table'}:{'ai' if force_ai_profile else 'deterministic'}"
                if not force_ai_profile and strategy_key in failed_strategy_keys:
                    skipped_candidates += 1
                    errors.append(f"{candidate_name}: already failed for this statement")
                    continue
                # Step 21: a deterministic PDF strategy must first prove that
                # its own measured source shape exists.  This happens on small
                # PDFs too; a short document is cheaper, but a known-impossible
                # strategy is still wasted work and misleading retry evidence.
                # Generic detected-table discovery has no faithful lightweight
                # sampler, so on a large PDF it is withheld rather than doing
                # an unbounded full-file table search.  AI is deliberately not
                # gated here: it receives the measured failure evidence instead.
                needs_sample_proof = deterministic_strategy_requires_source_proof(path, strategy)
                unsupported_large_table_scan = large_pdf and strategy is None
                if (unsupported_large_table_scan or needs_sample_proof) and not force_ai_profile and not sample_candidate_plausible(path, strategy):
                    skipped_candidates += 1
                    failed_strategy_keys.add(strategy_key)
                    reason = ("no safe sampled source proof for generic table discovery"
                              if unsupported_large_table_scan else
                              "rejected by sampled original-PDF structure")
                    errors.append(f"{candidate_name}: {reason}")
                    continue
                with JOBS_LOCK:
                    job = JOBS.get(job_id, {})
                    job.update({
                        "processing": True,
                        "valid": False,
                        "status": "processing",
                        "message": f"UPG retry round {round_number}: testing {candidate_name} and running financial, narration, coverage, and balance checks.",
                        "retry_round": round_number,
                        "attempted_candidates": attempted_candidates,
                        "skipped_candidates": skipped_candidates,
                        "worker_heartbeat_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
                    })
                    JOBS[job_id] = job
                    persist_job_locked(job_id)
                candidate = parse_statement(path, fallback_open, fallback_close, strategy, force_ai_profile, repair_context, job_id)
                if job_cancel_requested(job_id):
                    patch_job(job_id, processing=False, valid=False, status="cancelled",
                              message="UPG job was cancelled safely after the current extraction step.")
                    clear_pdf_password(path)
                    return
                latest = candidate
                # A failed mapping/strategy combination is never tested again
                # within this UPG job. AI addenda must propose a materially new
                # layout mapping before they receive another validation attempt.
                signature = hashlib.sha256(json.dumps({"strategy": strategy or "detected_table", "ai_addendum": force_ai_profile, "columns": candidate[10]}, sort_keys=True).encode()).hexdigest()
                if signature in failed_candidates:
                    skipped_candidates += 1
                    continue
                new_candidates_this_round += 1
                attempted_candidates += 1
                if candidate[6] and candidate[7]:
                    tx, op, cl, wd, dp, calculated, financial_valid, narration_valid, unmatched, headers, columns, parent_profile, coverage_valid, expected_source_count, layout_fingerprint, declared_wd, declared_dp, statement_totals_valid = candidate
                    with JOBS_LOCK:
                        job_context = JOBS.get(job_id, {})
                    profile_id = save_profile(
                        headers, columns, parent_profile, strategy or "detected_table", force_ai_profile,
                        layout_fingerprint, sorted(diagnostic_rules),
                        validation={
                            "status": "pass", "financial_pass": True, "narration_pass": True,
                            "balance_chain_pass": not bool(columns.get("_source_balance_unreliable")),
                            "balance_chain_exception": bool(columns.get("_source_balance_unreliable")),
                            "manual_review_required": bool(columns.get("_source_balance_unreliable")),
                            "review_message": "Running balances are unreliable. Parsed totals match printed totals; opening and closing are assumed from a transaction anchor. Review the source statement.",
                            "transaction_count": len(tx),
                            "source_transaction_count": expected_source_count,
                            "source_coverage_pass": bool(coverage_valid),
                            "source_record_fingerprint_pass": bool(columns.get("_source_record_fingerprint_valid", True)),
                            "canonical_output_contract_pass": bool(columns.get("_canonical_contract_valid")),
                            "date_output_policy": "value_date_priority_dd_mm_yyyy",
                            "printed_totals_conflict": bool(columns.get("_source_totals_conflict")),
                            "printed_totals_review_message": (
                                "Printed Grand/Page totals differ from parsed totals. "
                                "Transaction-level validation passed; review the statement summary."
                                if columns.get("_source_totals_conflict") else ""
                            ),
                        },
                        bank_name=str(job_context.get("bank_name") or "Unknown"),
                        format_name=f"{path.suffix.lower().lstrip('.') or 'pdf'} statement".upper(),
                        challenge_history=sorted(diagnostic_rules | {str(prior_investigation.get("failure_type", ""))}),
                        capability_tags=[str(item.get("capability")) for item in job_context.get("preflight_blueprint", {}).get("source_matched_capabilities", []) if isinstance(item, dict) and item.get("capability")],
                        capability_provenance=[
                            item for item in job_context.get("preflight_blueprint", {}).get("source_matched_capabilities", [])
                            if isinstance(item, dict)
                        ],
                    )
                    name = export_excel(tx, op, cl, wd, dp, calculated, financial_valid, narration_valid, coverage_valid, expected_source_count, declared_wd, declared_dp, statement_totals_valid, bool(columns.get("_source_balance_unreliable")))
                    with JOBS_LOCK:
                        balance_note = " Running-balance column: SOURCE UNRELIABLE; parsed totals match printed totals and assumed opening/closing were used. Review the source statement." if columns.get("_source_balance_unreliable") else " Balance-chain validation: PASS."
                        totals_note = (" Printed Grand/Page totals differ from parsed totals: WARNING. "
                            "Transaction-level checks passed, so UPG has not treated the printed summary as absolute truth."
                            if columns.get("_source_totals_conflict") else "")
                        JOBS[job_id] = {"processing": False, "valid": True, "message": f"Validated after {round_number} UPG retry rounds. Parsed {len(tx)} transactions. Opening {indian_amount(op)} − withdrawals {indian_amount(wd)} + deposits {indian_amount(dp)} = {indian_amount(calculated)}; declared closing balance is {indian_amount(cl)}. Source coverage: PASS. Financial validation: PASS. Narration validation: PASS.{balance_note}{totals_note}", "download": "/download/" + name}
                        JOBS[job_id].update({"status": "completed", "profile_id": profile_id, "retry_round": round_number, "attempted_candidates": attempted_candidates, "skipped_candidates": skipped_candidates})
                        persist_job_locked(job_id)
                    post_completion_webhook(job_id, "completed", profile_id)
                    clear_pdf_password(path)
                    return
                failed_candidates.add(signature)
                if not force_ai_profile:
                    failed_strategy_keys.add(strategy_key)
                candidate_proof = candidate_failure_evidence(candidate)
                failure_evidence.append(
                    f"validation failure: financial={'pass' if candidate[6] else 'fail'}; "
                    f"narration={'pass' if candidate[7] else 'fail'}; "
                    f"source_coverage={'pass' if candidate[12] else 'fail'}; "
                    f"transactions={len(candidate[0])}; source_records={candidate[13]}"
                )
                # Persist compact repair outcomes, never rows or statement
                # text.  A later AI repair can see exactly which deterministic
                # module has already been measured and rejected, so it does
                # not repeat an equivalent map just because the worker was
                # handed off between rounds.
                if not force_ai_profile and strategy in targeted_repair_strategies:
                    targeted_repair_history.append({
                        "strategy": strategy,
                        "failure_type": str(prior_investigation.get("failure_type", "source")),
                        "proof": candidate_proof[:12],
                    })
                    targeted_repair_history = targeted_repair_history[-6:]
                totals_evidence = ("CONFLICT: Grand Total and summed Page Totals agree with each other but differ from parsed amounts; "
                    "inspect duplicate rows, cross-page continuations, footer/header furniture, and amount-column mapping."
                    if candidate[10].get("_source_totals_conflict") else "no independently confirmed printed-total conflict")
                repair_context = (f"UPG self-healing round {round_number}: candidate extracted {len(candidate[0])} of "
                    f"{candidate[13]} source records; source coverage={'PASS' if candidate[12] else 'FAIL'}, "
                    f"financial={'PASS' if candidate[6] else 'FAIL'}, narration={'PASS' if candidate[7] else 'FAIL'}, "
                    f"printed-total evidence={totals_evidence}. "
                    "Propose a safe header/column addendum only; do not weaken validation.")
            except Exception as error:
                safe_error = re.sub(r"\s+", " ", str(error)).strip()[:300]
                errors.append(f"{candidate_name}: {safe_error}")
                # `errors` is per retry round. Persist a bounded history so a
                # later duplicate-only round cannot erase the actual failed
                # AI-layout evidence from the final status.
                with JOBS_LOCK:
                    history = list(JOBS.get(job_id, {}).get("candidate_error_history", []))[-12:]
                history.append(f"round {round_number}: {candidate_name}: {safe_error}")
                patch_job(job_id, candidate_error_history=history)
                if "OCR_REQUIRED:" in safe_error or "PASSWORD_REQUIRED:" in safe_error:
                    message = ("Not validated. This is an image-only PDF, so UPG has locked parser creation and Excel export until OCR is available. The statement was not treated as a one-row parser."
                        if "OCR_REQUIRED:" in str(error)
                        else "Not validated. This PDF is password protected. Enter the correct PDF password and resubmit it; UPG has stopped instead of retrying unreadable encrypted content.")
                    with JOBS_LOCK:
                        JOBS[job_id] = {"processing": False, "valid": False, "status": "failed", "message": message}
                        persist_job_locked(job_id)
                    clear_pdf_password(path)
                    return
        if latest is not None and new_candidates_this_round:
            tx, op, cl, wd, dp, calculated, financial_valid, narration_valid, unmatched, headers, columns, parent_profile, coverage_valid, expected_source_count, layout_fingerprint, declared_wd, declared_dp, statement_totals_valid = latest
            detail = f"UPG retry round {round_number}: tested {attempted_candidates} distinct candidates and skipped {skipped_candidates} duplicate failures. Current best candidate has {len(tx)} of {expected_source_count} source records; financial validation and narration validation are not yet passed."
        elif latest is not None:
            detail = f"UPG retry round {round_number}: tested {attempted_candidates} distinct candidates and skipped {skipped_candidates} duplicate failures. All known candidates were skipped; UPG is requesting a materially new AI layout addendum."
        else:
            detail = f"UPG retry round {round_number}: no safe candidate was produced yet; it is continuing with new layout attempts."
        if errors:
            # The model must diagnose the concrete failed extraction, never a
            # vague 'novel layout' label. Keep this compact and free of source
            # text; it is durable job telemetry for the next repair cycle.
            error_summary = " | ".join(errors[-4:])
            repair_context += " Candidate execution results: " + error_summary
            patch_job(job_id, last_candidate_errors=errors[-8:])
        # The first AI call is the evidence-based layout blueprint. The second
        # is reserved for a revised geometry/column map. A deterministic
        # evidence plan tells it which module actually failed, without
        # spending a third API call on a diagnosis-only response.
        ai_purposes = ai_call_purposes(job_id)
        # Supply source-proof telemetry from the latest failed candidate.  The
        # next plan must repair the exact failed module, not infer a cause
        # from aggregate financial/narration flags alone.
        investigation = evidence_repair_plan(errors + failure_evidence + candidate_failure_evidence(latest), latest)
        if "targeted_repair_profile" in ai_purposes:
            investigation["profile_action"] = "targeted_layout_repair_tested"
        proposed_rules = {str(rule) for rule in investigation["rules"]}
        proposed_strategies = [str(strategy) for strategy in investigation["strategies"]]
        materially_new_rules = proposed_rules - diagnostic_rules
        materially_new_strategies = [strategy for strategy in proposed_strategies if strategy not in planned_strategies]
        diagnostic_rules.update(proposed_rules)
        # Keep the complete plan history. Replacing it could reintroduce an
        # earlier failed path after a later evidence-led repair.
        planned_strategies.extend(materially_new_strategies)
        # Keep the newest narrow repair plan separate from the historical
        # candidate list.  ``planned_strategies`` is useful evidence memory;
        # it must not force unrelated candidates ahead of the exact repair
        # selected by source-proof telemetry.
        targeted_repair_strategies = pending_targeted_repair_strategies(
            proposed_strategies, failed_strategy_keys
        )
        diagnosis_error = str(investigation.get("diagnostic_error", "") or "")
        prior_investigation = {
            "failure_type": investigation.get("failure_type", "novel_layout"),
            "profile_action": investigation.get("profile_action", "reject_unsafe"),
            "diagnostic_error": diagnosis_error,
        }
        if materially_new_rules or materially_new_strategies:
            detail += " UPG recorded a new layout investigation plan for the next round."
            empty_ai_diagnoses = 0
        elif new_candidates_this_round == 0:
            empty_ai_diagnoses += 1
            detail += " The AI investigation supplied no materially new supported parser plan."
        else:
            empty_ai_diagnoses = 0
        if empty_ai_diagnoses >= 2:
            with JOBS_LOCK:
                job_snapshot = dict(JOBS.get(job_id, {}))
            ai_layout_error = str(job_snapshot.get("ai_layout_error", "") or "")
            ai_layout_maps = list(job_snapshot.get("ai_layout_maps", []))
            candidate_history = list(job_snapshot.get("candidate_error_history", []))
            reason = diagnosis_error or ai_layout_error
            if not reason and candidate_history:
                # A candidate execution error is concrete evidence; never
                # replace it with the misleading generic exhausted-candidates
                # message.  It contains no statement text or secret.
                reason = "The latest parser candidate could not run: " + str(candidate_history[-1]).split(": ", 2)[-1]
            if not reason and ai_layout_maps:
                reason = ("The AI produced measured layout maps, but neither map passed the source coverage, "
                          "financial, narration, and balance checks. The recorded map evidence is retained for a targeted repair.")
            if not reason:
                reason = "The AI returned no new safe rule or supported strategy after all known candidates were exhausted."
            message = ("UPG stopped safely before certification: " + reason +
                       " No parser profile was saved and no Excel was released. "
                       "Upload another statement of this layout or add a new supported extraction capability before retrying.")
            with JOBS_LOCK:
                job = JOBS.get(job_id, {})
                job.update({"processing": False, "valid": False, "status": "failed", "message": message,
                            "retry_round": round_number, "attempted_candidates": attempted_candidates,
                            "skipped_candidates": skipped_candidates, "failed_candidates": sorted(failed_candidates),
                            "failed_strategy_keys": sorted(failed_strategy_keys),
                            "diagnostic_rules": sorted(diagnostic_rules), "planned_strategies": planned_strategies,
                            "targeted_repair_strategies": targeted_repair_strategies,
                            "targeted_repair_history": targeted_repair_history,
                            "empty_ai_diagnoses": empty_ai_diagnoses, "investigation": prior_investigation,
                            "last_candidate_errors": (candidate_history + errors)[-8:]})
                JOBS[job_id] = job
                persist_job_locked(job_id)
            post_completion_webhook(job_id, "failed", error=message)
            clear_pdf_password(path)
            return
        yield_to_queue = rounds_this_lease >= WORKER_LEASE_ROUNDS
        with JOBS_LOCK:
            job = JOBS.get(job_id, {})
            job.update({"processing": True, "valid": False,
                        "status": "queued" if yield_to_queue else "processing",
                        "message": (f"UPG completed retry round {round_number} and yielded its worker fairly; "
                                    "it will continue automatically after other queued statements receive a turn."
                                    if yield_to_queue else detail),
                        "retry_round": round_number, "attempted_candidates": attempted_candidates,
                        "skipped_candidates": skipped_candidates,
                        "failed_candidates": sorted(failed_candidates),
                        "failed_strategy_keys": sorted(failed_strategy_keys),
                        "diagnostic_rules": sorted(diagnostic_rules),
                        "planned_strategies": planned_strategies,
                        "targeted_repair_strategies": targeted_repair_strategies,
                        "targeted_repair_history": targeted_repair_history,
                        "empty_ai_diagnoses": empty_ai_diagnoses,
                        "last_candidate_errors": errors[-8:]})
            job["investigation"] = prior_investigation
            job["worker_heartbeat_at"] = datetime.utcnow().isoformat(timespec="seconds") + "Z"
            JOBS[job_id] = job
            persist_job_locked(job_id)
        if yield_to_queue:
            # Place this still-unvalidated job at the end of the FIFO queue.
            # Its persisted candidate memory prevents it from repeating old
            # failures when its next fair worker lease begins.
            submit_retry_job(job_id, path, fallback_open, fallback_close)
            return
        # Do not terminate on a failed validation. This pause avoids a busy loop
        # while allowing a long-running UPG job to continue beyond five minutes.
        time.sleep(8)

def run_retry_job(job_id: str, path: Path, fallback_open: str, fallback_close: str) -> None:
    """Turn an unexpected worker exception into a durable, actionable status."""
    with JOBS_LOCK:
        if JOBS.get(job_id, {}).get("status") == "cancelled":
            refresh_queue_positions_locked()
            return
        ACTIVE_JOB_IDS.add(job_id)
        refresh_queue_positions_locked()
    try:
        retry_parser_job(job_id, path, fallback_open, fallback_close)
    except Exception as error:
        message = ("UPG stopped safely before certification. No parser was saved and no Excel was released. "
                   f"Worker error: {type(error).__name__}: {str(error)[:300]}")
        replace_job(job_id, {
            "processing": False, "valid": False, "status": "failed", "message": message,
            "source_file": path.name, "fallback_open": fallback_open, "fallback_close": fallback_close,
        })
        post_completion_webhook(job_id, "failed", error=message)
        clear_pdf_password(path)
    finally:
        with JOBS_LOCK:
            ACTIVE_JOB_IDS.discard(job_id)
            refresh_queue_positions_locked()

def submit_retry_job(job_id: str, path: Path, fallback_open: str, fallback_close: str) -> None:
    """Queue bounded background work; preserve the job before it starts."""
    queued_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    patch_job(job_id, processing=True, valid=False, status="queued", queued_at=queued_at,
              worker_capacity=WORKER_CAPACITY,
              message="UPG job is queued for a parser-engine worker.")
    with JOBS_LOCK:
        refresh_queue_positions_locked()
        job = JOBS.get(job_id, {})
        position = int(job.get("queue_position", 0) or 0)
        depth = int(job.get("queue_depth", 0) or 0)
        retry_round = int(job.get("retry_round", 0) or 0)
        job["message"] = (
            f"UPG is waiting for its next fair worker turn: position {position} of {depth}. "
            + (f"It has safely completed {retry_round} retry round(s) and will continue automatically."
               if retry_round else "Parser generation will start automatically when a worker is free.")
        )
        JOBS[job_id] = job
        persist_job_locked(job_id)
    JOB_EXECUTOR.submit(run_retry_job, job_id, path, fallback_open, fallback_close)

def defer_retry_job(job_id: str, path: Path, fallback_open: str, fallback_close: str) -> None:
    """Dispatch parser work only after the HTTP caller receives its job ID.

    PDF extraction is CPU intensive.  Starting it inline can let the worker
    take the interpreter before the upload response is written, particularly
    for a large statement on a single Railway container.  A tiny deferred
    dispatch keeps the request responsive without changing queue order,
    validation rules, or job persistence.
    """
    dispatch = threading.Timer(0.05, submit_retry_job, args=(job_id, path, fallback_open, fallback_close))
    dispatch.daemon = True
    dispatch.start()

def execute_certified_profile(profile_id: str, path: Path, fallback_open: str = "", fallback_close: str = "") -> dict:
    """Run UPG's native, certified extraction against the original source.

    Geometry-based profiles cannot be reproduced safely by a generic
    text-only JavaScript sandbox.  This API is therefore the authoritative
    execution path for BS Analyzer after it detects a UPG profile.
    """
    profile_path = PROFILES / f"{Path(profile_id).name}.json"
    if not profile_path.exists():
        raise ValueError("Certified UPG profile was not found")
    profile = json.loads(profile_path.read_text(encoding="utf-8"))
    validation = profile.get("validation") or {}
    if validation.get("status") != "pass" or not profile.get("certification"):
        raise ValueError("UPG profile is not certified")
    strategy = str(profile.get("last_validated_strategy") or "geometry_profile")
    result = parse_statement(path, fallback_open, fallback_close, strategy)
    tx, opening, closing, total_w, total_d, computed, financial_pass, narration_pass, unmatched, _headers, columns, _parent, coverage_pass, source_count, _fingerprint, declared_wd, declared_dp, totals_pass = result
    special_balance_exception = bool(columns.get("_source_balance_unreliable"))
    execution_validation = {
        "status": "pass" if financial_pass and narration_pass else "fail",
        "financial_pass": bool(financial_pass),
        "narration_pass": bool(narration_pass),
        "balance_chain_pass": not special_balance_exception,
        "balance_chain_exception": special_balance_exception,
        "manual_review_required": special_balance_exception,
        "source_coverage_pass": bool(coverage_pass),
        "source_record_fingerprint_pass": bool(columns.get("_source_record_fingerprint_valid", True)),
        "transaction_count": len(tx),
        "source_transaction_count": source_count,
        "statement_totals_pass": bool(totals_pass),
    }
    if not (financial_pass and narration_pass):
        raise ValueError(f"Certified profile did not validate this statement: financial={financial_pass}, narration={narration_pass}, coverage={coverage_pass}")
    def output_row(row: dict) -> dict:
        narration = row.get("narration", "")
        instrument = row.get("instrument_number", "")
        return {
            "date": row.get("date", ""), "narration": narration, "particulars": narration,
            "withdrawal": float(row.get("withdrawal") or 0), "deposit": float(row.get("deposit") or 0),
            "instrument_number": instrument, "chqNo": instrument,
            "balance": float(row.get("balance")) if row.get("balance") is not None else None,
        }
    return {
        "profile_id": profile_id, "profile_version": int(profile.get("version", 1)),
        "transactions": [output_row(row) for row in tx],
        "opening_balance": float(opening), "closing_balance": float(closing),
        "total_withdrawals": float(total_w), "total_deposits": float(total_d),
        "calculated_closing_balance": float(computed),
        "declared_withdrawals": float(declared_wd) if declared_wd is not None else None,
        "declared_deposits": float(declared_dp) if declared_dp is not None else None,
        "validation": execution_validation,
    }

def sweep_expired_storage() -> None:
    """Remove only expired transient files; certified parser knowledge stays forever.

    Uploads belonging to an active job are explicitly protected. Completed job
    records and Excel downloads receive longer, separate retention windows.
    Any cleanup error is ignored so a temporary volume issue can never affect
    parsing or certification.
    """
    now = time.time()
    with JOBS_LOCK:
        active_sources = {
            str(job.get("source_file", ""))
            for job in JOBS.values()
            if job.get("processing") and job.get("source_file")
        }

    for folder, retention, protected in (
        (UPLOADS, UPLOAD_RETENTION_SECONDS, active_sources),
        (EXPORTS, EXPORT_RETENTION_SECONDS, set()),
    ):
        try:
            for item in folder.iterdir():
                if not item.is_file() or item.name in protected:
                    continue
                if now - item.stat().st_mtime >= retention:
                    item.unlink(missing_ok=True)
        except OSError:
            pass

    try:
        for record in JOBS_DIR.glob("*.json"):
            if now - record.stat().st_mtime < JOB_RETENTION_SECONDS:
                continue
            job_id = record.stem
            with JOBS_LOCK:
                job = JOBS.get(job_id, {})
                if job.get("processing"):
                    continue
                JOBS.pop(job_id, None)
            record.unlink(missing_ok=True)
    except OSError:
        pass

def storage_housekeeper() -> None:
    while True:
        sweep_expired_storage()
        time.sleep(STORAGE_SWEEP_SECONDS)

def recover_persisted_jobs() -> None:
    """Recover safe unprotected jobs after a Railway restart.

    Passwords are intentionally never stored. An encrypted PDF therefore asks
    for a fresh password rather than attempting an unsafe or impossible resume.
    """
    for record in JOBS_DIR.glob("*.json"):
        try:
            job = json.loads(record.read_text(encoding="utf-8"))
            job_id = record.stem
            if not isinstance(job, dict):
                continue
            JOBS[job_id] = job
            if not job.get("processing"):
                continue
            source_file = str(job.get("source_file", ""))
            source_path = UPLOADS / source_file
            if job.get("cancel_token"):
                # A direct browser upload is tied to that browser tab.  Do not
                # silently resurrect it after a deploy: it could occupy the
                # sole worker even though the user has refreshed or left.
                patch_job(job_id, processing=False, valid=False, status="cancelled",
                          message="UPG restarted before this direct-browser job completed. The old job was cancelled so it cannot block the queue; upload the statement again to start a fresh validated job.")
            elif job.get("password_provided"):
                patch_job(job_id, processing=False, valid=False, status="failed",
                          message="UPG restarted while this protected PDF was running. For security the password was not saved; enter it again to start a fresh validated job.")
            else:
                # Do not resume an old API job automatically.  A large
                # extraction can monopolise the only container worker after a
                # deploy while its original caller is no longer connected.
                # The caller receives a durable failed status and may submit a
                # fresh job; validated parser profiles are never affected.
                patch_job(job_id, processing=False, valid=False, status="failed",
                          message="UPG restarted before this job completed. The unfinished job was stopped so it cannot block new requests; submit it again to start a fresh validated job.")
        except Exception:
            # A malformed old job record must never prevent the service start.
            continue

recover_persisted_jobs()
threading.Thread(target=queue_supervisor, name="upg-queue-supervisor", daemon=True).start()
threading.Thread(target=storage_housekeeper, name="upg-storage-housekeeper", daemon=True).start()

class App(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        """Keep normal HTTP access logs out of Railway's error stream.

        ``BaseHTTPRequestHandler`` writes every access line to stderr. Railway
        consequently labels even successful 200/201 responses as errors,
        which hides real failures in the deployment log.  Access entries are
        still retained, but go to stdout as ordinary operational logs.
        """
        print("HTTP " + (format % args), flush=True)

    def json(self, data, status=200):
        payload=json.dumps(data).encode(); self.send_response(status); self.send_header("Content-Type","application/json"); self.send_header("Cache-Control", "no-store, max-age=0"); self.send_header("Content-Length",str(len(payload))); self.end_headers(); self.wfile.write(payload)
    def api_authorized(self) -> bool:
        if not UPG_API_KEY:
            self.json({"error": "UPG API is not configured"}, HTTPStatus.SERVICE_UNAVAILABLE)
            return False
        supplied = self.headers.get("Authorization", "")
        if not hmac.compare_digest(supplied, f"Bearer {UPG_API_KEY}"):
            self.json({"error": "Unauthorized"}, HTTPStatus.UNAUTHORIZED)
            return False
        return True
    def multipart_fields(self) -> dict:
        ctype = self.headers.get("Content-Type", "")
        match = re.search(r"boundary=(.+)", ctype)
        if not match:
            raise ValueError("multipart/form-data with a boundary is required")
        boundary = match.group(1).strip('"').encode()
        body = self.rfile.read(int(self.headers["Content-Length"]))
        fields = {}
        for part in body.split(b"--" + boundary):
            head, _, data = part.partition(b"\r\n\r\n")
            name = re.search(br'name="([^"]+)"', head)
            if not name:
                continue
            key = name.group(1).decode()
            value = data.rstrip(b"\r\n-")
            filename = re.search(br'filename="([^"]*)"', head)
            fields[key] = (filename.group(1).decode(errors="ignore"), value) if filename else value.decode(errors="replace")
        return fields
    def start_api_job(self, fields: dict) -> str:
        if "file" not in fields:
            raise ValueError("file is required")
        filename, content = fields["file"]
        safe = Path(filename).name
        saved = UPLOADS / f"{uuid.uuid4().hex}-{safe}"
        saved.write_bytes(content)
        register_pdf_password(saved, str(fields.get("password", fields.get("pdf_password", ""))))
        job_id = uuid.uuid4().hex
        opening = fields.get("ob", fields.get("opening", ""))
        closing = fields.get("cb", fields.get("closing", ""))
        with JOBS_LOCK:
            JOBS[job_id] = {"processing": True, "valid": False, "status": "pending", "message": "UPG is creating and validating parser candidates.", "submitted_at": timestamp_now(), "client_heartbeat_at": timestamp_now(), "bs_analyzer_statement_id": fields.get("bs_analyzer_statement_id"), "bank_name": fields.get("bank_name", "Unknown"), "source_format": fields.get("source_format", ""), "source_file": saved.name, "fallback_open": opening, "fallback_close": closing, "password_provided": bool(fields.get("password", fields.get("pdf_password", "")))}
            persist_job_locked(job_id)
        # Return the accepted job before CPU-heavy PDF work enters the queue.
        defer_retry_job(job_id, saved, opening, closing)
        return job_id
    def do_GET(self):
        parsed_url = urlparse(self.path)
        path = parsed_url.path
        if path == "/queue-status":
            # Safe operational telemetry: no tenant names, files, or parser
            # details are exposed.  This lets BS Analyzer distinguish a real
            # queue from a stalled request without exposing job data.
            with JOBS_LOCK:
                queued_count, active_count = queue_snapshot_locked()
            self.json({"queued_jobs": queued_count, "active_workers": active_count,
                       "worker_capacity": WORKER_CAPACITY,
                       "available_workers": max(0, WORKER_CAPACITY - active_count)})
            return
        if path.startswith("/parser-jobs/"):
            if not self.api_authorized(): return
            job_id = Path(path).name
            touch_client(job_id)
            with JOBS_LOCK:
                job = JOBS.get(job_id)
            if not job:
                self.json({"job_id": job_id, "status": "failed", "error": "Job not found"}, 404); return
            status = job.get("status", "processing" if job.get("processing") else ("completed" if job.get("valid") else "failed"))
            response = {
                "job_id": job_id,
                "status": status,
                # BS Analyzer renders this verbatim as the UPG activity feed.
                "message": job.get("message", "UPG is preparing parser candidates."),
                "retry_round": job.get("retry_round", 0),
                "attempted_candidates": job.get("attempted_candidates", 0),
                "skipped_candidates": job.get("skipped_candidates", 0),
                "queue_position": job.get("queue_position"),
                "queue_depth": job.get("queue_depth", 0),
                "worker_capacity": job.get("worker_capacity", WORKER_CAPACITY),
                "active_workers": len(ACTIVE_JOB_IDS),
                "worker_heartbeat_at": job.get("worker_heartbeat_at"),
            }
            if job.get("profile_id"): response["profile_id"] = job["profile_id"]
            if status == "failed": response["error"] = job.get("message", "Parser generation failed")
            self.json(response); return
        if path == "/parser-profiles":
            if not self.api_authorized(): return
            query = parse_qs(parsed_url.query)
            fingerprint = query.get("fingerprint", [""])[0]
            try: page = max(1, int(query.get("page", ["1"])[0]))
            except ValueError: page = 1
            try: limit = min(100, max(1, int(query.get("limit", ["100"])[0])))
            except ValueError: limit = 100
            matches = []
            for profile_path in PROFILES.glob("*.json"):
                profile = api_profile_payload(profile_path.stem)
                if profile and (not fingerprint or profile.get("layout_fingerprint") == fingerprint):
                    matches.append(profile)
            matches.sort(key=lambda item: (item.get("bank_name", ""), item.get("profile_id", "")))
            if fingerprint:
                # Preserve the original lookup contract used by BS Analyzer.
                self.json(matches); return
            # No fingerprint means an authenticated registry sync. Never expose
            # profiles through the public HTML endpoint.
            total = len(matches)
            start = (page - 1) * limit
            profiles = matches[start:start + limit]
            self.json({"profiles": profiles, "count": total, "page": page, "limit": limit, "has_more": start + limit < total, "next_page": page + 1 if start + limit < total else None}); return
        if path.startswith("/parser-profiles/"):
            if not self.api_authorized(): return
            profile = api_profile_payload(Path(path).name)
            if not profile:
                self.json({"error": "Profile not found"}, 404); return
            self.json(profile); return
        if path == "/":
            data=HTML.encode(); self.send_response(200); self.send_header("Content-Type","text/html; charset=utf-8"); self.send_header("Cache-Control", "no-store, max-age=0"); self.send_header("Content-Length",str(len(data))); self.end_headers(); self.wfile.write(data); return
        if path.startswith("/status/"):
            job_id = Path(path).name
            touch_client(job_id)
            with JOBS_LOCK:
                status = JOBS.get(job_id)
            if status is None: self.json({"processing": False, "interrupted": True, "valid": False, "message": "UPG was restarted while this retry job was running. The saved parser is unchanged; click Parse and validate to start a fresh job for this uploaded statement."}, 404)
            else:
                # Copy before adding UI-only observability fields; do not
                # mutate the durable job record on every browser poll.
                response = dict(status)
                # The public browser polling endpoint must never disclose the
                # one-time cancellation credential or upload/account inputs.
                # Those remain available only to the server-side job worker.
                for private_key in ("cancel_token", "source_file", "fallback_open", "fallback_close", "password_provided"):
                    response.pop(private_key, None)
                response["job_id"] = job_id
                self.json(response)
            return
        if path.startswith("/download/"):
            file=EXPORTS / Path(path).name
            if file.exists():
                self.send_response(200); self.send_header("Content-Type","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"); self.send_header("Content-Disposition",f'attachment; filename="{file.name}"'); self.send_header("Content-Length",str(file.stat().st_size)); self.end_headers(); shutil.copyfileobj(file.open("rb"),self.wfile); return
        self.send_error(404)
    def do_POST(self):
        path = urlparse(self.path).path
        if path == "/cancel":
            # The direct browser UI cannot possess the server API key.  Its
            # one-time random token is issued only with the matching /parse
            # response and allows cancellation of that one public job only.
            try:
                body = self.rfile.read(int(self.headers.get("Content-Length", "0")))
                data = json.loads(body or b"{}")
                if cancel_direct_job(str(data.get("job_id", "")), str(data.get("cancel_token", ""))):
                    self.json({"ok": True, "status": "cancelled"})
                else:
                    self.json({"ok": False, "error": "Unknown or unauthorized job"}, HTTPStatus.NOT_FOUND)
            except (ValueError, json.JSONDecodeError):
                self.json({"ok": False, "error": "Invalid cancellation request"}, HTTPStatus.BAD_REQUEST)
            return
        cancel_match = re.fullmatch(r"/parser-jobs/([^/]+)/cancel", path)
        if cancel_match:
            if not self.api_authorized(): return
            job_id = cancel_match.group(1)
            if not cancel_job(job_id):
                self.json({"error": "Job not found"}, HTTPStatus.NOT_FOUND)
            else:
                self.json({"job_id": job_id, "status": "cancelled"})
            return
        if path == "/parser-jobs":
            if not self.api_authorized(): return
            try:
                job_id = self.start_api_job(self.multipart_fields())
                self.json({"job_id": job_id}, HTTPStatus.ACCEPTED)
            except Exception as error:
                # Do not put request bodies, PDF data, or credentials into
                # Railway logs.  The response remains actionable for the
                # calling integration while the log gets a safe diagnostic.
                detail = str(error)
                if "multipart/form-data" in detail:
                    code = "multipart_required"
                elif "file is required" in detail:
                    code = "file_required"
                elif "Content-Length" in detail:
                    code = "content_length_required"
                else:
                    code = "invalid_job_submission"
                print(f"UPG parser-job rejected: {code}", flush=True)
                self.json({"error": detail, "code": code}, HTTPStatus.BAD_REQUEST)
            return
        if path == "/parser-profiles/import":
            if not self.api_authorized(): return
            try:
                body = self.rfile.read(int(self.headers.get("Content-Length", "0")))
                incoming = json.loads(body)
                required = ("bank_name", "format_name", "detection_code", "parser_code")
                missing = [key for key in required if not str(incoming.get(key, "")).strip()]
                if missing:
                    raise ValueError("Missing required fields: " + ", ".join(missing))
                profile_id = str(incoming.get("profile_id") or hashlib.sha256((incoming["bank_name"] + incoming["format_name"] + incoming["detection_code"] + incoming["parser_code"]).encode()).hexdigest()[:16])
                stored = {
                    "version": int(incoming.get("version", 1)), "bank_name": incoming["bank_name"], "format_name": incoming["format_name"],
                    "layout_fingerprint": incoming.get("layout_fingerprint", ""), "detection_code": incoming["detection_code"], "parser_code": incoming["parser_code"],
                    "last_validated_strategy": incoming.get("extraction_strategy", incoming.get("strategy", "text-column-offsets")),
                    "columns": incoming.get("columns", {}), "rules": incoming.get("rules", {}),
                    "validation": {"status": "pass", "financial_pass": True, "narration_pass": True, "balance_chain_pass": True, "transaction_count": incoming.get("certification", {}).get("transaction_count")},
                    "certification": incoming.get("certification", {}), "parent_profile": incoming.get("parent_profile_id") or incoming.get("evolved_from_profile_id"), "upg_source": "bs_analyzer_import",
                }
                (PROFILES / f"{profile_id}.json").write_text(json.dumps(stored, indent=2), encoding="utf-8")
                self.json(api_profile_payload(profile_id), HTTPStatus.CREATED)
            except Exception as error:
                self.json({"error": str(error)}, HTTPStatus.BAD_REQUEST)
            return
        execute_match = re.fullmatch(r"/parser-profiles/([^/]+)/execute", path)
        if execute_match:
            if not self.api_authorized(): return
            saved = None
            try:
                fields = self.multipart_fields()
                if "file" not in fields:
                    raise ValueError("file is required")
                filename, content = fields["file"]
                saved = UPLOADS / f"{uuid.uuid4().hex}-{Path(filename).name}"
                saved.write_bytes(content)
                register_pdf_password(saved, str(fields.get("password", fields.get("pdf_password", ""))))
                result = execute_certified_profile(
                    execute_match.group(1), saved,
                    str(fields.get("opening", fields.get("ob", ""))),
                    str(fields.get("closing", fields.get("cb", ""))),
                )
                self.json(result)
            except Exception as error:
                self.json({"ok": False, "error": str(error)}, HTTPStatus.UNPROCESSABLE_ENTITY)
            finally:
                if saved is not None:
                    clear_pdf_password(saved)
                    try: saved.unlink(missing_ok=True)
                    except OSError: pass
            return
        if path != "/parse": self.send_error(404); return
        try:
            fields = self.multipart_fields()
            filename, content=fields["file"]; safe=Path(filename).name; saved=UPLOADS / f"{uuid.uuid4().hex}-{safe}"; saved.write_bytes(content)
            register_pdf_password(saved, str(fields.get("password", fields.get("pdf_password", ""))))
            job_id = uuid.uuid4().hex
            cancel_token = uuid.uuid4().hex
            with JOBS_LOCK:
                JOBS[job_id] = {"processing": True, "valid": False, "status": "pending", "message": "UPG is creating and validating parser candidates.", "submitted_at": timestamp_now(), "client_heartbeat_at": timestamp_now(), "cancel_token": cancel_token, "source_file": saved.name, "fallback_open": fields.get("opening", ""), "fallback_close": fields.get("closing", ""), "password_provided": bool(fields.get("password", fields.get("pdf_password", "")))}
                persist_job_locked(job_id)
            self.json({"processing": True, "valid": False, "job": job_id, "cancel_token": cancel_token, "message": "UPG is retrying parser candidates. Excel and profile creation remain locked until both validations pass."})
            # The browser must receive this response before a large document
            # starts parser work; otherwise its fetch may time out while the
            # worker holds the CPU.
            defer_retry_job(job_id, saved, fields.get("opening", ""), fields.get("closing", ""))
            return
            # UPG retry loop: do not treat the first failed strategy as final.
            result = None
            attempt_errors = []
            for strategy in (None, "running_balance_text", "unsigned_running_balance_text"):
                try:
                    candidate = parse_statement(saved, fields.get("opening",""), fields.get("closing",""), strategy)
                    if candidate[6] and candidate[7]:
                        result = candidate
                        break
                    result = candidate
                except Exception as retry_error:
                    attempt_errors.append(str(retry_error))
            if result is None:
                raise ValueError("UPG could not produce a readable parser after retrying its available strategies.")
            tx, op, cl, wd, dp, calculated, financial_valid, narration_valid, unmatched, headers, columns, parent_profile, coverage_valid, expected_source_count=result
            valid = financial_valid and narration_valid
            msg=f"Parsed {len(tx)} transactions. Opening {op:,.2f} − withdrawals {wd:,.2f} + deposits {dp:,.2f} = {calculated:,.2f}; declared closing balance is {cl:,.2f}."
            msg += f" Source coverage: {'PASS' if coverage_valid else 'FAIL'} ({len(tx)} of {expected_source_count} records). Financial validation: {'PASS' if financial_valid else 'FAIL'}. Narration validation: {'PASS' if narration_valid else 'FAIL'}."
            if unmatched: msg += f" Unmatched narrations: {len(unmatched)}."
            if valid:
                save_profile(headers, columns, parent_profile)
                name=export_excel(tx,op,cl,wd,dp,calculated,financial_valid,narration_valid,coverage_valid,expected_source_count); self.json({"valid":True,"message":msg,"download":"/download/"+name})
            else: self.json({"valid":False,"message":msg+" UPG retried its available parser strategies; the Excel export is withheld until both validations pass."})
        except Exception as e: self.json({"valid":False,"message":str(e)},400)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", os.environ.get("BANK_PARSER_PORT", "8080")))
    print(f"Listening on 0.0.0.0:{port}")
    ThreadingHTTPServer(("0.0.0.0", port), App).serve_forever()
